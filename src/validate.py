#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""数据进门验证：run.py 算数之前先整体体检 6 个数据源，格式不对就报"哪个源、哪一列、Excel 第几行"。

规格全部按【真实导出】校准（原始素材/数据存档/2026真实导出 + 陆总原始202407 实测）：
- 项目明细/下单/回款记录/内部译员：智云导出，sheet 名与列名固定；日期可能是文本("2026-07-02")、
  整数(20260105) 或日期单元格；金额可能是文本("1234.56")、int、float——形态都合法，解析不出才算错。
- 收单台账：一年一 sheet(如"2026")，同工作簿还有 长沙装修/往年/Sheet1 等无关 sheet（正常，不读）；
  收单月份=文本"01"或数字；收单日期=整数20260105；表尾可能拖空表头列（正常）。
- 手填与调整：明昊自建模板（宽表：项目=行、YYYY-MM=列），项目名必须在 config.manual_items 里。

分级：error=会算错数/整块数据丢失，直接拦下不出报表；warn=可疑但不阻塞，打印出来给人看。
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

import loaders
import columns as columns_mod

MAX_ROWS_SHOWN = 5  # 每类问题最多点名前几行，其余汇总成"等N行"


@dataclass
class Issue:
    source: str  # 哪个数据源
    level: str  # error / warn
    message: str  # 哪一列、第几行、什么问题、怎么修


@dataclass
class Report:
    issues: list[Issue] = field(default_factory=list)

    def error(self, source: str, message: str) -> None:
        self.issues.append(Issue(source, "error", message))

    def warn(self, source: str, message: str) -> None:
        self.issues.append(Issue(source, "warn", message))

    @property
    def errors(self) -> list[Issue]:
        return [i for i in self.issues if i.level == "error"]

    @property
    def warns(self) -> list[Issue]:
        return [i for i in self.issues if i.level == "warn"]


def _rows_desc(rows: list[int], total: int | None = None) -> str:
    total = total if total is not None else len(rows)
    head = "、".join(f"第{r}行" for r in rows[:MAX_ROWS_SHOWN])
    return f"{head}（共{total}行）" if total > len(rows[:MAX_ROWS_SHOWN]) else head


def _open_sheet(rep: Report, source: str, path: Path, expect_sheet: str | None):
    """打开工作簿；smoke 检查 sheet 名是否符合真实导出惯例（不对只 warn，仍读激活 sheet）。"""
    if not path.exists():
        rep.error(source, f"文件不存在：{path.name}（应放在 数据/ 目录，来源见 数据/README.md）")
        return None
    try:
        wb = openpyxl.load_workbook(path, data_only=True)  # 绝不 read_only（智云 dimension 标签不可信）
    except Exception as e:
        rep.error(source, f"文件打不开（{type(e).__name__}: {e}）——可能没导出完整或不是 xlsx")
        return None
    ws = wb.active
    if expect_sheet and ws.title != expect_sheet:
        rep.warn(source, f"激活 sheet 名是「{ws.title}」，真实导出应为「{expect_sheet}」——确认没导错表")
    return ws


def _scan_sheet_rows(it, i_date, i_amt) -> tuple[list[int], list[int], int]:
    """逐行扫日期/金额；返回 (坏日期行号, 坏金额行号, 空日期数)。"""
    bad_date: list[int] = []
    bad_amt: list[int] = []
    empty_date = 0
    for rowno, row in enumerate(it, start=2):
        if all(v is None for v in row):
            continue
        dv = row[i_date] if i_date < len(row) else None
        if dv is None or not str(dv).strip():
            empty_date += 1
        elif loaders.parse_date_parts(dv) is None:
            bad_date.append(rowno)
        if loaders.amount_parse_fails(row[i_amt] if i_amt < len(row) else None):
            bad_amt.append(rowno)
    return bad_date, bad_amt, empty_date


def _scan_zhiyun_sheet(rep: Report, source: str, ws, date_col: str, amount_col: str, required: list[str]) -> None:
    """智云导出通用校验：必需列在 + 必需列不重名 + 逐行日期/金额可解析（带 Excel 行号）。"""
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(it, [])]
    missing = [c for c in required if c not in header]
    if missing:
        rep.error(source, f"缺必需列：{missing}（实际表头：{[h for h in header if h]}）——导出格式变了或导错文件")
        return
    dup = [c for c in required if header.count(c) > 1]
    if dup:
        rep.error(source, f"必需列重名：{dup}——无法确定读哪一列，先在源文件里改名去重")
        return
    i_date, i_amt = header.index(date_col), header.index(amount_col)
    bad_date, bad_amt, empty_date = _scan_sheet_rows(it, i_date, i_amt)
    if bad_date:
        rep.error(
            source, f"「{date_col}」列有日期解析不出的行：{_rows_desc(bad_date)}——这些行会被整条剔除、不计入任何周期"
        )
    if bad_amt:
        rep.error(source, f"「{amount_col}」列有非数字金额：{_rows_desc(bad_amt)}——会被按 0 计")
    if empty_date:
        rep.warn(source, f"「{date_col}」列有 {empty_date} 行为空（行会被剔除；智云导出一般不该有空日期，抽查一下）")


def _scan_ledger_rows(rows, year: int, lcols, known_cats) -> tuple[list[int], list[int], dict[str, list[int]]]:
    import periods

    c_amt, c_cat = lcols["含税金额"], lcols["对应报表大类"]
    c_d, c_m = lcols["收单日期"], lcols["收单月份"]
    bad_date: list[int] = []
    bad_amt: list[int] = []
    bad_cat: dict[str, list[int]] = {}
    for rowno, row in enumerate(rows[1:], start=2):
        if all(v is None for v in row):
            continue
        amt_raw = row[c_amt] if len(row) > c_amt else None
        if loaders.amount_parse_fails(amt_raw):
            bad_amt.append(rowno)
        has_amount = loaders.parse_amount(amt_raw) != 0.0
        if not has_amount:
            continue  # 没金额的行本来就不参与计算，下面的检查只对有金额的行有意义
        if periods.ledger_row_date(row, year, lcols) is None:
            rawd = row[c_d] if len(row) > c_d else None
            rawm = row[c_m] if len(row) > c_m else None
            if (rawd is not None and str(rawd).strip()) or (rawm is not None and str(rawm).strip()):
                bad_date.append(rowno)
        cat_raw = row[c_cat] if len(row) > c_cat else None
        cat = str(cat_raw).strip() if cat_raw not in (None, "") else ""
        if cat and cat not in known_cats:
            bad_cat.setdefault(cat, []).append(rowno)
    return bad_date, bad_amt, bad_cat


def _scan_ledger_rows(rows, year, lcols, known_cats) -> tuple[list[int], list[int], dict]:
    import periods

    c_amt, c_cat = lcols["含税金额"], lcols["对应报表大类"]
    c_d, c_m = lcols["收单日期"], lcols["收单月份"]
    bad_date: list[int] = []
    bad_amt: list[int] = []
    bad_cat: dict[str, list[int]] = {}
    for rowno, row in enumerate(rows[1:], start=2):
        if all(v is None for v in row):
            continue
        amt_raw = row[c_amt] if len(row) > c_amt else None
        if loaders.amount_parse_fails(amt_raw):
            bad_amt.append(rowno)
        has_amount = loaders.parse_amount(amt_raw) != 0.0
        if not has_amount:
            continue
        if periods.ledger_row_date(row, year, lcols) is None:
            rawd = row[c_d] if len(row) > c_d else None
            rawm = row[c_m] if len(row) > c_m else None
            if (rawd is not None and str(rawd).strip()) or (rawm is not None and str(rawm).strip()):
                bad_date.append(rowno)
        cat_raw = row[c_cat] if len(row) > c_cat else None
        cat = str(cat_raw).strip() if cat_raw not in (None, "") else ""
        if cat and cat not in known_cats:
            bad_cat.setdefault(cat, []).append(rowno)
    return bad_date, bad_amt, bad_cat


def _validate_ledger(rep: Report, cfg: dict, path: Path, year: int) -> None:
    src = "收单台账"
    if not path.exists():
        rep.error(src, f"文件不存在：{path.name}")
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    if str(year) not in wb.sheetnames:
        rep.error(src, f"找不到「{year}」sheet（现有：{wb.sheetnames}）——新一年的 sheet 还没建，找总账会计确认")
        return
    ws = wb[str(year)]
    rows = list(ws.iter_rows(values_only=True))
    try:
        lcols = columns_mod.resolve_ledger_columns(rows[0])
    except ValueError as e:
        rep.error(src, str(e))
        return
    known_cats = set(cfg["expense_categories_included"]) | set(cfg["expense_categories_excluded"])
    bad_date, bad_amt, bad_cat = _scan_ledger_rows(rows, year, lcols, known_cats)
    if bad_amt:
        rep.error(src, f"「含税金额」列有非数字：{_rows_desc(bad_amt)}——会被按 0 计")
    if bad_date:
        rep.error(
            src,
            f"「收单日期/收单月份」都解析不出的行：{_rows_desc(bad_date)}——这些费用会无声消失（收单月份要填数字如 01，别填「1月」）",
        )
    for cat, rs in sorted(bad_cat.items(), key=lambda x: -len(x[1])):
        rep.error(
            src,
            f"「对应报表大类」出现口径外的值「{cat}」：{_rows_desc(rs)}——不在 8 大类里（可能是错别字），这些费用不会计入任何科目",
        )

def _scan_manual_rows(rows, header, month_idx, known_items, i_item) -> tuple[list[str], list[str]]:
    unknown: list[str] = []
    bad_val: list[str] = []
    for rowno, r in enumerate(rows[1:], start=2):
        item = str(r[i_item]).strip() if i_item < len(r) and r[i_item] is not None else ""
        if not item:
            continue
        if item not in known_items:
            unknown.append(f"第{rowno}行「{item}」")
            continue
        for i in month_idx:
            v = r[i] if i < len(r) else None
            if loaders.amount_parse_fails(v):
                bad_val.append(f"第{rowno}行·{header[i]}列")
    return unknown, bad_val


def _manual_month_idx(header, row0) -> dict:
    month_pat = re.compile(r"\d{4}[-/]\d{1,2}")
    return {i: h for i, h in enumerate(header) if month_pat.fullmatch(h) or (hasattr(row0[i], "year"))}


def _validate_manual(rep: Report, cfg: dict, path: Path, year: int) -> None:
    src = "手填与调整"
    if not path.exists():
        rep.warn(src, f"文件不存在：{path.name}——全部手填项按 0 计，利润会虚高")
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["手填与调整"] if "手填与调整" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        rep.error(src, "表是空的")
        return
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    if "项目" not in header:
        rep.error(src, f"缺「项目」列（实际表头：{[h for h in header if h]}）——模板格式见 数据/README.md")
        return
    month_idx = _manual_month_idx(header, rows[0])
    if not month_idx:
        rep.error(src, "没有任何 YYYY-MM 月份列——手填数会整表读不进")
        return
    known_items = {it["name"] for it in cfg["manual_items"]}
    i_item = header.index("项目")
    unknown, bad_val = _scan_manual_rows(rows, header, month_idx, known_items, i_item)
    if unknown:
        rep.warn(
            src,
            f"「项目」列有 config 里不认识的名字：{'、'.join(unknown[:MAX_ROWS_SHOWN])}——这些行不会被读取（改名或在 config.manual_items 里登记）",
        )
    if bad_val:
        rep.error(src, f"有非数字的手填值：{'、'.join(bad_val[:MAX_ROWS_SHOWN])}——会被按 0 计")

def validate_all(cfg: dict, year: int, root: Path | None = None) -> Report:
    """开算前整体验证 6 个数据源。返回 Report；有 error 就不该往下算。"""
    rep = Report()
    base = loaders.data_dir(cfg, root)
    cc = cfg["columns"]

    try:
        pdp = loaders.resolve_project_detail_path(cfg, root)
    except FileNotFoundError as e:
        rep.error("项目明细(智云)", str(e))
        pdp = None
    if pdp is not None and pdp.suffix.lower() == ".csv":
        rep.warn("项目明细(智云)", "当前用的是 CSV 版（值级校验只支持 xlsx）——建议统一用智云导出的 xlsx")
    elif pdp is not None:
        ws = _open_sheet(rep, "项目明细(智云)", pdp, "项目明细")
        if ws is not None:
            _scan_zhiyun_sheet(
                rep,
                "项目明细(智云)",
                ws,
                cc["project_delivery_date"],
                cc["project_revenue"],
                [cc["project_delivery_date"], cc["project_revenue"], cc["project_cost"]],
            )

    ws = _open_sheet(rep, "下单(智云)", base / cfg["files"]["orders"], "下单")
    if ws is not None:
        _scan_zhiyun_sheet(
            rep, "下单(智云)", ws, cc["order_date"], cc["order_amount"], [cc["order_date"], cc["order_amount"]]
        )

    ws = _open_sheet(rep, "回款记录(智云)", base / cfg["files"]["receipts"], "回款记录")
    if ws is not None:
        _scan_zhiyun_sheet(
            rep,
            "回款记录(智云)",
            ws,
            cc["receipt_date"],
            cc["receipt_amount"],
            [cc["receipt_date"], cc["receipt_amount"]],
        )

    ws = _open_sheet(rep, "内部译员(智云)", base / cfg["files"]["inhouse"], "任务")
    if ws is not None:
        _scan_zhiyun_sheet(
            rep,
            "内部译员(智云)",
            ws,
            cc["inhouse_date"],
            cc["inhouse_amount"],
            [cc["inhouse_date"], cc["inhouse_amount"], cc["inhouse_type"]],
        )

    _validate_ledger(rep, cfg, base / cfg["files"]["ledger"], year)
    _validate_manual(rep, cfg, base / cfg["files"]["manual"], year)
    return rep


def print_report(rep: Report) -> None:
    if not rep.issues:
        print("=== 数据进门验证 ===\n  ✓ 6 个数据源格式全部正常")
        return
    print("=== 数据进门验证 ===")
    for i in rep.errors:
        print(f"  ✗ [{i.source}] {i.message}")
    for i in rep.warns:
        print(f"  ⚠ [{i.source}] {i.message}")
