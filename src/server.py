#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""内网双端服务（FastAPI + uvicorn）：用户端只读 + 管理员控制台（明细编辑/手填/年度预算/调整台账）。

- 用户端 `/`：账号+密码登录，按 数据/看板账号.json 权限分流（管理员→/admin、整体→整体页、BU→本 BU 页）。
- 管理员端 `/admin`：账号 lushasha（或任何权限=管理员的号）+ 密码；经手人=登录账号。
- `/api/detail`：明细数据，**仅管理员会话内可用**（服务端挡，未登录 401；非前端藏）。
- `/api/health`：最近一次运行日志（体检状态条数据源）。

安全实现用标准库：会话 HMAC 签名 token；账号明文存 数据/看板账号.json（不进 git）。
会话签名密钥存 数据/管理员密钥.json（只保留 cookie_key；旧 salt/pw_hash 字段读时忽略）。
"""
from __future__ import annotations

import base64
import hashlib
import hmac
import json
import os
import re
import threading
import time
from pathlib import Path

from fastapi import Body, FastAPI, Form, HTTPException, Query, Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles

import loaders
import accounts
import bu
import db
import core
import profit
import ingest
import render
import assets
import version as product_version
import updater
import api_v1

# v1.4 静态资源（CSS/JS/壳）：与 run.py 同级 static/
STATIC_DIR = Path(__file__).resolve().parent.parent / "static"

COOKIE = "kanban_session"
VCOOKIE = "kanban_view"   # 查看端会话：主体=登录账号名（v8.0）
SESSION_TTL = 24 * 3600
# 管理员会话看内嵌看板时隐藏「🔑密码」自改入口（管理员改密走 /admin 设置页，避免误改）
_HIDE_PW_STYLE = '<style>#pwBtn{display:none!important}</style>'
# 兼容旧测试/文档引用（v8.0 起管理员口令在 看板账号.json，不再走密钥哈希）
DEFAULT_PW = os.environ.get("KANBAN_ADMIN_PW", accounts.DEFAULT_ADMIN_PW)
DEFAULT_VIEW_PW = accounts.DEFAULT_VIEW_PW
DEFAULT_ADMIN_ACCOUNT = "lushasha"

# 服务内存态：当前汇总 + 渲染好的两端页面 + 上次规范化的原始记录（供秒级重算）
# refreshing/last_refresh：后台「立即更新」的进行中标记与最近一次结果（/api/refresh_status 用）
_state: dict = {"summary": None, "user_html": "", "admin_html": "", "built_at": None, "records": None,
                "refreshing": None, "last_refresh": None, "bu_pages": {}}
_LOCK = threading.Lock()  # 写库/重算全局互斥（03：写库一把锁，运行中排队）
_EXPORT_LOCK = threading.Lock()  # 导出截图互斥：Playwright 整页截图是重活，同一时刻只跑一张，连发返回 429


# ---------------- 密钥文件（仅会话签名 cookie_key；口令改由 accounts 管） ----------------
def _secret_path(cfg, root=None) -> Path:
    return loaders.data_dir(cfg, root) / "管理员密钥.json"


def _load_or_init_secret(cfg, root=None) -> dict:
    """读/建会话签名密钥。旧文件可能还带 salt/pw_hash/viewer_*（v7.x），读时保留不删、不再使用。"""
    p = _secret_path(cfg, root)
    if p.exists():
        try:
            sec = json.loads(p.read_text(encoding="utf-8"))
            if sec.get("cookie_key"):
                return sec
        except (OSError, ValueError):
            pass
    sec = {"cookie_key": os.urandom(32).hex()}
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(sec, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[server] 已生成会话密钥文件：{p}（账号口令见 数据/看板账号.json）")
    return sec


def _save_secret(cfg, root, sec: dict) -> None:
    p = _secret_path(cfg, root)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(sec, ensure_ascii=False, indent=2), encoding="utf-8")


# ---------------- 会话 token（HMAC 签名，含过期） ----------------
def _make_token(sec: dict, user: str, now: float | None = None) -> str:
    """签发会话。user=登录账号名（管理员 cookie 与查看 cookie 都存账号，权限运行时再查）。"""
    now = time.time() if now is None else now
    payload = f"{user}|{int(now + SESSION_TTL)}".encode()
    b64 = base64.urlsafe_b64encode(payload)
    sig = hmac.new(bytes.fromhex(sec["cookie_key"]), b64, hashlib.sha256).hexdigest()
    return b64.decode() + "." + sig


def _check_token_raw(sec: dict, token: str, now: float | None = None) -> str | None:
    """校验 HMAC+过期，返回主体字符串（账号名）；无效 None。不查权限。"""
    now = time.time() if now is None else now
    if not token or "." not in token:
        return None
    b64, sig = token.rsplit(".", 1)
    expect = hmac.new(bytes.fromhex(sec["cookie_key"]), b64.encode(), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expect, sig):
        return None
    try:
        user, exp = base64.urlsafe_b64decode(b64.encode()).decode().split("|", 1)
    except (ValueError, TypeError):
        return None
    if float(exp) < now:
        return None
    return user or None


def _check_token(sec: dict, token: str, now: float | None = None) -> str | None:
    """管理员会话 cookie → 账号名（是否仍是管理员由调用方再查账号表）。"""
    return _check_token_raw(sec, token, now)


def _check_vsubject(sec: dict, token: str, now: float | None = None) -> str | None:
    """查看端会话（cookie kanban_view）：返回登录账号名；无效 None。
    v8.0 起主体=账号（不再是 main/bu:xxx）；权限在请求时从账号表解析。"""
    return _check_token_raw(sec, token, now)

# ---------------- 渲染缓存 ----------------
def _publish(cfg, summary, html, bu_pages=None):
    _state["summary"] = summary
    _state["user_html"] = html
    _state["admin_html"] = _admin_page(html, summary, cfg)
    if bu_pages is not None:
        _state["bu_pages"] = bu_pages
    _state["built_at"] = time.strftime("%Y-%m-%d %H:%M:%S")


def _apply_profile(html: str, profile: str) -> str:
    """视图档案注入（兼容保留）。线上 view_profile 恒 full，本函数对 full/空/非法不改；
    若传入 executive 仍可换根节点属性一次（旧测试/回放兼容，非产品路径）。"""
    if not html or profile == accounts.VIEW_FULL or profile not in accounts.VIEW_PROFILES:
        return html
    return html.replace('data-profile="full"', f'data-profile="{profile}"', 1)


def _do_full(cfg, root, trigger) -> dict:
    today = loaders.pinned_today(cfg)
    summary, html, ing, bu_pages = core.generate(cfg, today, trigger=trigger)
    _state["records"] = ing.get("records")  # 缓存原始记录供秒级重算
    _publish(cfg, summary, html, bu_pages)
    return ing


def _do_recompute(cfg, root) -> None:
    if not _state.get("records"):
        _do_full(cfg, root, "manual")
        return
    today = loaders.pinned_today(cfg)
    logo = assets.load_logo_base64(cfg)
    conn = db.connect(cfg, root)
    try:
        ingest.reapply(cfg, conn, _state["records"], today)
        summary = core.summary_from_conn(cfg, conn, today)
        bu_pages = core.build_bu_pages(cfg, conn, today, logo, root)
        core.attach_unassigned(cfg, conn, today, summary, root)
    finally:
        conn.close()
    html = render.render_dashboard(summary, cfg, logo)
    _publish(cfg, summary, html, bu_pages)


def refresh(cfg, root=None, trigger="manual") -> dict:
    """完整更新（fetch+重读xlsx+重建+重放）+ 渲染两端，刷新缓存。启动与 /api/refresh 用。"""
    with _LOCK:
        return _do_full(cfg, root, trigger)


def start_refresh_async(cfg, root=None, trigger="manual") -> bool:
    """后台线程跑完整更新（在线抓开着约80秒，不能同步阻塞按钮）。
    拿到锁→起线程返回 True；拿不到（已在更新）返回 False。进度看 _state["refreshing"]/["last_refresh"]。"""
    if not _LOCK.acquire(blocking=False):
        return False
    _state["refreshing"] = {"started_at": time.strftime("%Y-%m-%d %H:%M:%S"), "trigger": trigger}

    def _job():
        t0 = time.time()
        try:
            ing = _do_full(cfg, root, trigger)
            _state["last_refresh"] = {"status": "ok", "result": ing.get("result"),
                                      "seconds": round(time.time() - t0, 1),
                                      "finished_at": time.strftime("%Y-%m-%d %H:%M:%S")}
        except Exception as e:  # 失败也要落状态，前端能看到为啥
            _state["last_refresh"] = {"status": "error", "detail": f"{type(e).__name__}: {e}",
                                      "seconds": round(time.time() - t0, 1),
                                      "finished_at": time.strftime("%Y-%m-%d %H:%M:%S")}
        finally:
            _state["refreshing"] = None
            _LOCK.release()

    threading.Thread(target=_job, daemon=True).start()
    return True


# ---------------- 设置（config.json 可改项：自动更新时间/备份保留天数/在线抓开关） ----------------
_TIME_RE = re.compile(r"^([01]\d|2[0-3]):[0-5]\d$")
SCHTASK_NAME = "经营驾驶舱每日更新"  # 与 注册每日更新.bat 里的 TN 一致

EDITABLE_SETTINGS = ("schedule_time", "backup_keep_days", "zhiyun_auto_fetch")
MAX_SCHEDULE_TIMES = 6  # 每天最多几个自动更新时间点（09:30/12:00/17:30… 3 个已够，留余量）


def normalize_schedule_times(value, fallback=None) -> list[str]:
    """把「多次更新时间」规范成有序去重的 HH:MM 列表（②多次更新时间）。
    接受：list（元素 HH:MM）/ 单个 "HH:MM" / 顿号·逗号·分号·空白分隔的串。
    每个 HH:MM 走 _TIME_RE 校验（24 小时制）→ 去重、升序；空/非法/超上限 → ValueError。"""
    if value is None:
        value = fallback
    if isinstance(value, str):
        value = re.split(r"[、，,;；\s]+", value.strip())
    if not isinstance(value, (list, tuple)):
        raise ValueError("更新时间格式不对（应为 HH:MM 列表）")
    seen, out = set(), []
    for v in value:
        s = str(v).strip()
        if not s:
            continue
        if not _TIME_RE.match(s):
            raise ValueError(f"更新时间「{s}」格式须为 HH:MM（24小时制），如 09:30")
        if s not in seen:
            seen.add(s)
            out.append(s)
    if not out:
        raise ValueError("至少保留一个自动更新时间点")
    if len(out) > MAX_SCHEDULE_TIMES:
        raise ValueError(f"自动更新时间点最多 {MAX_SCHEDULE_TIMES} 个")
    return sorted(out)


def get_schedule_times(cfg) -> list[str]:
    """读当前配置的更新时间列表：优先 schedule_times（新），缺失则从旧 schedule_time 单值推导。"""
    raw = cfg.get("schedule_times")
    if raw:
        try:
            return normalize_schedule_times(raw)
        except ValueError:
            pass
    try:
        return normalize_schedule_times(cfg.get("schedule_time") or "09:30")
    except ValueError:
        return ["09:30"]


def _schtask_command(root=None) -> str:
    """计划任务要跑的命令：<当前解释器> <run.py> --scheduled（部署机上解释器=venv python）。"""
    import sys
    py = sys.executable or "python"
    runpy = str((root or loaders.ROOT) / "run.py")
    return f'"{py}" "{runpy}" --scheduled'


def _win_task_names(n: int) -> list[str]:
    """n 个时间点对应的计划任务名：第 1 个=主名（铁律，须与 .bat 一致），其余 _2.._n。"""
    return [SCHTASK_NAME] + [f"{SCHTASK_NAME}_{i}" for i in range(2, n + 1)]


def _win_sync_schedule(times: list[str], root=None) -> str:
    """Windows：把计划任务同步成 times 里每个时间点各一个任务（主名 + _2.._n）。
    先试 /Change（改已存在任务的时间，通常不需提权）→ 不存在则 /Create（可能需提权）；
    再删掉多出来的编号任务（时间点变少时）。全程 best-effort + try/except，
    **永不抛异常打断保存**；返回人读备注。非 Windows 不会走到这里。"""
    import subprocess
    tr = _schtask_command(root)
    names = _win_task_names(len(times))
    created = changed = failed = 0
    try:
        for name, t in zip(names, times):
            r = subprocess.run(["schtasks", "/Change", "/TN", name, "/ST", t],
                               capture_output=True, timeout=15)
            if r.returncode == 0:
                changed += 1
                continue
            rc = subprocess.run(["schtasks", "/Create", "/TN", name, "/SC", "DAILY",
                                 "/ST", t, "/TR", tr, "/F"], capture_output=True, timeout=15)
            if rc.returncode == 0:
                created += 1
            else:
                failed += 1
        for i in range(len(times) + 1, MAX_SCHEDULE_TIMES + 2):  # 删多余编号任务
            subprocess.run(["schtasks", "/Delete", "/TN", f"{SCHTASK_NAME}_{i}", "/F"],
                           capture_output=True, timeout=15)
    except Exception:
        return "；⚠计划任务同步出错——请以管理员身份重跑 注册每日更新.bat"
    if failed:
        return (f"；⚠有 {failed} 个时间点没同步成（多半需管理员权限）"
                "——请以管理员身份重跑 注册每日更新.bat")
    return f"；计划任务已同步（{len(times)} 个时间点：{'、'.join(times)}）"


def _zhiyun_cfg_file(cfg, root=None) -> Path:
    return loaders.data_dir(cfg, root) / "智云配置.json"


def read_zhiyun_creds(cfg, root=None) -> dict:
    """读智云账号密码（给设置页显示；管理员会话内可见，与部署机本地明文配置同权限层级）。"""
    p = _zhiyun_cfg_file(cfg, root)
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
        return {"username": d.get("username", ""), "password": d.get("password", "")}
    except (OSError, ValueError):
        return {"username": "", "password": ""}


def save_zhiyun_creds(cfg, root, username: str, password: str) -> bool:
    """账号或密码变了才写：更新 username/password + 清掉旧会话（md_pss_id/account_id），
    下次更新自动用新账号登录、account_id 登录时自动获取——换陆总号只需在界面填这两样。
    返回是否发生了变更。"""
    p = _zhiyun_cfg_file(cfg, root)
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        d = {}
    if d.get("username") == username and d.get("password") == password:
        return False
    d["username"], d["password"] = username, password
    d["md_pss_id"] = ""      # 旧会话作废，强制新账号重登
    d["account_id"] = ""     # 登录时从页面全局变量自动取新账号的 GUID
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(d, ensure_ascii=False, indent=2), encoding="utf-8")
    return True


def read_zhiyun_conn(cfg, root=None) -> dict:
    """读智云连接配置的**生效值**（内置默认 ZHIYUN_DEFAULTS + 本地覆盖合并后）：服务器地址 + 四表ID。"""
    from ingest import fetch_zhiyun
    zy = fetch_zhiyun._load_zhiyun_cfg(cfg, root)
    tables = {s: str(((zy.get("tables") or {}).get(s) or {}).get("worksheetId", ""))
              for s in fetch_zhiyun.SOURCES}
    return {"base_url": zy.get("base_url", ""), "tables": tables}


def save_zhiyun_conn(cfg, root, base_url: str, tables: dict) -> bool:
    """保存界面填的服务器地址/四表ID到 数据/智云配置.json 覆盖层。
    与内置默认相同的项**删除覆盖**（文件保持精简、跟着代码默认走）；不同才写覆盖。
    改了服务器地址顺带清旧会话（token 绑服务器）。返回生效值是否发生变更。"""
    from ingest import fetch_zhiyun
    defaults = fetch_zhiyun.ZHIYUN_DEFAULTS
    before = read_zhiyun_conn(cfg, root)
    base_url = str(base_url or "").strip().rstrip("/")
    if not base_url:
        raise ValueError("智云服务器地址不能为空")
    p = _zhiyun_cfg_file(cfg, root)
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        d = {}
    if base_url == defaults["base_url"]:
        d.pop("base_url", None)
    else:
        d["base_url"] = base_url
    over = d.get("tables") or {}
    for s in fetch_zhiyun.SOURCES:
        wid = str((tables or {}).get(s) or "").strip()
        if not wid:
            raise ValueError("四张表的表ID都不能为空")
        if wid == defaults["tables"][s]["worksheetId"]:
            if s in over:
                over[s].pop("worksheetId", None)
                if not over[s]:
                    over.pop(s)
        else:
            over.setdefault(s, {})["worksheetId"] = wid
    if over:
        d["tables"] = over
    else:
        d.pop("tables", None)
    if base_url != before["base_url"]:
        d["md_pss_id"] = ""  # 换服务器旧 token 必失效，强制重登
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(d, ensure_ascii=False, indent=2), encoding="utf-8")
    after = read_zhiyun_conn(cfg, root)
    return after != before


def save_settings(cfg, root, payload: dict) -> dict:
    """校验并落盘设置（支持各卡就近保存：只传要改的字段即可）。
    改运行中 cfg + 重写 config.json。Windows 上改更新时间会顺手同步计划任务（多时间点=多任务）。"""
    # 更新时间：新字段 schedule_times（列表·②多次更新）优先；兼容旧 schedule_time（单值）
    changed_times = ("schedule_times" in payload) or ("schedule_time" in payload)
    if "schedule_times" in payload:
        times = normalize_schedule_times(payload["schedule_times"])
    elif "schedule_time" in payload:
        times = normalize_schedule_times(payload["schedule_time"])
    else:
        times = get_schedule_times(cfg)
    st = times[0]  # 旧字段镜像=最早的时间点（向后兼容 .bat / 读单值的地方）
    if "backup_keep_days" in payload:
        try:
            keep = int(payload.get("backup_keep_days"))
        except (TypeError, ValueError):
            raise ValueError("备份保留天数须为整数")
    else:
        keep = int(cfg.get("backup_keep_days", 30))
    if not (1 <= keep <= 365):
        raise ValueError("备份保留天数须在 1~365 之间")
    auto = bool(payload.get("zhiyun_auto_fetch", cfg.get("zhiyun_auto_fetch", False))) \
        if "zhiyun_auto_fetch" in payload else bool(cfg.get("zhiyun_auto_fetch", False))

    cfg["schedule_time"], cfg["backup_keep_days"], cfg["zhiyun_auto_fetch"] = st, keep, auto
    cfg["schedule_times"] = times
    # 落到机器本地覆盖文件（数据/本地配置.json），**绝不写 config.json** → git 工作区干净 → 一键更新可用。
    updates = {"schedule_time": st, "schedule_times": times,
               "backup_keep_days": keep, "zhiyun_auto_fetch": auto}
    # 收单台账共享盘路径（部署机专属·界面填）：传了才改，一并落覆盖文件
    if "ledger_share_path" in payload:
        lsp = str(payload.get("ledger_share_path") or "").strip()
        cfg["ledger_share_path"] = lsp
        updates["ledger_share_path"] = lsp
    loaders.write_local_config(cfg, root, updates)

    zu, zp = payload.get("zhiyun_username"), payload.get("zhiyun_password")
    cred_note = ""
    if zu is not None and zp is not None:
        zu, zp = str(zu).strip(), str(zp)
        if not zu or not zp:
            raise ValueError("智云账号和密码都不能为空")
        if save_zhiyun_creds(cfg, root, zu, zp):
            cred_note = "；智云账号已更新（下次更新自动用新账号登录）"

    # 智云连接配置（服务器地址/四表ID·内置默认可界面覆盖）：两键都传才处理（界面总是整组提交）
    conn_note = ""
    if "zhiyun_base_url" in payload or "zhiyun_tables" in payload:
        cur = read_zhiyun_conn(cfg, root)
        bu = payload.get("zhiyun_base_url", cur["base_url"])
        tb = dict(cur["tables"])
        for k, v in (payload.get("zhiyun_tables") or {}).items():
            if k in tb:
                tb[k] = v
        if save_zhiyun_conn(cfg, root, bu, tb):
            conn_note = "；智云连接配置已更新（下次更新生效）"

    note = "已保存" + cred_note + conn_note
    # 仅当本次真的提交了更新时间时才动计划任务（各卡就近保存）
    if changed_times:
        import sys
        if sys.platform == "win32":
            note += _win_sync_schedule(times, root)
        else:
            note += f"（本机非 Windows：{len(times)} 个时间点在部署机上生效）"
    return {"schedule_time": st, "schedule_times": times,
            "backup_keep_days": keep, "zhiyun_auto_fetch": auto,
            "ledger_share_path": cfg.get("ledger_share_path", ""), "note": note}


def recompute(cfg, root=None) -> None:
    """**秒级重算**（保存调整/手填后）：缓存记录重置标准表→重放→重算→重渲染，不读 xlsx。"""
    with _LOCK:
        _do_recompute(cfg, root)


# ---------------- 配置变更留痕（C3）：写接口 diff → 人读摘要 → db.log_config_change ----------------
def _audit(cfg, root, account, changes) -> None:
    """写配置变更留痕。changes=(类别,摘要) 或其列表；空摘要跳过。摘要绝不含密码明文（调用方脱敏）。"""
    if not changes:
        return
    if isinstance(changes, tuple):
        changes = [changes]
    conn = db.connect(cfg, root)
    try:
        for cat, summ in changes:
            db.log_config_change(conn, account, cat, summ)
    finally:
        conn.close()


def _join_summary(prefix: str, items: list[str], cap: int = 8) -> str:
    """把多条变更并成一句人读摘要；超过 cap 条只列前 cap 条 + 总数（防超长）。"""
    if len(items) <= cap:
        return prefix + "；".join(items)
    return prefix + "；".join(items[:cap]) + f"；等共 {len(items)} 项"


def _diff_bu_config(old_bus: list, new_bus: list,
                    old_alloc: bool = False, new_alloc: bool = False) -> list:
    """销售归属/BU 结构/分摊比例变化 → [(类别,摘要)]（old/new 均规范化 bus 列表）。"""
    def sale_map(bus):
        m = {}
        for b in bus:
            for s in b.get("销售") or []:
                m[str(s).strip()] = b["name"]
        return m

    om, nm = sale_map(old_bus), sale_map(new_bus)
    moves = [f"{s} {om.get(s) or '未归属'}→{nm.get(s) or '未归属'}"
             for s in sorted(set(om) | set(nm)) if om.get(s) != nm.get(s)]
    onames = {b["name"] for b in old_bus}
    nnames = {b["name"] for b in new_bus}
    oown = {b["name"]: "、".join(b.get("负责人") or []) for b in old_bus}
    nown = {b["name"]: "、".join(b.get("负责人") or []) for b in new_bus}
    struct = ([f"新增 BU {x}" for x in sorted(nnames - onames)]
              + [f"删除 BU {x}" for x in sorted(onames - nnames)]
              + [f"{x} 负责人改为「{nown.get(x) or '（空）'}」"
                 for x in sorted(nnames & onames) if oown.get(x) != nown.get(x)])
    out = []
    if moves:
        out.append(("销售归属", _join_summary("销售归属：", moves)))
    if struct:
        out.append(("BU配置", _join_summary("BU配置：", struct)))
    # 分摊开关 + 各 BU 比例（不存敏感值，只记百分比数字）
    alloc_lines = []
    if bool(old_alloc) != bool(new_alloc):
        alloc_lines.append(f"公共费用分摊 {'开' if new_alloc else '关'}←{'开' if old_alloc else '关'}")
    orat = {b["name"]: b.get("分摊比例") for b in old_bus}
    nrat = {b["name"]: b.get("分摊比例") for b in new_bus}
    for nm in sorted(set(orat) | set(nrat)):
        o, n = orat.get(nm), nrat.get(nm)
        if o != n:
            def _fmt(v):
                return "空" if v is None else f"{v:g}%"
            alloc_lines.append(f"{nm} {_fmt(o)}→{_fmt(n)}")
    if alloc_lines:
        out.append(("分摊", _join_summary("分摊：", alloc_lines)))
    return out


def _diff_accounts(old_accs: list, new_accs: list) -> list:
    """账号增删改（含改权限/改密码/改显示名）→ [(账号,摘要)]。密码只记「改密码」不记值。"""
    om = {a["账号"]: a for a in old_accs}
    nm = {a["账号"]: a for a in new_accs}
    lines = [f"新增 {a}（{nm[a].get('权限')}）" for a in sorted(set(nm) - set(om))]
    lines += [f"删除 {a}" for a in sorted(set(om) - set(nm))]
    for a in sorted(set(om) & set(nm)):
        o, n = om[a], nm[a]
        chg = []
        if o.get("权限") != n.get("权限"):
            chg.append(f"权限 {o.get('权限')}→{n.get('权限')}")
        if str(o.get("密码")) != str(n.get("密码")):
            chg.append("改密码")
        if (o.get("显示名") or "") != (n.get("显示名") or ""):
            chg.append("改显示名")
        if chg:
            lines.append(f"{a} " + "、".join(chg))
    return [("账号", _join_summary("账号：", lines))] if lines else []


def _manual_items_json(cfg: dict | None = None) -> str:
    """手填项目名列表 JSON（注入管理端 JS；与 config.manual_items 同步）。"""
    import json as _json
    items = [it["name"] for it in (cfg or {}).get("manual_items") or []] or [
        "营销人力成本", "管理人力成本", "研发人力成本", "财务费用补充", "PM人力成本", "VM人力成本",
        "实际内部译员成本", "税费损失", "技术流量成本", "其他（生产成本）", "其他损益"]
    return _json.dumps(items, ensure_ascii=False, separators=(",", ":"))


def _admin_page(dash_html: str, summary: dict, cfg: dict | None = None) -> str:
    """管道跑通后标记「管理端可进完整台」（truthy 写入 _state['admin_html']）。
    页面本体只在 static/admin/，此处不再生成整页 HTML。"""
    return "ready"


def _admin_static_html() -> str:
    """管理端完整台骨架：static/admin/admin.html（CSS/JS 外链；手填清单由 /admin/app.js 注入）。"""
    p = STATIC_DIR / "admin" / "admin.html"
    if not p.is_file():
        raise FileNotFoundError(f"缺少管理端静态页：{p}")
    return p.read_text(encoding="utf-8")


def _bootstrap_page() -> str:
    """首次部署引导页（F-02）：仅 static/admin/bootstrap.html。"""
    p = STATIC_DIR / "admin" / "bootstrap.html"
    if not p.is_file():
        raise FileNotFoundError(f"缺少引导页：{p}")
    return p.read_text(encoding="utf-8")


def admin_ui_source() -> str:
    """供测试搜锚点：admin.html + admin.js + admin.css 拼接（非运行路径）。"""
    parts = []
    for name in ("admin.html", "admin.js", "admin.css", "bootstrap.html"):
        p = STATIC_DIR / "admin" / name
        if p.is_file():
            parts.append(p.read_text(encoding="utf-8"))
    return "\n".join(parts)


# 测试兼容别名：历史上断言搜 server._ADMIN_CONSOLE，现指向 static 拼接
def __getattr__(name: str):
    if name == "_ADMIN_CONSOLE":
        return admin_ui_source()
    raise AttributeError(name)




def _run_reasons(report: dict) -> list[str]:
    """从最近一次管道运行日志（体检JSON=report）推导"为啥黄/红"。
    与 ingest._log_run 判定口径一致：fetch 走本地副本/无源、过期调整。
    注意：这是「管道运行」信号（黄/红），与「数据体检」的未填分类等（警）是两套，别糊在一起。"""
    report = report or {}
    reasons: list[str] = []
    fetch = report.get("fetch", {}) or {}
    st = fetch.get("status")
    if st == "no_source":
        reasons.append("收单台账无可用数据源（共享路径与本地副本都没有）→ 判红")
    elif st and st != "fetched":
        reasons.append(f"收单台账未从共享路径拉取、走本地副本（状态：{st}）")
    adj = report.get("adjust", {}) or {}
    if adj.get("expired", 0):
        reasons.append(f"{adj['expired']} 条调整「过期疑似」（源头已改、调整未套用）→ 去『异常处理·数据修正』看")
    if adj.get("missing", 0):
        reasons.append(f"{adj['missing']} 条调整定位键失配未套用（源头行删了/改了金额，剔除或改值没生效）→ 去『异常处理·数据修正』人工复核")
    return reasons


_LOGIN_HTML = """<!doctype html><html lang="zh"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>管理员登录 · 甲骨易智能经营罗盘</title>
<script>try{{if(localStorage.getItem("cockpit-theme")==="light")document.documentElement.classList.add("theme-light")}}catch(e){{}}</script>
<style>
:root{{--bg:#0f172a;--card:#1e293b;--fg:#e2e8f0;--mut:#94a3b8;--line:#334155;--input-bg:#0f172a;--hint:#64748b;--err:#f87171}}
html.theme-light{{--bg:#eef1f5;--card:#fff;--fg:#1d2836;--mut:#525c68;--line:#e3e8ef;--input-bg:#fff;--hint:#64748b;--err:#dc2626}}
body{{font-family:-apple-system,system-ui,sans-serif;background:var(--bg);color:var(--fg);
display:flex;min-height:100vh;align-items:center;justify-content:center;margin:0;position:relative}}
.card{{background:var(--card);padding:32px;border-radius:12px;width:300px;box-shadow:0 8px 30px rgba(0,0,0,.2);border:1px solid var(--line)}}
h1{{font-size:18px;margin:0 0 20px}}label{{font-size:13px;color:var(--mut)}}
input{{width:100%;box-sizing:border-box;margin:6px 0 16px;padding:9px;border-radius:7px;
border:1px solid var(--line);background:var(--input-bg);color:var(--fg);font-size:14px}}
button[type=submit]{{width:100%;padding:10px;border:0;border-radius:7px;background:#8b5cf6;color:#fff;
font-size:15px;cursor:pointer}}.err{{color:var(--err);font-size:13px;margin-bottom:10px}}
.hint{{color:var(--hint);font-size:12px;margin-top:12px}}
#themeBtn{{position:fixed;top:14px;right:16px;background:transparent;border:1px solid var(--line);color:var(--fg);
padding:6px 12px;border-radius:8px;font-size:13px;cursor:pointer}}
#themeBtn:hover{{border-color:#8b5cf6}}
</style></head>
<body>
<button type="button" id="themeBtn" title="深色/浅色（全局同步）">◑ 浅色</button>
<form class="card" method="post" action="/admin/login">
<h1>管理员端登录</h1>{err}
<label>账号</label><input name="account" value="{account}" autocomplete="username" autofocus>
<label>密码</label><input type="password" name="password" autocomplete="current-password">
<button type="submit">进入</button>
<div class="hint">管理员账号见「看板账号」表（默认 lushasha）。</div></form>
<script>
(function(){{var r=document.documentElement,b=document.getElementById("themeBtn");
function setL(l){{r.classList.toggle("theme-light",!!l);if(b)b.textContent=l?"◐ 深色":"◑ 浅色";}}
try{{setL(localStorage.getItem("cockpit-theme")==="light");}}catch(e){{}}
if(b)b.onclick=function(){{var l=!r.classList.contains("theme-light");try{{localStorage.setItem("cockpit-theme",l?"light":"dark");}}catch(e){{}}setL(l);}};
window.addEventListener("storage",function(e){{if(e.key==="cockpit-theme")setL(e.newValue==="light");}});}})();
</script></body></html>"""



def _login_page(err: str = "", account: str = "") -> str:
    err_html = f'<div class="err">{err}</div>' if err else ""
    acct = str(account or DEFAULT_ADMIN_ACCOUNT).replace("&", "&amp;").replace("<", "&lt;").replace('"', "&quot;")
    return _LOGIN_HTML.format(err=err_html, account=acct)


# 查看端登录页（v8.0）：账号+密码，按权限分流（管理员→/admin、整体→整体页、BU→本 BU 页）。
_VIEW_LOGIN_HTML = """<!doctype html><html lang="zh"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>看板登录 · 甲骨易智能经营罗盘</title>
<script>try{{if(localStorage.getItem("cockpit-theme")==="light")document.documentElement.classList.add("theme-light")}}catch(e){{}}</script>
<style>
:root{{--bg:#0f172a;--card:#1e293b;--fg:#e2e8f0;--mut:#94a3b8;--line:#334155;--input-bg:#0f172a;--hint:#64748b;--err:#f87171}}
html.theme-light{{--bg:#eef1f5;--card:#fff;--fg:#1d2836;--mut:#525c68;--line:#e3e8ef;--input-bg:#fff;--hint:#64748b;--err:#dc2626}}
body{{font-family:-apple-system,system-ui,sans-serif;background:var(--bg);color:var(--fg);
display:flex;min-height:100vh;align-items:center;justify-content:center;margin:0;position:relative}}
.card{{background:var(--card);padding:32px;border-radius:12px;width:300px;box-shadow:0 8px 30px rgba(0,0,0,.2);border:1px solid var(--line)}}
h1{{font-size:18px;margin:0 0 20px}}label{{font-size:13px;color:var(--mut)}}
input{{width:100%;box-sizing:border-box;margin:6px 0 16px;padding:9px;border-radius:7px;
border:1px solid var(--line);background:var(--input-bg);color:var(--fg);font-size:14px}}
button[type=submit]{{width:100%;padding:10px;border:0;border-radius:7px;background:#8b5cf6;color:#fff;
font-size:15px;cursor:pointer}}.err{{color:var(--err);font-size:13px;margin-bottom:10px}}
.hint{{color:var(--hint);font-size:12px;margin-top:12px}}
#themeBtn{{position:fixed;top:14px;right:16px;background:transparent;border:1px solid var(--line);color:var(--fg);
padding:6px 12px;border-radius:8px;font-size:13px;cursor:pointer}}
#themeBtn:hover{{border-color:#8b5cf6}}
</style></head>
<body>
<button type="button" id="themeBtn" title="深色/浅色（全局同步）">◑ 浅色</button>
<form class="card" method="post" action="/login">
<h1>看板登录</h1>{err}
<label>账号</label><input name="account" value="{account}" autocomplete="username" autofocus>
<label>密码</label><input type="password" name="password" autocomplete="current-password">
<button type="submit">进入</button>
<div class="hint">账号密码问财务部管理员要；登录后可自己改密码。忘记密码找管理员重置。</div></form>
<script>
(function(){{var r=document.documentElement,b=document.getElementById("themeBtn");
function setL(l){{r.classList.toggle("theme-light",!!l);if(b)b.textContent=l?"◐ 深色":"◑ 浅色";}}
try{{setL(localStorage.getItem("cockpit-theme")==="light");}}catch(e){{}}
if(b)b.onclick=function(){{var l=!r.classList.contains("theme-light");try{{localStorage.setItem("cockpit-theme",l?"light":"dark");}}catch(e){{}}setL(l);}};
window.addEventListener("storage",function(e){{if(e.key==="cockpit-theme")setL(e.newValue==="light");}});}})();
</script></body></html>"""


def _view_login_page(err: str = "", account: str = "") -> str:
    err_html = f'<div class="err">{err}</div>' if err else ""
    acct = str(account).replace("&", "&amp;").replace("<", "&lt;").replace('"', "&quot;")
    return _VIEW_LOGIN_HTML.format(err=err_html, account=acct)




# ---------------- FastAPI 应用 ----------------
def create_app(cfg, root=None) -> FastAPI:
    app = FastAPI(title="甲骨易智能经营罗盘", docs_url=None, redoc_url=None, openapi_url=None)
    sec = _load_or_init_secret(cfg, root)
    # 确保账号文件存在（部署零配置）
    accounts.load_accounts(cfg, root, create=True)
    # 静态 CSS/JS（展示层外置；内容来自 theme/render 原样抽取）
    if STATIC_DIR.is_dir():
        app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

    def _user(request: Request) -> str | None:
        """管理员会话：cookie 主体=账号名，且账号表里权限仍是「管理员」。经手人=该账号。"""
        name = _check_token(sec, request.cookies.get(COOKIE, ""))
        if not name:
            return None
        acc = accounts.find_account(cfg, root, name)
        return name if accounts.is_admin(acc) else None

    def _vacct(request: Request) -> str | None:
        """查看端会话：返回登录账号名（权限运行时再解析）。"""
        return _check_vsubject(sec, request.cookies.get(VCOOKIE, ""))

    def _vacc_row(request: Request) -> dict | None:
        name = _vacct(request)
        return accounts.find_account(cfg, root, name) if name else None

    def _can_view_main(request: Request) -> bool:
        """整体页/全公司口径：整体权限账号 或 管理员会话。BU 账号不行。"""
        if _user(request):
            return True
        acc = _vacc_row(request)
        return accounts.is_main(acc)

    def _can_view_bu(request: Request, bu_name: str) -> bool:
        if _user(request):
            return True
        acc = _vacc_row(request)
        if not acc:
            return False
        if accounts.is_main(acc):
            return True
        return accounts.can_see_bu(acc, bu_name)  # 多 BU：在其绑定名单内即可

    def _bu_switcher_html(my_names, current: str) -> str:
        """多 BU 账号看 BU 页时顶部的「我的 BU」切换条：**只列该账号绑定且仍存在的 BU**
        （绝不列他 BU，铁律12）。单个绑定不出条。"""
        from urllib.parse import quote

        def esc(s):
            return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
        existing = [n for n in my_names if n in _state.get("bu_pages", {})]
        if len(existing) <= 1:
            return ""
        links = "".join(
            f'<a class="bu-nav-a" href="/bu/{quote(n)}"'
            + (' aria-current="page" style="border-color:var(--blue)"' if n == current else "")
            + f'>{esc(n)}</a>' for n in existing)
        return ('<div class="bu-nav" role="navigation" aria-label="我的 BU 分页">'
                '<span class="bu-nav-label">我的 BU</span>'
                '<span class="bu-nav-links">' + links + '</span></div>')

    def _bu_view_html(name: str, my_names=None, hide_pw: bool = False,
                      profile: str = accounts.VIEW_FULL) -> str:
        """渲染某 BU 页 + 可选注入（管理员隐藏自改密码 / 多 BU 账号的切换条）。缺页返回空串。
        profile：视图档案，末尾注入根节点 data-profile（BU 账号看=executive 精简，管理员看=full）。"""
        page = _state.get("bu_pages", {}).get(name)
        if not page:
            return ""
        parts = []
        if hide_pw:
            parts.append(_HIDE_PW_STYLE)
        if my_names:
            parts.append(_bu_switcher_html(my_names, name))
        html = page["html"]
        if any(parts):
            html = html.replace('<div class="wrap">', "".join(parts) + '<div class="wrap">', 1)
        return _apply_profile(html, profile)

    def _set_vcookie(resp, account: str):
        resp.set_cookie(VCOOKIE, _make_token(sec, account), max_age=SESSION_TTL,
                        httponly=True, samesite="lax")
        return resp

    def _set_acookie(resp, account: str):
        resp.set_cookie(COOKIE, _make_token(sec, account), max_age=SESSION_TTL,
                        httponly=True, samesite="lax")
        return resp

    def _use_fetch_shell() -> bool:
        """v1.4：已登录用户端经 shell fetch /api/v1/cockpit/view 拿同源 HTML。
        KANBAN_LEGACY_INLINE=1 或 unittest 运行时 → 直接吐 HTML（保旧测试与对照）。"""
        if os.environ.get("KANBAN_LEGACY_INLINE", "0") == "1":
            return False
        import sys
        if "unittest" in sys.modules or "pytest" in sys.modules:
            return False
        return (STATIC_DIR / "shell.html").is_file()

    def _shell_or_html(html: str):
        if _use_fetch_shell() and html:
            return FileResponse(STATIC_DIR / "shell.html", media_type="text/html; charset=utf-8")
        return HTMLResponse(html or "<h1>数据尚未生成，请稍候刷新</h1>")

    @app.get("/", response_class=HTMLResponse)
    def user_page(request: Request):
        """看板统一入口（v8.0 / v1.4 shell）：
        管理员会话 → 整体页；整体权限 → 整体页（带 BU 入口条）；BU 权限 → 本 BU 页；
        未登录 → 登录页。v1.4 已登录走 static/shell.html → fetch 像素级 HTML。"""
        if _user(request):
            return _shell_or_html(_main_with_nav(hide_pw=True) or "")
        acc = _vacc_row(request)
        if acc:
            if accounts.is_main(acc):
                return _shell_or_html(_main_with_nav(profile=accounts.view_profile(acc)) or "")
            names = accounts.bu_names_of(acc)  # 多 BU：绑定名单（旧单 BU 账号=[该名]）
            if names:
                existing = [n for n in names if n in _state.get("bu_pages", {})]
                if not existing:
                    return HTMLResponse(_view_login_page(
                        "你绑定的 BU 已被管理员移除，请重新登录或联系管理员"))
                # BU 账号：仍直接出 BU HTML（隔离铁律；壳只服务整体页 API）
                return HTMLResponse(_bu_view_html(existing[0], names,
                                                  profile=accounts.view_profile(acc)))
            # 管理员账号误走查看 cookie：引导去 /admin
            if accounts.is_admin(acc):
                return RedirectResponse("/admin", status_code=303)
        return HTMLResponse(_view_login_page())

    def _main_with_nav(hide_pw: bool = False, profile: str = accounts.VIEW_FULL) -> str:
        """整体页 + BU 入口条（只有整体/管理员会话能拿到本页，无泄漏面）。
        看端不展示「未归属 BU」文案（管理端设置页「BU 数据归属」仍提示待配置）。
        hide_pw=True（管理员会话看）：隐藏右上「🔑密码」自改密码入口——管理员改密码走 /admin「设置→账号与权限」，
        避免在内嵌看板里误改（管理员本无查看会话，点了也只会 401，属确认无用的入口）。
        profile：视图档案（full=管理员完整 / executive=整体·姜总精简），末尾按档案注入根节点 data-profile。"""
        html = _state["user_html"]
        if not html:
            return html
        from urllib.parse import quote

        def _esc(s):
            return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
        parts = []
        if hide_pw:
            parts.append(_HIDE_PW_STYLE)
        names = list(_state.get("bu_pages", {}))
        if names:
            links = "".join(f'<a class="bu-nav-a" href="/bu/{quote(n)}">{_esc(n)}</a>' for n in names)
            parts.append('<div class="bu-nav" role="navigation" aria-label="BU 分页">'
                         '<span class="bu-nav-label">业务 BU 分页</span>'
                         '<span class="bu-nav-links">' + links + '</span></div>')
        if parts:
            html = html.replace('<div class="wrap">', "".join(parts) + '<div class="wrap">', 1)
        return _apply_profile(html, profile)

    @app.post("/login")
    def viewer_login(account: str = Form(""), password: str = Form("")):
        """账号+密码登录，按权限分流：管理员→/admin；整体→/；BU→/。
        账号不存在与密码错同一文案。"""
        account = account.strip()
        acc = accounts.authenticate(cfg, root, account, password)
        if not acc:
            return HTMLResponse(_view_login_page("账号或密码不正确", account), status_code=401)
        accounts.mark_login(cfg, root, account)
        if accounts.is_admin(acc):
            return _set_acookie(RedirectResponse("/admin", status_code=303), account)
        return _set_vcookie(RedirectResponse("/", status_code=303), account)

    @app.get("/bu/{name}", response_class=HTMLResponse)
    def bu_page(name: str, request: Request):
        """BU 页：本 BU 权限账号（含多 BU 绑定）/ 整体账号 / 管理员可看；未登录出登录页；不存在 404。"""
        page = _state.get("bu_pages", {}).get(name)
        if not page:
            raise HTTPException(status_code=404, detail="Not Found")
        if not _can_view_bu(request, name):
            return HTMLResponse(_view_login_page())
        if _user(request):  # 管理员看 BU 页也隐藏「🔑密码」自改入口（与整体页一致）
            return HTMLResponse(page["html"].replace(
                '<div class="wrap">', _HIDE_PW_STYLE + '<div class="wrap">', 1))
        # 多 BU 账号：注入「我的 BU」切换条（只列其绑定的 BU）；整体账号 bu_names_of=[] 不注入
        vacc = _vacc_row(request)
        my = accounts.bu_names_of(vacc)
        return HTMLResponse(_bu_view_html(name, my, profile=accounts.view_profile(vacc)))

    # ---------- v1.4 JSON API（只序列化 summary，不算账）----------
    @app.get("/api/v1/session")
    def api_v1_session(request: Request):
        admin = _user(request)
        if admin:
            acc = accounts.find_account(cfg, root, admin)
            return api_v1.session_public(acc, is_admin_session=True)
        acc = _vacc_row(request)
        if not acc:
            raise HTTPException(status_code=401, detail="未登录")
        return api_v1.session_public(acc)

    @app.post("/api/v1/login")
    def api_v1_login(payload: dict = Body(default={})):
        account = str(payload.get("account") or "").strip()
        password = str(payload.get("password") or "")
        acc = accounts.authenticate(cfg, root, account, password)
        if not acc:
            raise HTTPException(status_code=401, detail="账号或密码不正确")
        accounts.mark_login(cfg, root, account)
        if accounts.is_admin(acc):
            sess = api_v1.session_public(acc, is_admin_session=True)
            resp = JSONResponse({"ok": True, "redirect": "/admin", "session": sess})
            return _set_acookie(resp, account)
        sess = api_v1.session_public(acc)
        redir = "/"
        if not accounts.is_main(acc):
            names = accounts.bu_names_of(acc)
            if names:
                redir = f"/bu/{names[0]}"
        resp = JSONResponse({"ok": True, "redirect": redir, "session": sess})
        return _set_vcookie(resp, account)

    @app.post("/api/v1/logout")
    def api_v1_logout():
        resp = JSONResponse({"ok": True})
        resp.delete_cookie(COOKIE)
        resp.delete_cookie(VCOOKIE)
        return resp

    @app.get("/api/v1/cockpit")
    def api_v1_cockpit(request: Request):
        """整体驾驶舱 JSON（数字与 golden 全等；前端/飞书等复用）。"""
        if not (_vacct(request) or _user(request)):
            raise HTTPException(status_code=401, detail="未登录")
        if not _can_view_main(request):
            raise HTTPException(status_code=403, detail="无整体驾驶舱权限")
        summary = _state.get("summary")
        if not summary:
            raise HTTPException(status_code=503, detail="数据尚未生成")
        out = api_v1.cockpit_payload(summary, scope="整体")
        if _state.get("built_at"):
            out.setdefault("meta", {})["built_at"] = _state["built_at"]
        return out

    @app.get("/api/v1/cockpit/bu/{name}")
    def api_v1_cockpit_bu(name: str, request: Request):
        if not (_vacct(request) or _user(request)):
            raise HTTPException(status_code=401, detail="未登录")
        if not _can_view_bu(request, name):
            raise HTTPException(status_code=403, detail="无权查看该 BU")
        page = (_state.get("bu_pages") or {}).get(name)
        if not page:
            raise HTTPException(status_code=404, detail="BU 不存在或未配置")
        summary = page.get("summary")
        if not summary:
            # 基准版 bu_pages 仅有 html：现场重算会动 core——此处 503 提示需带 summary 的发布
            raise HTTPException(status_code=503, detail="该 BU 尚无 JSON 快照（请更新数据）")
        return api_v1.cockpit_payload(summary, scope="BU", bu_name=name)

    @app.get("/api/v1/cockpit/view", response_class=HTMLResponse)
    def api_v1_cockpit_view(request: Request):
        """像素级同源：返回与 / 完全同一套 render_dashboard HTML（只读缓存/现算展示层）。"""
        if not (_vacct(request) or _user(request)):
            raise HTTPException(status_code=401, detail="未登录")
        if not _can_view_main(request):
            raise HTTPException(status_code=403, detail="无整体驾驶舱权限")
        html = _state.get("user_html") or ""
        if not html:
            raise HTTPException(status_code=503, detail="数据尚未生成")
        # 与 / 整体页一致：注入 BU 条与 profile
        if _user(request):
            return HTMLResponse(_main_with_nav(hide_pw=True) or html)
        acc = _vacc_row(request)
        return HTMLResponse(_main_with_nav(profile=accounts.view_profile(acc)) or html)

    @app.post("/api/my_passwd")
    def api_my_passwd(request: Request, payload: dict = Body(default={})):
        """看的人自改密码（整体页/BU 页右上 🔑）：验旧设新，写回 看板账号.json 明文。"""
        name = _vacct(request)
        if not name:
            raise HTTPException(status_code=401, detail="请先登录看板")
        old, new = str(payload.get("old") or ""), str(payload.get("new") or "")
        err = accounts.change_password(cfg, root, name, old, new)
        if err:
            raise HTTPException(status_code=400, detail=err)
        _audit(cfg, root, name, ("密码", f"账号 {name} 自改密码"))  # C3：不记密码内容
        return {"note": "密码已修改"}

    @app.get("/api/accounts")
    def api_accounts_get(request: Request):
        """账号表（管理员会话）：含明文密码。绝不出现在其他出口。"""
        _require(request)
        rows = [accounts.public_row(a, with_password=True) for a in accounts.load_accounts(cfg, root)]
        return {"accounts": rows, "count": len(rows),
                "master_account": accounts.MASTER_ACCOUNT}

    @app.post("/api/accounts")
    def api_accounts_post(request: Request, payload: dict = Body(default={})):
        """保存账号表（管理员）。至少保留一个管理员；总账号不可删。C3：变更留痕（密码只记「改密码」）。"""
        user = _require(request)
        raw = payload.get("accounts")
        if not isinstance(raw, list):
            raise HTTPException(status_code=400, detail="accounts 须为列表")
        if len(raw) > 50:
            raise HTTPException(status_code=400, detail="账号数量过多（上限 50）")
        old_accs = accounts.load_accounts(cfg, root, create=False)
        try:
            saved = accounts.save_accounts(cfg, root, raw)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        _audit(cfg, root, user, _diff_accounts(old_accs, saved))
        rows = [accounts.public_row(a, with_password=True) for a in saved]
        return {"accounts": rows, "count": len(rows), "note": "已保存",
                "master_account": accounts.MASTER_ACCOUNT}

    @app.get("/admin", response_class=HTMLResponse)
    def admin_page(request: Request):
        """管理员控制台：仅 static/admin（+ /admin/app.js）。
        _state['admin_html'] 仅作「是否已首次取数成功」标记（truthy=完整台，空=引导页）。"""
        if _user(request):
            # 数据未生成（空机器首次部署）→ 引导页：填智云账号→立即更新→自动进完整管理端（F-02）
            if not _state.get("admin_html"):
                return HTMLResponse(_bootstrap_page())
            return HTMLResponse(_admin_static_html())
        return HTMLResponse(_login_page())

    @app.get("/admin/app.js")
    def admin_app_js(request: Request):
        """管理端应用 JS：磁盘 static/admin/admin.js 与抽取常量一致，
        仅将 __MANUAL_ITEMS__ 换成当前 config 手填项 JSON（纯注入、不算账）。"""
        from fastapi.responses import Response
        js_path = STATIC_DIR / "admin" / "admin.js"
        if not js_path.is_file():
            raise HTTPException(status_code=404, detail="admin.js missing")
        raw = js_path.read_text(encoding="utf-8")
        body = raw.replace("__MANUAL_ITEMS__", _manual_items_json(cfg))
        return Response(body, media_type="application/javascript; charset=utf-8",
                        headers={"Cache-Control": "no-store"})

    @app.post("/admin/login")
    def admin_login(account: str = Form(""), password: str = Form(""),
                    identity: str = Form("")):  # identity 兼容旧表单字段名，忽略
        account = (account or identity or "").strip()
        acc = accounts.authenticate(cfg, root, account, password)
        if not acc or not accounts.is_admin(acc):
            return HTMLResponse(_login_page("账号或密码不正确", account), status_code=401)
        accounts.mark_login(cfg, root, account)
        return _set_acookie(RedirectResponse("/admin", status_code=303), account)

    @app.get("/admin/logout")
    def admin_logout():
        resp = RedirectResponse("/admin", status_code=303)
        resp.delete_cookie(COOKIE)
        return resp

    @app.get("/api/detail_export")
    def api_detail_export(request: Request, table: str = Query("收入明细"),
                          month: str | None = None, q: str | None = None):
        """当前筛选结果导出 Excel（.xlsx · 管理员；上限 5000 行，避免拖垮）。
        表头+行与明细页一致；月份/搜索条件与页面筛选相同。"""
        _require(request)
        import io
        from urllib.parse import quote
        from fastapi.responses import Response
        import openpyxl
        conn = _conn()
        try:
            d = db.query_detail(conn, table, month, q, page=1, page_size=5000)
        except KeyError as e:
            raise HTTPException(status_code=400, detail=str(e))
        finally:
            conn.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        # sheet 名最多 31 字，去掉 Excel 非法字符
        safe = "".join(c for c in str(table) if c not in r'[]:*?/\\')[:31] or "明细"
        ws.title = safe
        cols = d["columns"]
        ws.append(list(cols))
        for r in d["rows"]:
            ws.append([r.get(c, "") if r.get(c, "") is not None else "" for c in cols])
        # 首行粗体 + 简单列宽
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for i, col in enumerate(cols, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max(len(str(col)) + 4, 10), 28)
        bio = io.BytesIO()
        wb.save(bio)
        raw = bio.getvalue()
        day = time.strftime("%Y%m%d")
        fname = f"{safe}_{day}.xlsx"
        # RFC 5987：中文文件名用 filename*
        cd = f"attachment; filename=\"export.xlsx\"; filename*=UTF-8''{quote(fname)}"
        return Response(
            content=raw,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": cd},
        )

    @app.get("/api/detail")
    def api_detail(request: Request, table: str = Query("收入明细"), month: str | None = None,
                   q: str | None = None, page: int = 1, page_size: int = 50,
                   unclassified: bool = False, unfilled_dept: bool = False):
        user = _user(request)
        if not user:
            raise HTTPException(status_code=401, detail="需要管理员登录")
        conn = db.connect(cfg, root)
        try:
            try:
                return JSONResponse(db.query_detail(conn, table, month, q, page, page_size,
                                                    unclassified, unfilled_dept))
            except KeyError as e:
                raise HTTPException(status_code=400, detail=str(e))
        finally:
            conn.close()

    @app.get("/api/daily")
    def api_daily(request: Request, start: str = Query(""), end: str = Query(""), top: int = Query(10)):
        """按天明细（用户端「明细」入口·迭代计划13批次B）：任意日期区间的逐日下单/回款 + 期内排名。
        v7.8 起要求整体页/管理员会话（全公司口径出口，BU 会话不给——否则 BU 链接持有者可绕过页面隔离）；
        **纯只读**、无任何写路径；金额显示串全部后端算好（铁律2）。入参严格校验：ISO日期、start<=end、区间≤366天。"""
        if not _can_view_main(request):
            raise HTTPException(status_code=401, detail="请先登录看板")
        import datetime as _dt
        try:
            s = _dt.date.fromisoformat(start)
            e = _dt.date.fromisoformat(end)
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="日期格式须为 YYYY-MM-DD")
        if e < s:
            raise HTTPException(status_code=400, detail="结束日期须不早于开始日期")
        if (e - s).days > 366:
            raise HTTPException(status_code=400, detail="区间最长 366 天")
        top = max(1, min(2000, int(top)))   # 排名条数：默认前10，「其余点开看明细」传 2000 拿全量
        conn = db.connect(cfg, root)
        try:
            orders = db.load_orders(cfg, conn)
            receipts = db.load_receipts(cfg, conn)
        finally:
            conn.close()
        import charts
        import profit as _profit
        # 与全年预渲染一致：有销售→BU 映射则多算 orders_by_bu（看端时间段查询统一按 BU）
        sales_to_bu = None
        try:
            import bu as _bu
            bucfg = _bu.load_bu_config(cfg, root)
            if bucfg and bucfg.get("bus"):
                sales_to_bu = {}
                for b in bucfg["bus"]:
                    for sal in (b.get("销售") or []):
                        sales_to_bu.setdefault(str(sal).strip(), b["name"])
                if not sales_to_bu:
                    sales_to_bu = None
        except Exception:
            sales_to_bu = None
        d = _profit.compute_daily(orders, receipts, cfg["columns"], s, e, top=top,
                                  sales_to_bu=sales_to_bu)

        def _wan(v):   # 显示串：与排名卡一致（负数全角−）
            return ("−" if v < 0 else "") + charts.fmt_wan(abs(v)) + "万"

        for row in d["days"]:
            row["orders_disp"], row["receipts_disp"] = _wan(row.pop("orders")), _wan(row.pop("receipts"))
        t = d["totals"]
        t["orders_disp"], t["receipts_disp"] = _wan(t.pop("orders")), _wan(t.pop("receipts"))
        for rk in d["rankings"].values():
            for it in rk["items"]:
                it["disp"] = _wan(it.pop("amount"))
            if rk.get("others"):
                rk["others"]["disp"] = _wan(rk["others"].pop("amount"))
            if rk.get("unfilled"):
                rk["unfilled"]["disp"] = _wan(rk["unfilled"].pop("amount"))
            rk.pop("total", None)   # 用不到就不下发，防前端拿去做运算
        return {"start": start, "end": end, **d}

    @app.get("/api/profit_ranking")
    def api_profit_ranking(request: Request, dim: str = Query(""), start: str = Query(""),
                           end: str = Query(""), top: int = Query(5000)):
        """板块③「收入与毛利结构」全量明细（「其余 N 个」点开）：确认口径 收入/毛利 按客户/销售。
        与 /api/daily 同为全公司口径出口——要整体页/管理员会话（BU 会话 401，防绕过页面隔离）；
        **纯只读**；金额/毛利率显示串全部后端算好（铁律2）。入参严格校验：dim∈{customer,sales}、ISO 日期、区间≤366天。"""
        if not _can_view_main(request):
            raise HTTPException(status_code=401, detail="请先登录看板")
        name_col = {"customer": "客户", "sales": "销售"}.get(dim)
        if not name_col:
            raise HTTPException(status_code=400, detail="dim 须为 customer 或 sales")
        import datetime as _dt
        try:
            s = _dt.date.fromisoformat(start)
            e = _dt.date.fromisoformat(end)
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="日期格式须为 YYYY-MM-DD")
        if e < s:
            raise HTTPException(status_code=400, detail="结束日期须不早于开始日期")
        if (e - s).days > 366:
            raise HTTPException(status_code=400, detail="区间最长 366 天")
        top = max(1, min(5000, int(top)))
        conn = db.connect(cfg, root)
        try:
            project = db.load_project_detail(cfg, conn)
        finally:
            conn.close()
        import charts
        import profit as _profit
        vat = cfg["tax"]["vat_rate"]
        rk = _profit.compute_profit_ranking(project, name_col, cfg["columns"], s, e, vat, top=top)

        def _wan(v):
            return ("−" if v < 0 else "") + charts.fmt_wan(abs(v)) + "万"

        def _mg(it):
            # 陆总0714：改叫「系统成本率」；按销售的率先不显示（防"人力算不算"连锁追问）
            if dim == "sales":
                return ""
            cp = it.get("cost_pct")
            return f"系统成本率 {cp:.0f}%" if cp is not None else "系统成本率 —"

        items = [{"name": it["name"], "revenue_disp": _wan(it["revenue"]), "margin_disp": _mg(it)}
                 for it in rk["items"]]
        if rk.get("unfilled"):
            uf = rk["unfilled"]
            items.append({"name": "（未填）", "revenue_disp": _wan(uf["revenue"]),
                          "margin_disp": _mg(uf), "unfilled": True})
        return {"dim": dim, "start": start, "end": end, "items": items}

    @app.get("/api/exceptions")
    def api_exceptions(request: Request):
        """异常处理「总览」计数（管理员）。体检黄红是运行信号，留在 /api/health，不在这。"""
        _require(request)  # 同函数作用域下文定义，调用时已存在
        conn = db.connect(cfg, root)
        try:
            return db.exceptions_summary(conn)
        finally:
            conn.close()

    @app.get("/api/order_depts")
    def api_order_depts(request: Request):
        """下单表已出现过的部门清单（「下单未填部门」归类下拉用）。"""
        _require(request)
        conn = db.connect(cfg, root)
        try:
            return db.list_order_depts(conn)
        finally:
            conn.close()

    @app.get("/api/health")
    def api_health():
        """体检状态条数据源（公开：只给绿/黄/红 + 时间 + 各源行数，不含金额/客户名）。"""
        conn = db.connect(cfg, root)
        try:
            run_log = db.latest_run(conn)
        finally:
            conn.close()
        meta = (_state.get("summary") or {}).get("meta", {})
        health = meta.get("health", {})
        result = (run_log or {}).get("结果")               # 黄/红/绿：管道运行日志
        reasons = _run_reasons((run_log or {}).get("体检", {}))  # 「黄/红」：为啥（fetch/过期调整）
        # A3：未归属销售>0 → 至少判黄 + 顶栏短原因（沿用 v8.0 机制；不覆盖已判红）
        n_un = int((meta.get("unassigned") or {}).get("count") or 0)
        if n_un > 0:
            reasons = [f"{n_un} 名销售未归属 BU（业务不进任何 BU 页，各 BU 合计小于全公司）"] + reasons
            if result in ("绿", None):   # 未判红时至少判黄（无运行日志时 result 为 None 也升黄）
                result = "黄"
        return {
            "result": result,
            "run_time": (run_log or {}).get("时间"),
            "built_at": _state.get("built_at"),
            "sources": health.get("sources", []),
            "warnings": health.get("warnings", []),          # 「警」：数据体检（未填分类等）
            "run_reasons": reasons,
        }

    def _require(request: Request) -> str:
        user = _user(request)
        if not user:
            raise HTTPException(status_code=401, detail="需要管理员登录")
        return user

    def _conn():
        return db.connect(cfg, root)

    @app.post("/api/refresh")
    def api_refresh(request: Request):
        """立即更新=完整 pipeline（fetch+重读+重建+重放），后台线程跑、立即返回（在线抓约80秒）。
        运行中互斥，重复点返回进行中；进度轮询 /api/refresh_status。"""
        _require(request)
        if not start_refresh_async(cfg, root, "manual"):
            return JSONResponse({"status": "running", "detail": "更新进行中，请稍候"}, status_code=409)
        return {"status": "started", "refreshing": _state["refreshing"]}

    @app.get("/api/refresh_status")
    def api_refresh_status(request: Request):
        _require(request)
        return {"running": bool(_state["refreshing"]), "refreshing": _state["refreshing"],
                "last": _state["last_refresh"], "built_at": _state["built_at"],
                "zhiyun_auto_fetch": bool(cfg.get("zhiyun_auto_fetch"))}

    @app.get("/export.png")
    def api_export_png(request: Request, blk: str = ""):
        """导出=当前所选周期的整页 PNG（服务端 Playwright 截图）。v7.8 起要求整体页/管理员会话
        （导出的是全公司主页）。"""
        if not _can_view_main(request):
            raise HTTPException(status_code=401, detail="请先登录看板")
        html = _state.get("user_html")
        if not html:
            raise HTTPException(status_code=503, detail="页面尚未构建，稍后再试")
        keys = set(((_state.get("summary") or {}).get("periods") or {}).keys())
        if blk and keys and blk not in keys:
            raise HTTPException(status_code=400, detail="未知周期")
        if not _EXPORT_LOCK.acquire(blocking=False):
            raise HTTPException(status_code=429, detail="正在生成另一张导出图，请稍候几秒再点")
        try:
            png = _screenshot_png(html, blk)
        except HTTPException:
            raise
        except Exception as e:  # noqa: BLE001 chromium 未装/超时等
            raise HTTPException(status_code=503,
                                detail=f"截图失败（{type(e).__name__}: {e}）；部署机需先 playwright install chromium")
        finally:
            _EXPORT_LOCK.release()
        label = blk or ((_state.get("summary") or {}).get("meta") or {}).get("year_key", "")
        from urllib.parse import quote
        fn = quote(f"甲骨易智能经营罗盘_{label}_{time.strftime('%Y%m%d_%H%M')}.png")
        return Response(content=png, media_type="image/png",
                        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fn}",
                                 "X-Filename": fn})

    @app.get("/bu/{name}/export.png")
    def api_bu_export_png(name: str, request: Request, blk: str = ""):
        """BU 页导出（迭代22·D5）：截该 BU 页整页 PNG。会话闸=能看该 BU 才能导（铁律12：
        截图源就是该 BU 已过滤页面，天然不含他 BU 数据）。"""
        page = _state.get("bu_pages", {}).get(name)
        if not page:
            raise HTTPException(status_code=404, detail="Not Found")
        if not _can_view_bu(request, name):
            raise HTTPException(status_code=401, detail="请先登录看板")
        keys = set(((_state.get("summary") or {}).get("periods") or {}).keys())
        if blk and keys and blk not in keys:
            raise HTTPException(status_code=400, detail="未知周期")
        if not _EXPORT_LOCK.acquire(blocking=False):
            raise HTTPException(status_code=429, detail="正在生成另一张导出图，请稍候几秒再点")
        try:
            png = _screenshot_png(page["html"], blk)
        except HTTPException:
            raise
        except Exception as e:  # noqa: BLE001
            raise HTTPException(status_code=503,
                                detail=f"截图失败（{type(e).__name__}: {e}）；部署机需先 playwright install chromium")
        finally:
            _EXPORT_LOCK.release()
        label = blk or ((_state.get("summary") or {}).get("meta") or {}).get("year_key", "")
        from urllib.parse import quote
        fn = quote(f"甲骨易智能经营罗盘_{name}_{label}_{time.strftime('%Y%m%d_%H%M')}.png")
        return Response(content=png, media_type="image/png",
                        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fn}",
                                 "X-Filename": fn})

    @app.get("/api/history")
    def api_history(request: Request):
        """历史页面快照列表（按天，倒序）。供管理员端「历史快照」页。"""
        _require(request)
        bdir = loaders.data_dir(cfg, root) / "备份"
        out = []
        for p in sorted(bdir.glob("页面_*.html"), reverse=True):
            d = p.stem.split("_")[1]
            out.append({"day": d, "label": f"{d[:4]}-{d[4:6]}-{d[6:]}",
                        "saved_at": time.strftime("%Y-%m-%d %H:%M", time.localtime(p.stat().st_mtime)),
                        "kb": round(p.stat().st_size / 1024)})
        return out

    @app.get("/api/history/{day}")
    def api_history_page(request: Request, day: str):
        """回看某天的看板页面（当天最后一次更新的原样快照）。"""
        _require(request)
        if not re.fullmatch(r"\d{8}", day):
            raise HTTPException(status_code=400, detail="日期格式须为 YYYYMMDD")
        p = loaders.data_dir(cfg, root) / "备份" / f"页面_{day}.html"
        if not p.exists():
            raise HTTPException(status_code=404, detail="该日无页面快照")
        return HTMLResponse(p.read_text(encoding="utf-8"))

    @app.get("/api/bu_config")
    def api_bu_config_get(request: Request):
        """BU 配置（管理员会话）：BU 清单/负责人/销售名单/分摊比例 + 分摊总开关。"""
        _require(request)
        bucfg = bu.load_bu_config(cfg, root) or {"bus": [], "公共费用分摊启用": False}
        return {"bus": bucfg["bus"], "count": len(bucfg["bus"]),
                "公共费用分摊启用": bool(bucfg.get("公共费用分摊启用"))}

    @app.get("/api/sales_pool")
    def api_sales_pool(request: Request):
        """四源销售池（管理员·A1 归属页）：供批量/拖拽归属。含配置里有、库里暂无的名字（rows=0）。
        每人带当年下单笔数+金额参考串（服务端算好=铁律2）；顶层带 A3 未归属计数+当年未归属下单额。"""
        _require(request)
        today = loaders.pinned_today(cfg)
        conn = db.connect(cfg, root)
        try:
            from_db = db.list_salespeople(conn)
            ostats = db.order_stats_by_sales(conn, today.year)
            snap = core.unassigned_snapshot(cfg, conn, today, root)
        finally:
            conn.close()
        by = {x["name"]: x["rows"] for x in from_db}
        bucfg = bu.load_bu_config(cfg, root) or {"bus": []}
        for b in bucfg.get("bus", []):
            for s in b.get("销售") or []:
                s = str(s).strip()
                if s and s not in by:
                    by[s] = 0

        def _ref(name):
            st = ostats.get(name)
            if not st or not st["count"]:
                return {"orders_count": 0, "ref_disp": "当年无下单"}
            return {"orders_count": st["count"],
                    "ref_disp": f'{st["count"]} 笔 · {core._unassigned_wan(st["amount"])[1:]}'}

        people = [{"name": n, "rows": by[n], **_ref(n)}
                  for n in sorted(by.keys(), key=lambda k: (-by[k], k))]
        return {"sales": people, "count": len(people), **snap}

    @app.post("/api/bu_config")
    def api_bu_config_post(request: Request, payload: dict = Body(default={})):
        """保存 BU 数据归属 + 公共费用分摊，并立即重算重渲染 BU 页（一人一 BU）。C3：变更留痕。"""
        user = _require(request)
        bus = payload.get("bus")
        if not isinstance(bus, list):
            raise HTTPException(status_code=400, detail="bus 须为列表")
        if len(bus) > 20:
            raise HTTPException(status_code=400, detail="BU 数量过多（上限 20）")
        old = bu.load_bu_config(cfg, root) or {"bus": [], "公共费用分摊启用": False}
        old_bus, old_alloc = old["bus"], bool(old.get("公共费用分摊启用"))
        new_alloc = bool(payload.get("公共费用分摊启用", False))
        try:
            saved = bu.save_bu_config(cfg, root, bus, 公共费用分摊启用=new_alloc)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        recompute(cfg, root)
        _audit(cfg, root, user, _diff_bu_config(
            old_bus, saved["bus"], old_alloc, bool(saved.get("公共费用分摊启用"))))
        return {"bus": saved["bus"], "count": len(saved["bus"]),
                "公共费用分摊启用": bool(saved.get("公共费用分摊启用")),
                "note": "已保存并重算"}

    @app.get("/api/config_changes")
    def api_config_changes(request: Request, category: str | None = None, limit: int = 200):
        """C3 操作记录（管理员）：配置变更留痕倒序，可按类别筛。仅摘要，无密码明文。"""
        _require(request)
        conn = db.connect(cfg, root)
        try:
            return {"changes": db.list_config_changes(conn, category or None, limit),
                    "categories": list(db.CONFIG_CHANGE_CATEGORIES)}
        finally:
            conn.close()

    @app.get("/api/version")
    def api_version(request: Request):
        """产品版本号 + 面向用户的更新日志（管理员会话）。
        版本号=根目录 VERSION（现 1.0-beta 公测 Beta），与 git 开发号(v8.x)分开、不给普通用户看。"""
        _require(request)
        return product_version.version_info()

    @app.get("/api/update/check")
    def api_update_check(request: Request):
        """④ 检测远端有没有新版本（管理员会话）：git fetch + 比对 HEAD 与 <update_remote>/分支。
        对标的远端由 config `update_remote` 决定（默认 origin；部署机从 Gitee clone 则 origin 即 Gitee）。
        只读、带护栏（非仓库/分叉/脏工作区不给更新），返回是否有新版本与"要更新啥"。"""
        _require(request)
        return updater.check_update(loaders.ROOT, remote=cfg.get("update_remote") or "origin")

    @app.post("/api/update/apply")
    def api_update_apply(request: Request):
        """④ 一键更新（管理员会话）：复检护栏 → git pull --ff-only <update_remote> → 触发看门狗重启。
        拉取成功才重启（进程以退出码 42 退出，看门狗用新代码拉起）；失败原样返回不重启。"""
        user = _require(request)
        res = updater.apply_update(loaders.ROOT, remote=cfg.get("update_remote") or "origin")
        if res.get("ok"):
            _audit(cfg, root, user,
                   ("更新", f"一键更新 {res.get('from') or '?'}→{res.get('to') or '?'}"
                            f"（{res.get('pulled') or 0} 个提交）"))
            updater.request_restart()   # 后台延时退出→看门狗重启；HTTP 响应先发回
            res["restarting"] = True
        return res

    @app.get("/api/settings")
    def api_settings_get(request: Request):
        _require(request)
        out = {k: cfg.get(k) for k in EDITABLE_SETTINGS}
        out["schedule_times"] = get_schedule_times(cfg)  # ②多次更新：列表（缺失从旧单值推导）
        creds = read_zhiyun_creds(cfg, root)
        out["zhiyun_username"], out["zhiyun_password"] = creds["username"], creds["password"]
        out["zhiyun_conn"] = read_zhiyun_conn(cfg, root)  # 服务器地址+四表ID（内置默认+本地覆盖的生效值）
        out["ledger_share_path"] = cfg.get("ledger_share_path", "")  # 收单台账共享盘路径（界面填·落本地覆盖）
        bdir = loaders.data_dir(cfg, root) / "备份"
        baks = (sorted(bdir.glob("看板_*.db")) + sorted(bdir.glob("页面_*.html"))) if bdir.exists() else []
        out["backup_stats"] = {"count": len(baks),
                               "mb": round(sum(p.stat().st_size for p in baks) / 1048576, 1)}
        return out

    @app.post("/api/settings")
    def api_settings_post(request: Request, payload: dict = Body(default={})):
        user = _require(request)
        old_times = get_schedule_times(cfg)
        old_keep = cfg.get("backup_keep_days")
        old_lsp = cfg.get("ledger_share_path")
        try:
            res = save_settings(cfg, root, payload)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        chg = []  # C3：设置变更留痕（智云账号只记「已更换」不记值）
        if ("schedule_times" in payload or "schedule_time" in payload) \
                and res["schedule_times"] != old_times:
            chg.append(f"更新时间 {'、'.join(old_times) or '—'}→{'、'.join(res['schedule_times'])}")
        if "backup_keep_days" in payload and res["backup_keep_days"] != old_keep:
            chg.append(f"备份保留 {old_keep}→{res['backup_keep_days']} 天")
        if "智云账号已更新" in (res.get("note") or ""):
            chg.append("智云账号已更换")
        if "智云连接配置已更新" in (res.get("note") or ""):
            chg.append("智云连接配置已更改（服务器/表ID）")
        # 台账路径含内网服务器名（敏感）→ 只记「已更改」不落值（铁律16）
        if "ledger_share_path" in payload and str(payload.get("ledger_share_path") or "").strip() != str(old_lsp or "").strip():
            chg.append("收单台账共享盘路径已更改")
        if chg:
            _audit(cfg, root, user, ("设置", "设置：" + "；".join(chg)))
        return res

    @app.post("/api/adjust")
    def api_adjust(request: Request, payload: dict = Body(default={})):
        user = _require(request)
        conn = _conn()
        try:
            aid = db.add_adjustment(conn, user, payload.get("目标表", ""),
                                    payload.get("定位键", ""), payload.get("字段", ""),
                                    payload.get("新值", ""), payload.get("原因", ""),
                                    payload.get("类型", "改值"))
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        finally:
            conn.close()
        recompute(cfg, root)
        return {"status": "ok", "adj_id": aid, "built_at": _state["built_at"]}

    @app.post("/api/adjust/{adj_id}/revoke")
    def api_revoke(request: Request, adj_id: int):
        _require(request)
        conn = _conn()
        try:
            ok = db.revoke_adjustment(conn, adj_id)
        finally:
            conn.close()
        recompute(cfg, root)
        return {"status": "ok" if ok else "noop", "built_at": _state["built_at"]}

    @app.post("/api/adjust/expired/revoke_all")
    def api_revoke_all_expired(request: Request):
        """批量撤销全部「过期疑似」=一键听源头新值。前端走"点按钮→确认保存"两步，这里只管执行。"""
        _require(request)
        conn = _conn()
        try:
            n = db.revoke_expired_adjustments(conn)
        finally:
            conn.close()
        if n:
            recompute(cfg, root)
        return {"status": "ok", "revoked": n, "built_at": _state["built_at"]}

    @app.post("/api/adjust/{adj_id}/rearm")
    def api_rearm(request: Request, adj_id: int):
        """坚持我的数（仅过期疑似、仅逐条）：原值刷新为源头现值→重新生效→立即重算。"""
        _require(request)
        conn = _conn()
        try:
            db.rearm_adjustment(conn, adj_id)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        finally:
            conn.close()
        recompute(cfg, root)
        return {"status": "ok", "built_at": _state["built_at"]}

    @app.get("/api/adjustments")
    def api_adjustments(request: Request):
        _require(request)
        conn = _conn()
        try:
            return db.list_adjustments(conn)
        finally:
            conn.close()

    @app.get("/api/manual")
    def api_manual_get(request: Request, month: str | None = None, scope: str = "全公司"):
        _require(request)
        conn = _conn()
        try:
            return db.get_manual(conn, month, 范围=scope or "全公司")
        finally:
            conn.close()

    @app.post("/api/manual")
    def api_manual_set(request: Request, payload: dict = Body(default={})):
        user = _require(request)
        item = payload.get("项目", "")
        if item not in {it["name"] for it in cfg["manual_items"]}:
            raise HTTPException(status_code=400, detail=f"未知手填项目：{item}")
        try:
            金额 = float(payload.get("金额"))
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="金额须为数字")
        scope = str(payload.get("范围") or "全公司").strip() or "全公司"
        conn = _conn()
        try:
            db.set_manual(conn, payload.get("归属月", ""), item, 金额, user, 范围=scope)
        finally:
            conn.close()
        recompute(cfg, root)
        return {"status": "ok", "built_at": _state["built_at"]}

    @app.post("/api/manual_batch")
    def api_manual_batch(request: Request, payload: dict = Body(default={})):
        """批量手填：payload={归属月, 范围?, items:[{项目,金额,范围?}]}，只重算一遍。"""
        user = _require(request)
        month = payload.get("归属月", "")
        default_scope = str(payload.get("范围") or "全公司").strip() or "全公司"
        items = payload.get("items") or []
        if not isinstance(items, list) or not items:
            raise HTTPException(status_code=400, detail="items 不能为空")
        names = {it["name"] for it in cfg["manual_items"]}
        conn = _conn()
        try:
            n = 0
            for it in items:
                item = (it or {}).get("项目", "")
                if item not in names:
                    raise HTTPException(status_code=400, detail=f"未知手填项目：{item}")
                try:
                    金额 = float((it or {}).get("金额"))
                except (TypeError, ValueError):
                    raise HTTPException(status_code=400, detail=f"金额须为数字：{item}")
                sc = str((it or {}).get("范围") or default_scope).strip() or "全公司"
                db.set_manual(conn, month, item, 金额, user, 范围=sc)
                n += 1
        finally:
            conn.close()
        recompute(cfg, root)
        return {"status": "ok", "count": n, "built_at": _state["built_at"]}

    def _alloc_month_payload(conn, month: str) -> dict:
        """某月分摊面板数据：BU 名单（与设置页同源）+ 比例 + 本月公共费用总额/剩余（显示串后端下发·铁律2）。"""
        import datetime as _dt
        import columns as _columns
        try:
            y, m = int(month[:4]), int(month[5:7])
            assert 1 <= m <= 12 and month[4] == "-"
        except (ValueError, AssertionError, IndexError):
            raise HTTPException(status_code=400, detail="归属月格式须为 YYYY-MM")
        bucfg = bu.load_bu_config(cfg, root) or {"bus": []}
        bu_names = [b["name"] for b in bucfg["bus"]]
        # 陆总0714：该月没填 → 回显沿用的最近填写月比例（inherited_from 标来源；保存即固化到本月）
        ratios, src_month = db.effective_alloc_month(conn, month)
        inherited_from = src_month if (src_month and src_month != month) else None
        lh, lr = db.load_ledger(cfg, conn)
        month_total = 0.0
        if lh:
            lcols = _columns.resolve_ledger_columns(lh)
            public_rows = profit.filter_ledger_rows_by_pc(lh, lr, {"公共"})
            start = _dt.date(y, m, 1)
            end = _dt.date(y, m + 1, 1) - _dt.timedelta(days=1) if m < 12 else _dt.date(y, 12, 31)
            led, _ = profit.compute_ledger_expenses(public_rows, y, start, end, cfg, lcols)
            month_total = round(sum(float(v or 0) for v in led.values()), 2)
        known = {b: p for b, p in ratios.items() if b in set(bu_names)}
        sum_pct = round(sum(known.values()), 1)
        remain_pct = round(max(0.0, 100.0 - sum_pct), 1)
        remain_amt = round(month_total * remain_pct / 100.0, 2)
        orphans = sorted(set(ratios) - set(bu_names))
        return {"month": month, "bus": bu_names, "ratios": known,
                "inherited_from": inherited_from,
                "orphans": orphans,
                "month_total": month_total,
                "month_total_disp": f"{month_total:,.2f}",
                "sum_pct": sum_pct, "remain_pct": remain_pct,
                "remain_amt_disp": f"{remain_amt:,.2f}"}

    @app.get("/api/alloc_ratios")
    def api_alloc_get(request: Request, month: str = ""):
        _require(request)
        conn = _conn()
        try:
            return _alloc_month_payload(conn, month)
        finally:
            conn.close()

    @app.post("/api/alloc_ratios")
    def api_alloc_set(request: Request, payload: dict = Body(default={})):
        """写某月分摊比例（管理员）。payload={归属月, ratios:{BU:比例%|null}}。
        约束：BU 须在设置页 BU 名单内；单值 0~100；已知 BU 合计 ≤100（容差 0.05）。null=删行不分摊。"""
        user = _require(request)
        month = str(payload.get("归属月") or "").strip()
        ratios = payload.get("ratios")
        if not isinstance(ratios, dict) or not ratios:
            raise HTTPException(status_code=400, detail="ratios 不能为空")
        bucfg = bu.load_bu_config(cfg, root) or {"bus": []}
        known = {b["name"] for b in bucfg["bus"]}
        vals: dict[str, float | None] = {}
        for b, v in ratios.items():
            b = str(b).strip()
            if b not in known:
                raise HTTPException(status_code=400, detail=f"未知 BU：{b}（以设置页 BU 名单为准）")
            if v is None or v == "":
                vals[b] = None
                continue
            try:
                fv = float(v)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail=f"比例须为数字：{b}")
            if not (0 <= fv <= 100):
                raise HTTPException(status_code=400, detail=f"比例须在 0~100：{b}")
            vals[b] = round(fv, 1)
        conn = _conn()
        try:
            # 合并基准=该月生效比例（含沿用值·陆总0714）；保存时把生效全集固化进本月，
            # 否则只改一个 BU 会让其余 BU 的沿用比例丢失（本月一旦有行，沿用即不再兜底）
            merged, _src = db.effective_alloc_month(conn, month)
            merged = {b: p for b, p in merged.items() if b in known}
            for b, v in vals.items():
                if v is None:
                    merged.pop(b, None)
                else:
                    merged[b] = v
            total = sum(p for b, p in merged.items() if b in known)
            if total > 100.05:
                raise HTTPException(status_code=400,
                                    detail=f"该月各 BU 比例合计 {total:g}% 超过 100%，请调整（可以小于 100%，剩余留公司层）")
            for b in known:
                db.set_alloc_ratio(conn, month, b, merged.get(b), user)
            out = _alloc_month_payload(conn, month)
        finally:
            conn.close()
        _audit(cfg, root, user, ("分摊", f"公共费用分摊：{month} 比例已更改（合计 {out['sum_pct']:g}%）"))
        recompute(cfg, root)
        out.update({"status": "ok", "built_at": _state["built_at"]})
        return out

    def _detax_payload(conn) -> dict:
        """费用去税率录入页数据：可去税类别（含全年金额参考·降序）+ 已填税率。"""
        import charts  # 局部导入避免循环依赖（与本模块其余 charts 用法一致）
        cats = db.list_detax_categories(conn, cfg)
        rates = db.load_detax_rates(conn)
        return {
            "categories": [{"category": c["category"],
                            "amount_disp": charts.fmt_wan(c["amount"]) + "万"} for c in cats],
            "rates": rates,
        }

    @app.get("/api/detax_rates")
    def api_detax_get(request: Request):
        _require(request)
        conn = _conn()
        try:
            return _detax_payload(conn)
        finally:
            conn.close()

    @app.post("/api/detax_rates")
    def api_detax_set(request: Request, payload: dict = Body(default={})):
        """写费用去税率（管理员·全局一套·陆总0714）。payload={rates:{费用类别:税率%|null}}。
        税率 0~100；null/空/0 → 删行=该类别不去税（等价默认，页面数字回归红线中性）。"""
        user = _require(request)
        rates = payload.get("rates")
        if not isinstance(rates, dict) or not rates:
            raise HTTPException(status_code=400, detail="rates 不能为空")
        vals: dict[str, float | None] = {}
        for cat, v in rates.items():
            cat = str(cat).strip()
            if not cat:
                raise HTTPException(status_code=400, detail="费用类别不能为空")
            if v is None or v == "":
                vals[cat] = None
                continue
            try:
                fv = float(v)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail=f"去税率须为数字：{cat}")
            if not (0 <= fv <= 100):
                raise HTTPException(status_code=400, detail=f"去税率须在 0~100：{cat}")
            vals[cat] = round(fv, 2)
        conn = _conn()
        try:
            for cat, v in vals.items():
                db.set_detax_rate(conn, cat, v, user)
            out = _detax_payload(conn)
        finally:
            conn.close()
        changed = "、".join(f"{c}={v if v is not None else '清除'}" for c, v in vals.items())
        _audit(cfg, root, user, ("去税", f"费用去税率已更改：{changed}"))
        recompute(cfg, root)
        out.update({"status": "ok", "built_at": _state["built_at"]})
        return out

    @app.get("/api/budget")
    def api_budget_get(request: Request, year: str | None = None):
        _require(request)
        conn = _conn()
        try:
            return db.get_budget(conn, year)
        finally:
            conn.close()

    @app.post("/api/budget")
    def api_budget_set(request: Request, payload: dict = Body(default={})):
        user = _require(request)
        metric = payload.get("指标", "")
        if metric not in db.BUDGET_METRICS:
            raise HTTPException(status_code=400, detail=f"未知预算指标：{metric}")
        year = str(payload.get("年份", "")).strip()
        if not (year.isdigit() and len(year) == 4):
            raise HTTPException(status_code=400, detail="年份须为4位数字")
        try:
            金额 = float(payload.get("金额"))
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="金额须为数字")
        scope = str(payload.get("范围", "全公司")).strip() or "全公司"
        if metric == "费用年预算" and scope == "全公司":
            raise HTTPException(status_code=400, detail="费用年预算须指定部门（范围）")
        # 业务目标允许 全公司 或 BU 名；费用年预算允许部门名
        conn = _conn()
        try:
            db.set_budget(conn, year, metric, 金额, user, 范围=scope)
        finally:
            conn.close()
        recompute(cfg, root)
        return {"status": "ok", "built_at": _state["built_at"]}

    @app.post("/api/budget_batch")
    def api_budget_batch(request: Request, payload: dict = Body(default={})):
        """批量业绩目标：payload={items:[{年份,指标,金额,范围?}]}，一次重算。"""
        user = _require(request)
        items = payload.get("items") or []
        if not isinstance(items, list) or not items:
            raise HTTPException(status_code=400, detail="items 不能为空")
        conn = _conn()
        try:
            n = 0
            for it in items:
                it = it or {}
                metric = it.get("指标", "")
                if metric not in db.BUDGET_METRICS:
                    raise HTTPException(status_code=400, detail=f"未知预算指标：{metric}")
                year = str(it.get("年份", "")).strip()
                if not (year.isdigit() and len(year) == 4):
                    raise HTTPException(status_code=400, detail="年份须为4位数字")
                try:
                    金额 = float(it.get("金额"))
                except (TypeError, ValueError):
                    raise HTTPException(status_code=400, detail=f"金额须为数字：{metric}")
                scope = str(it.get("范围", "全公司")).strip() or "全公司"
                if metric == "费用年预算" and scope == "全公司":
                    raise HTTPException(status_code=400, detail="费用年预算须指定部门（范围）")
                if "毛利率" in metric and (金额 < 0 or 金额 > 100):
                    raise HTTPException(status_code=400, detail=f"毛利率须为 0~100：{metric}")
                db.set_budget(conn, year, metric, 金额, user, 范围=scope)
                n += 1
        finally:
            conn.close()
        recompute(cfg, root)
        return {"status": "ok", "count": n, "built_at": _state["built_at"]}

    @app.get("/api/budget_depts")
    def api_budget_depts(request: Request):
        _require(request)
        conn = _conn()
        try:
            return db.list_budget_depts(conn)
        finally:
            conn.close()

    @app.get("/api/adjust_fields")
    def api_adjust_fields(request: Request):
        """R1：各明细表可调整字段（schema 黑名单制推导），管理员端字段下拉数据源。"""
        _require(request)
        return db.adjustable_fields()

    return app


def _screenshot_png(html: str, blk: str = "", width: int = 1440) -> bytes:
    """把用户页 HTML 在无头浏览器里渲开并整页截图。blk 非空=先切到该周期视图。
    reduced_motion 关掉全部动效（粒子/扫描线/生长动画），截出来是静止完整帧。"""
    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        br = p.chromium.launch(headless=True)
        try:
            ctx = br.new_context(viewport={"width": width, "height": 900},
                                 reduced_motion="reduce", device_scale_factor=2)
            pg = ctx.new_page()
            pg.set_content(html, wait_until="load")
            if blk:
                pg.evaluate(
                    "k=>{document.querySelectorAll('.pv').forEach(x=>{"
                    "x.style.display=x.getAttribute('data-blk')===k?'':'none';});"
                    "var b=document.getElementById('periodBtn');"
                    "if(b)b.childNodes[0].textContent=k+' ';}", blk)
            # 截图里去掉纯装饰/交互件（粒子层、导出与主题按钮）
            pg.add_style_tag(content=".particles,#exportBtn,#themeBtn{display:none!important}")
            pg.wait_for_timeout(400)
            return pg.screenshot(full_page=True, type="png")
        finally:
            br.close()


def serve(cfg=None, root=None):
    cfg = cfg or loaders.load_config()
    print("[server] 首次构建页面（跑管道+渲染）……")
    try:
        refresh(cfg, root)
        print(f"[server] 就绪 built_at={_state['built_at']}")
    except Exception as e:  # 数据有问题也让服务起来、页面提示
        print(f"[server] ⚠ 构建失败：{type(e).__name__}: {e}（服务仍启动，修数据后 /api/refresh 或重启）")
    app = create_app(cfg, root)
    import uvicorn
    host = cfg.get("server_host", "0.0.0.0")
    # 环境变量 KANBAN_PORT 可覆盖端口（本机多会话调试时避开 config 固定端口，不影响部署默认值）
    port = int(os.environ.get("KANBAN_PORT") or cfg.get("server_port", 8018))
    print(f"[server] 内网服务：用户端 http://<本机IP>:{port}/   管理员端 http://<本机IP>:{port}/admin")

    # 看门狗回滚配套：正常起服务 N 秒后清掉「更新回滚点」标记 = 确认这版没崩、无需回滚。
    # （若这版更新后启动即崩，进程活不到清标记，看门狗见标记仍在→自动回滚上一版本。）
    def _confirm_update_good():
        time.sleep(20)
        try:
            import updater
            updater.clear_rollback_marker(loaders.ROOT)
        except Exception:
            pass
    threading.Thread(target=_confirm_update_good, daemon=True).start()

    uvicorn.run(app, host=host, port=port, log_level="info")


if __name__ == "__main__":
    serve()
