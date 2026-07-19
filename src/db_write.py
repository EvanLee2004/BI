#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""存储写路径 / 管道用 SQL（任务书43·阶段二）：业务层零裸 SQL，SQL 只进 db.py + schema.py + 本模块。

SQLite 方言见 docs/softeng/08 换库须知；真换库时优先改本文件与 schema。
"""
from __future__ import annotations

import json
import sqlite3

import money
import schema

# 允许的 std 表（防 SQL 注入：动态表名白名单）
_STD = frozenset(schema.STD_TABLE_NAMES)

_STD_INSERT_COLS: dict[str, list[str]] = {
    "std_收入明细": [
        "定位键",
        "订单号",
        "客户",
        "业务线",
        "销售",
        "整单交付日期",
        "交付额",
        "项目成本",
        "归属月",
        "原值_交付日期",
        "原值_归属月",
    ],
    "std_下单": ["定位键", "订单号", "下单日期", "下单预估额", "部门", "销售", "客户", "归属月", "原值_归属月"],
    "std_回款": ["定位键", "回款ID", "到账日期", "到账金额", "客户", "销售", "归属月", "原值_归属月"],
    "std_内部译员": ["定位键", "任务ID", "任务提交日期", "结算金额", "译员类型", "译员姓名", "销售", "归属月", "原值_归属月"],
    "std_费用明细": [
        "定位键",
        "收单月份",
        "收单日期",
        "含税金额",
        "业务BU",
        "对应报表大类",
        "预算明细费用类型",
        "预算归属部门",
        "事项",
        "提单人",
        "提单人部门",
        "业务员",
        "配音费合同号",
        "归属月",
        "原值_归属月",
    ],
}

_STD_ORDER = ["std_收入明细", "std_下单", "std_回款", "std_内部译员", "std_费用明细"]


def insert_std_records(conn: sqlite3.Connection, table: str, records: list[dict]) -> None:
    if table not in _STD_INSERT_COLS:
        raise KeyError(table)
    cols = _STD_INSERT_COLS[table]
    sql = f"INSERT INTO {table}({','.join(cols)}) VALUES({','.join('?' * len(cols))})"
    rows = []
    for r in records:
        rf = money.record_amounts_to_fen(table, r)
        rows.append(tuple(rf.get(c) for c in cols))
    conn.executemany(sql, rows)


def rebuild_std_tables(conn: sqlite3.Connection, records: dict) -> None:
    """清表+插入单事务（BEGIN IMMEDIATE）。"""
    try:
        conn.commit()
    except Exception:
        pass
    prev_iso = conn.isolation_level
    conn.isolation_level = None
    try:
        conn.execute("BEGIN IMMEDIATE")
        schema.reset_std_tables(conn, commit=False)
        for t in _STD_ORDER:
            insert_std_records(conn, t, records.get(t) or [])
        conn.execute("COMMIT")
    except Exception:
        try:
            conn.execute("ROLLBACK")
        except Exception:
            pass
        raise
    finally:
        conn.isolation_level = prev_iso


def insert_run_log(conn: sqlite3.Connection, now: str, trigger: str, 结果: str, log_body: dict) -> None:
    conn.execute(
        "INSERT INTO meta_运行日志(时间,触发方式,结果,体检JSON) VALUES(?,?,?,?)",
        (now, trigger, 结果, json.dumps(log_body, ensure_ascii=False)),
    )
    conn.commit()


def prune_run_logs(conn: sqlite3.Connection, keep_days: int = 365) -> int:
    """删除超过 keep_days 的运行日志。返回删除行数。"""
    keep_days = max(1, int(keep_days))
    # SQLite：时间列为 'YYYY-MM-DD HH:MM:SS' 文本，用 date 比较
    cur = conn.execute(
        "DELETE FROM meta_运行日志 WHERE date(substr(时间,1,10)) < date('now', ?)",
        (f"-{keep_days} days",),
    )
    n = cur.rowcount if cur.rowcount is not None and cur.rowcount >= 0 else 0
    conn.commit()
    return n


def vacuum_db(conn: sqlite3.Connection) -> None:
    """全量 VACUUM（月末快照后调用）。SQLite 方言。"""
    # VACUUM 不可在事务中
    try:
        conn.commit()
    except Exception:
        pass
    conn.execute("VACUUM")


def db_file_size_bytes(cfg: dict, root=None) -> int | None:
    import loaders

    p = loaders.data_dir(cfg, root) / cfg.get("db_path", "看板.db")
    try:
        return p.stat().st_size if p.is_file() else None
    except OSError:
        return None


def disk_free_ratio(path) -> float | None:
    """返回 path 所在盘剩余比例 0~1；失败 None。"""
    import shutil
    from pathlib import Path

    try:
        u = shutil.disk_usage(str(Path(path)))
        if u.total <= 0:
            return None
        return u.free / u.total
    except OSError:
        return None


def manual_row_count(conn: sqlite3.Connection) -> int:
    return int(conn.execute("SELECT COUNT(*) FROM manual_手填").fetchone()[0])


def upsert_manual_row(conn: sqlite3.Connection, 归属月, 项目, 金额, 填写时间, 经手人) -> None:
    conn.execute(
        "INSERT OR REPLACE INTO manual_手填(归属月,项目,金额,填写时间,经手人) VALUES(?,?,?,?,?)",
        (归属月, 项目, 金额, 填写时间, 经手人),
    )


# ---------- 调整重放 SQL ----------
def list_active_adjustments(conn: sqlite3.Connection):
    return conn.execute(
        "SELECT id,目标表,定位键,字段,原值,新值,类型 FROM adj_调整记录 WHERE 状态='生效' ORDER BY id"
    ).fetchall()


def count_locator_matches(conn: sqlite3.Connection, table: str, 定位键: str) -> list:
    if table not in _STD:
        return []
    return conn.execute(f"SELECT id FROM {table} WHERE 定位键=? AND 已删除=0", (定位键,)).fetchall()


def mark_adjustment_expired(conn: sqlite3.Connection, aid: int) -> None:
    conn.execute("UPDATE adj_调整记录 SET 状态='过期疑似' WHERE id=?", (aid,))


def soft_delete_by_locator(conn: sqlite3.Connection, table: str, 定位键: str) -> int:
    if table not in _STD:
        return 0
    cur = conn.execute(f"UPDATE {table} SET 已删除=1 WHERE 定位键=? AND 已删除=0", (定位键,))
    return cur.rowcount


def select_field_by_locator(conn: sqlite3.Connection, table: str, 字段: str, 定位键: str):
    if table not in _STD:
        return None
    # 字段白名单
    allowed = set(schema.ADJUSTABLE_FIELDS.get(table) or ()) | {"归属月", "收单日期", "收单月份"}
    if 字段 not in allowed:
        raise KeyError(字段)
    row = conn.execute(f"SELECT {字段} FROM {table} WHERE 定位键=? AND 已删除=0", (定位键,)).fetchone()
    return row[0] if row else None


def update_field_by_locator(conn: sqlite3.Connection, table: str, 字段: str, value, 定位键: str) -> None:
    if table not in _STD:
        return
    allowed = set(schema.ADJUSTABLE_FIELDS.get(table) or ()) | {"归属月"}
    if 字段 not in allowed:
        raise KeyError(字段)
    conn.execute(f"UPDATE {table} SET {字段}=? WHERE 定位键=? AND 已删除=0", (value, 定位键))


def select_ledger_date_parts(conn: sqlite3.Connection, 定位键: str):
    return conn.execute(
        "SELECT 收单日期,收单月份 FROM std_费用明细 WHERE 定位键=? AND 已删除=0", (定位键,)
    ).fetchone()


# ---------- 审计流水导出归档（不删，只导出；任务书43·阶段三）----------
_ARCHIVE_TABLES = {
    "manual_历史": ("时间", "经手人", "归属月", "项目", "旧值", "新值"),
    "manual_预算历史": ("时间", "经手人", "年份", "指标", "范围", "旧值", "新值"),
    "manual_配置变更": ("时间", "操作账号", "类别", "摘要"),
}


def export_audit_archive_xlsx(conn: sqlite3.Connection, year: str | int) -> bytes:
    """导出指定年的手填历史/预算历史/配置变更到一个 xlsx（多 sheet）。不删除库内行。"""
    import io
    import openpyxl

    y = str(year).strip()
    if not (y.isdigit() and len(y) == 4):
        raise ValueError("year 须为 4 位年份")
    wb = openpyxl.Workbook()
    first = True
    for table, cols in _ARCHIVE_TABLES.items():
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        ws.title = table.replace("manual_", "")[:31]
        ws.append(list(cols))
        coln = ",".join(cols)
        # 时间列以 YYYY 开头过滤该年
        rows = conn.execute(
            f"SELECT {coln} FROM {table} WHERE 时间 LIKE ? ORDER BY id",
            (f"{y}-%",),
        ).fetchall()
        for r in rows:
            ws.append([("" if v is None else v) for v in r])
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
