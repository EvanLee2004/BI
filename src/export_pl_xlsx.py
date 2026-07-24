#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""管理利润表 Excel 导出（2.4.1：金额=元；2.4.0 一页化版式）。

纯函数：从既有 summary + period_key 组装 xlsx bytes。
数据源用 domain.pl.structure.pl_structure 全量（含 impact 分 / is_pct），
金额列 = money.fen_to_yuan(impact) 数字元（#,##0.00）；百分比行写展示串。
禁止写页面 amt_disp 万元串；禁止再算税前/毛利。

一张 sheet：抬头块 + 主表大类行（加粗）+ 各大类 details 内嵌明细
（缩进 + 浅灰 + 字号小一号）。
"""

from __future__ import annotations

import io
import re
import time
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# 文件名非法字符
_FNAME_ILLEGAL = re.compile(r'[\\/:*?"<>|\s]+')

_BOLD = Font(bold=True, size=11)
_NORMAL = Font(bold=False, size=11)
_DETAIL = Font(bold=False, size=10, color="666666")
_META_LABEL = Font(bold=True, size=10)
_META_VALUE = Font(bold=False, size=10)
_HEADER = Font(bold=True, size=11)
_DETAIL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

# Excel 金额数字格式（元）
_YUAN_NUM_FMT = "#,##0.00"

_CALIBER = "与看板管理利润表当前筛选一致；金额单位：元（页面看板仍为万元展示）"


def safe_filename_part(s: str) -> str:
    """文件名安全片段：非法字符替换为 _。"""
    part = _FNAME_ILLEGAL.sub("_", str(s or "").strip())
    part = part.strip("._") or "x"
    return part[:80]


def _set_col_widths(ws: Worksheet, headers: list[str], *, extra: int = 8, cap: int = 36) -> None:
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(str(col)) + extra, 12), cap)


def _write_header_block(
    ws: Worksheet,
    *,
    scope_label: str,
    period_key: str,
    version: str,
    is_bu: bool,
    export_time: str | None,
) -> int:
    """写主表顶部抬头块（产品/VERSION/范围/周期/导出时间），返回下一写入行号（1-based）。"""
    when = export_time or time.strftime("%Y-%m-%d %H:%M:%S")
    info_rows = [
        ("产品", "甲骨易经营看板"),
        ("VERSION", str(version or "")),
        ("范围", str(scope_label or ("整体" if not is_bu else ""))),
        ("周期", str(period_key)),
        ("导出时间", when),
        ("口径", _CALIBER),
    ]
    row = 1
    for k, v in info_rows:
        ws.cell(row=row, column=1, value=k).font = _META_LABEL
        ws.cell(row=row, column=2, value=v).font = _META_VALUE
        row += 1
    # 空行分隔抬头与表头
    row += 1
    return row


def _amount_cell_value(item: dict[str, Any]) -> Any:
    """金额列写入值：百分比 → 展示串；否则 impact 分 → 元 float。"""
    import money

    if item.get("is_pct"):
        return item.get("amt_disp") if item.get("amt_disp") is not None else ""
    impact = item.get("impact")
    if impact is None:
        return ""
    return money.fen_to_yuan(impact)


def _write_amount_cell(cell, item: dict[str, Any], *, font: Font, fill: PatternFill | None = None) -> None:
    """写金额列：元数字 + #,##0.00；百分比写字符串。"""
    val = _amount_cell_value(item)
    cell.value = val
    cell.font = font
    if fill is not None:
        cell.fill = fill
    if not item.get("is_pct") and isinstance(val, (int, float)):
        cell.number_format = _YUAN_NUM_FMT


def _write_single_sheet(
    ws: Worksheet,
    rows: list[dict[str, Any]],
    details: dict[str, Any],
    *,
    scope_label: str,
    period_key: str,
    version: str,
    is_bu: bool,
    export_time: str | None,
) -> None:
    """一张 sheet：抬头 + 表头 + 大类（加粗）+ 内嵌明细（缩进浅灰）。"""
    start = _write_header_block(
        ws,
        scope_label=scope_label,
        period_key=period_key,
        version=version,
        is_bu=is_bu,
        export_time=export_time,
    )
    headers = ["科目", "金额", "说明"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=col, value=h)
        cell.font = _HEADER
    row_idx = start + 1

    for r in rows:
        name = r.get("name") or ""
        formula = r.get("formula") or ""
        ws.cell(row=row_idx, column=1, value=name).font = _BOLD
        _write_amount_cell(ws.cell(row=row_idx, column=2), r, font=_BOLD)
        ws.cell(row=row_idx, column=3, value=formula).font = _BOLD
        row_idx += 1

        open_key = r.get("open_key")
        if not open_key:
            continue
        block = details.get(str(open_key))
        if not isinstance(block, dict):
            continue
        for ln in block.get("lines") or []:
            if not isinstance(ln, dict):
                continue
            ln_name = ln.get("name") or ""
            # sub 行更深缩进（二级明细）
            indent = 2 if ln.get("sub") else 1
            c1 = ws.cell(row=row_idx, column=1, value=ln_name)
            c1.font = _DETAIL
            c1.alignment = Alignment(indent=indent)
            c1.fill = _DETAIL_FILL
            # 明细行无 is_pct；用 impact 写元
            _write_amount_cell(ws.cell(row=row_idx, column=2), ln, font=_DETAIL, fill=_DETAIL_FILL)
            c3 = ws.cell(row=row_idx, column=3, value="")
            c3.font = _DETAIL
            c3.fill = _DETAIL_FILL
            row_idx += 1

    _set_col_widths(ws, headers, extra=12, cap=40)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 36


def _struct_for_period(summary: dict, period_key: str, *, is_bu: bool) -> dict[str, Any]:
    """与 pack_pl_by_period 同入参组装，返回裁剪前 pl_structure（含 impact 分）。"""
    from domain.pl.structure import pl_structure

    meta = summary.get("meta") or {}
    P = summary.get("periods") or {}
    FT = summary.get("expense_fine_type") or {}
    yk = meta.get("year_key") or ""
    unc = (meta.get("unclassified") or {}).get("expense") or {}
    unc_amt = float(unc.get("amount") or 0) if unc else 0.0
    alloc = meta.get("public_allocation") or {"enabled": False}

    if period_key not in P:
        raise KeyError(period_key)
    p = P[period_key]
    if not isinstance(p, dict):
        raise KeyError(period_key)

    unc_use = unc_amt if ((not is_bu) and unc_amt > 0 and period_key == yk) else None
    return pl_structure(
        p,
        FT.get(period_key) or {},
        is_bu=is_bu,
        unclassified_amt=unc_use,
        alloc_meta=alloc if is_bu else None,
    )


def build_pl_xlsx_bytes(
    summary: dict,
    *,
    period_key: str,
    is_bu: bool,
    scope_label: str,
    version: str = "",
    export_time: str | None = None,
) -> bytes:
    """组装管理利润表 xlsx（单 sheet；金额=元）。

    Parameters
    ----------
    summary : dict
        整体或该 BU 的 summary（与页面 VM 同源）。
    period_key : str
        周期键（与 store.period / ?blk= 一致）。
    is_bu : bool
        是否按 BU 利润表结构（费用抽屉等）。
    scope_label : str
        「整体」或 BU 名，写入导出说明与文件名语义。
    version : str
        产品 VERSION，可空。
    export_time : str | None
        导出时间串；默认当前本地时间。

    Raises
    ------
    KeyError
        period_key 不在 summary.periods 中（路由层应转 400）。
    """
    struct = _struct_for_period(summary, period_key, is_bu=is_bu)
    rows = list(struct.get("rows") or [])
    details = dict(struct.get("details") or {})

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "管理利润表"
    _write_single_sheet(
        ws,
        rows,
        details,
        scope_label=scope_label,
        period_key=period_key,
        version=version,
        is_bu=is_bu,
        export_time=export_time,
    )

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def pl_xlsx_filename(*, scope_label: str, period_key: str, day: str | None = None) -> str:
    """RFC 文件名：管理利润表_{范围}_{period}_{YYYYMMDD}.xlsx"""
    d = day or time.strftime("%Y%m%d")
    scope = safe_filename_part(scope_label or "整体")
    period = safe_filename_part(period_key or "")
    return f"管理利润表_{scope}_{period}_{d}.xlsx"
