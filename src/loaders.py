#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""读取 6 个数据源（项目明细 / 收单台账 / 下单 / 回款 / 内部译员 / 手填与调整）+ 通用解析工具。

设计要点（沿用 v1.3 已验证做法，并按真实数据形态扩展）：
- 金额既可能是数字也可能是文本（陆总手动导出的是 '560.00' 文本），parse_amount 两种都吃。
- Excel 一律用完整加载模式，**绝不用 read_only=True**（智云导出的 xlsx `<dimension>` 标签谎报只有1格，
  流式读会静默只读1行，详见 docs/数据来源说明）。
- 列全部按表头文字定位，不认死列号。
- data_dir + period_pin 由 config 决定，测试/正式一键切换。
"""

from __future__ import annotations

import csv
import datetime
import json
import logging
from pathlib import Path
from typing import Any

import openpyxl

log = logging.getLogger("kanban.loaders")

ROOT = Path(__file__).resolve().parents[1]  # 程序根目录（config.json 所在层）

# 机器本地配置覆盖文件（放 data_dir 内，gitignore）：每台机不同 / 会在管理端改的设置
# （收单台账共享盘路径、更新时间、备份天数等）只落这里，**config.json 保持出厂默认永不被程序写脏**
# → git 工作区干净 → 一键更新的"工作区脏就拒绝"护栏不会被误触发（部署机才能用一键更新）。
LOCAL_CONFIG_NAME = "本地配置.json"

# 2.6.3·A2：本地配置允许覆盖的键白名单（路径类/环境类不可被覆盖，防影子库与环境串）
LOCAL_CONFIG_DENY_KEYS = frozenset({"data_dir", "db_path", "profiles"})
# 允许覆盖的运维/业务开关（其余未知键仍允许，便于扩展；仅拒绝上面危险键）
# 说明：白名单语义=「危险键拒绝」而非「仅允许列出键」——既防双拼又不大改现有本地配置面。

# 2.6.3·A4：本地配置损坏态（进程内；供 /api/v1/health 抬黄）
_LOCAL_CONFIG_CORRUPT: dict | None = None


# ---------------- 配置 / 时间 ----------------
def _local_config_path(base: Path, cfg: dict) -> Path:
    return base / cfg.get("data_dir", "数据") / LOCAL_CONFIG_NAME


# 启动/管道必需键（缺则清晰报错，避免深处 KeyError）
_REQUIRED_CONFIG_KEYS = (
    "data_dir",
    "files",
    "columns",
)


def validate_config(cfg: dict) -> None:
    """校验 config 形状；失败抛 ValueError（带键路径），不吞。"""
    if not isinstance(cfg, dict):
        raise ValueError("config 必须是 JSON 对象")
    for k in _REQUIRED_CONFIG_KEYS:
        if k not in cfg:
            raise ValueError(f"config.json 缺少必需键：{k}")
    if not isinstance(cfg["files"], dict):
        raise ValueError("config.files 必须是对象")
    for fk in ("project_detail_stem", "orders", "receipts", "inhouse", "ledger", "manual"):
        if fk not in cfg["files"]:
            raise ValueError(f"config.files 缺少：{fk}")
    if not isinstance(cfg["columns"], dict):
        raise ValueError("config.columns 必须是对象")
    for ck in (
        "project_delivery_date",
        "project_revenue",
        "project_cost",
        "project_line",  # 任务书64·D4：主线必需列（业务线映射）
        "order_date",
        "order_amount",
        "receipt_date",
        "receipt_amount",
        "inhouse_date",
        "inhouse_amount",
        "inhouse_type",
    ):
        if ck not in cfg["columns"]:
            raise ValueError(f"config.columns 缺少：{ck}")


def local_config_corrupt_status() -> dict | None:
    """2.6.3·A4：最近一次本地配置损坏信息（无则 None）。"""
    return dict(_LOCAL_CONFIG_CORRUPT) if _LOCAL_CONFIG_CORRUPT else None


def clear_local_config_corrupt_status() -> None:
    global _LOCAL_CONFIG_CORRUPT
    _LOCAL_CONFIG_CORRUPT = None


def _apply_local_overrides(cfg: dict, data: dict) -> None:
    """合并本地覆盖：拒绝 data_dir/db_path/profiles（2.6.3·A2）。"""
    for k, v in data.items():
        if v is None:
            continue
        if k in LOCAL_CONFIG_DENY_KEYS or k.startswith("_"):
            if k in LOCAL_CONFIG_DENY_KEYS:
                log.warning("本地配置忽略危险键 %s（不可覆盖）", k)
            continue
        cfg[k] = v


def _mark_local_config_corrupt(path: Path, reason: str, cfg: dict | None = None) -> None:
    """坏本地配置 → 进程旗标 + warning + 告警（2.6.3·A4，不许 except: pass 静默）。"""
    global _LOCAL_CONFIG_CORRUPT
    _LOCAL_CONFIG_CORRUPT = {"path": str(path), "reason": reason}
    log.warning("本地配置.json 损坏，退回 config.json 默认：%s (%s)", path, reason)
    try:
        import notify

        notify.maybe_alert_text(
            cfg or {},
            f"【经营看板告警】本地配置.json 损坏：{reason}；已退回出厂默认（含 ledger_share_path 等）。"
            f"请在管理端重存设置。路径：{path}",
        )
    except Exception:
        pass


def _apply_profile(cfg: dict) -> None:
    """2.6.3·D4：读环境变量 KANBAN_PROFILE=dev|staging|prod，套 config.profiles 覆盖。

    仅合并 profiles[name] 内非 None 键；不覆盖 profiles 自身。未设 env 或名未知 → 跳过。
    """
    import os

    name = (os.environ.get("KANBAN_PROFILE") or "").strip()
    if not name:
        return
    profiles = cfg.get("profiles")
    if not isinstance(profiles, dict):
        log.warning("KANBAN_PROFILE=%s 但 config.profiles 不是对象，忽略", name)
        return
    prof = profiles.get(name)
    if not isinstance(prof, dict):
        log.warning("KANBAN_PROFILE=%s 不在 config.profiles 中（可选：%s）", name, list(profiles.keys()))
        return
    for k, v in prof.items():
        if v is None or k in ("profiles",) or str(k).startswith("_"):
            continue
        cfg[k] = v
    cfg["_active_profile"] = name


def load_config(root: Path | None = None, *, strict: bool = True) -> dict:
    """读 config.json（出厂默认），套 KANBAN_PROFILE，再叠加机器本地覆盖（data_dir/本地配置.json，若有）。
    覆盖只认非 None 值；危险键 data_dir/db_path/profiles 不可被本地配置覆盖。
    坏文件：体检黄 + log.warning + 告警，退回 config.json 默认（2.6.3·A4）。
    config.json 本身只读不写。
    strict=True（默认）：校验必需键，缺则 ValueError。
    strict=False：仅读盘（updater 读 pip_mirror 等，不要求完整 schema）。
    """
    global _LOCAL_CONFIG_CORRUPT
    base = root or ROOT
    path = base / "config.json"
    try:
        with open(path, encoding="utf-8") as f:
            cfg = json.load(f)
    except FileNotFoundError as e:
        raise ValueError(f"找不到 config.json：{path}") from e
    except json.JSONDecodeError as e:
        raise ValueError(f"config.json JSON 无效：{e}") from e
    if not isinstance(cfg, dict):
        raise ValueError("config.json 必须是 JSON 对象")
    # 2.6.3·D4：环境 profile 先于本地配置（本地不可改 data_dir）
    _apply_profile(cfg)
    ov = _local_config_path(base, cfg)
    if ov.exists():
        try:
            data = json.loads(ov.read_text(encoding="utf-8"))
            if not isinstance(data, dict):
                _mark_local_config_corrupt(ov, "根节点不是 JSON 对象", cfg)
            else:
                # 飞书 webhook 功能已删：读盘时丢弃残留键，不进运行时 cfg
                data.pop("feishu_webhook_url", None)
                _apply_local_overrides(cfg, data)
                cfg.pop("feishu_webhook_url", None)
                _LOCAL_CONFIG_CORRUPT = None
        except (OSError, ValueError) as e:
            _mark_local_config_corrupt(ov, f"{type(e).__name__}: {e}", cfg)
    cfg.pop("feishu_webhook_url", None)
    if strict:
        validate_config(cfg)
    return cfg


def local_config_path(cfg: dict, root: Path | None = None) -> Path:
    return _local_config_path(root or ROOT, cfg)


def read_local_config(cfg: dict, root: Path | None = None) -> dict:
    """读机器本地覆盖文件为 dict（缺 → {}；坏 → {} 并标记损坏）。"""
    p = local_config_path(cfg, root)
    if not p.exists():
        return {}
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
        if not isinstance(d, dict):
            _mark_local_config_corrupt(p, "根节点不是 JSON 对象", cfg)
            return {}
        # 过滤危险键，调用方拿到的也是安全视图
        return {k: v for k, v in d.items() if k not in LOCAL_CONFIG_DENY_KEYS}
    except (OSError, ValueError) as e:
        _mark_local_config_corrupt(p, f"{type(e).__name__}: {e}", cfg)
        return {}


def write_local_config(cfg: dict, root: Path | None = None, updates: dict | None = None) -> dict:
    """把 updates 合并进机器本地覆盖文件并落盘；返回合并后的全量覆盖 dict。
    **只写这个 gitignore 文件，绝不动 config.json**（保持 git 工作区干净→一键更新可用）。
    拒绝写入 data_dir/db_path/profiles。
    """
    global _LOCAL_CONFIG_CORRUPT
    cur = read_local_config(cfg, root)
    for k, v in (updates or {}).items():
        if k in LOCAL_CONFIG_DENY_KEYS:
            log.warning("write_local_config 拒绝危险键 %s", k)
            continue
        # 飞书 webhook 功能已删：禁止再写入该键
        if k == "feishu_webhook_url":
            cur.pop("feishu_webhook_url", None)
            continue
        cur[k] = v
    cur.pop("feishu_webhook_url", None)
    p = local_config_path(cfg, root)
    p.parent.mkdir(parents=True, exist_ok=True)
    # 2.6.7 D-4：临时文件 + os.replace 原子落盘
    import os
    import tempfile

    text = json.dumps(cur, ensure_ascii=False, indent=2) + "\n"
    fd, tmp = tempfile.mkstemp(prefix=p.name + ".", suffix=".tmp", dir=str(p.parent))
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(text)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp, p)
    except Exception:
        try:
            os.unlink(tmp)
        except OSError:
            pass
        raise
    _LOCAL_CONFIG_CORRUPT = None
    return cur


def pinned_today(cfg: dict) -> datetime.date:
    """当前年月：config.period_pin 钉住则用它（测试=2024-07），否则用系统当天（正式版数据是当月的）。"""
    pin = cfg.get("period_pin")
    if pin and pin.get("year") and pin.get("month"):
        import calendar

        last = calendar.monthrange(pin["year"], pin["month"])[1]
        return datetime.date(pin["year"], pin["month"], min(datetime.date.today().day, last))
    return datetime.date.today()


def data_dir(cfg: dict, root: Path | None = None) -> Path:
    return (root or ROOT) / cfg["data_dir"]


# ---------------- 通用解析 ----------------
def parse_amount(val: Any) -> float:
    if val is None:
        return 0.0
    s = str(val).replace(",", "").strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def amount_parse_fails(val: Any) -> bool:
    """单元格非空、但 parse_amount 只能按 0 算的情形——供数据体检计数，别让坏值无声消失。"""
    if val is None:
        return False
    s = str(val).replace(",", "").strip()
    if not s:
        return False
    try:
        float(s)
        return False
    except ValueError:
        return True


_DATE_PARTS_CACHE: dict = {}


def parse_date_parts(val: Any) -> tuple[int, int, int] | None:
    """解析日期为 (年,月,日)，带结果缓存（周期矩阵含月区间后同一值会被解析几十次）。"""
    try:
        return _DATE_PARTS_CACHE[val]
    except KeyError:
        r = _parse_date_parts(val)
        _DATE_PARTS_CACHE[val] = r
        return r
    except TypeError:  # 不可哈希的怪值：不缓存直接算
        return _parse_date_parts(val)


def _valid_ymd(y: int, m: int, d: int) -> tuple[int, int, int] | None:
    """日历合法性：用 datetime.date 校验（含闰年/月长），非法日（如 2/30）返回 None 交给体检黄，不造假日期。"""
    try:
        datetime.date(int(y), int(m), int(d))
    except (TypeError, ValueError):
        return None
    return int(y), int(m), int(d)


def _ymd_from_digits_or_norm(s: str) -> tuple[int, int, int] | None:
    """从纯数字串或分隔日期串解析 (年,月,日)。"""
    digits = "".join(c for c in s if c.isdigit())
    if len(digits) >= 8:
        try:
            return _valid_ymd(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))
        except ValueError:
            return None
    norm = s.replace("/", "-").split("-")
    try:
        if len(norm) >= 3:
            return _valid_ymd(int(norm[0]), int(norm[1]), int(norm[2][:2]))
        if len(norm) >= 2:
            return _valid_ymd(int(norm[0]), int(norm[1]), 1)
    except ValueError:
        return None
    return None


def _parse_date_parts(val: Any) -> tuple[int, int, int] | None:
    """解析日期为 (年,月,日)。支持 datetime、YYYY-MM-DD、YYYY/MM/DD、YYYYMMDD。"""
    if val is None:
        return None
    if hasattr(val, "year") and hasattr(val, "month"):
        try:
            return int(val.year), int(val.month), int(getattr(val, "day", 1) or 1)
        except (TypeError, ValueError):
            return None
    s = str(val).strip()
    if not s:
        return None
    return _ymd_from_digits_or_norm(s)


def _header_index(header: list[str], path: Path, required: tuple[str, ...]) -> dict[str, int]:
    """表头 → 列号。重名列：是我们要读的列 → 直接报错（不能猜哪列对）；不用的列 → 保留第一处出现。
    智云导出真实出现过重名列（内部译员表有两个"PM"），所以不能一刀切全报错。"""
    dups = {h for h in header if h and header.count(h) > 1}
    bad = sorted(dups & set(required))
    if bad:
        raise ValueError(f"「{path.name}」必需列出现重名：{bad}\n无法确定读哪一列，请先在源文件里改名去重再导入。")
    idx: dict[str, int] = {}
    for i, h in enumerate(header):
        if h and h not in idx:
            idx[h] = i
    return idx


def _rows_as_dicts(path: Path, required: tuple[str, ...] = ()) -> list[dict[str, str]]:
    """Excel/CSV → list[dict]，键=表头文字，值=字符串。Excel 读激活的第一个 sheet（智云导出只有一份数据）。"""
    if path.suffix.lower() in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(path, data_only=True)  # 完整加载，绝不 read_only
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(it)]
        idx = _header_index(header, path, required)
        out: list[dict[str, str]] = []
        for row in it:
            if all(v is None for v in row):
                continue
            out.append({h: ("" if row[i] is None else str(row[i])) for h, i in idx.items() if i < len(row)})
        return out
    with open(path, encoding="utf-8-sig") as f:
        rdr = csv.reader(f)
        header = [str(h).strip() for h in next(rdr, [])]
        idx = _header_index(header, path, required)
        return [{h: (row[i] if i < len(row) else "") for h, i in idx.items()} for row in rdr if any(row)]


# ---------------- 各源 ----------------
def resolve_project_detail_path(cfg: dict, root: Path | None = None) -> Path:
    base = data_dir(cfg, root)
    stem = cfg["files"]["project_detail_stem"]
    for ext in (".xlsx", ".csv"):
        p = base / f"{stem}{ext}"
        if p.exists():
            return p
    raise FileNotFoundError(f"未找到项目明细：{base}/{stem}.xlsx 或 .csv（导出步骤见 docs/取数操作手册）。")


def _require_file(path: Path, name: str) -> Path:
    if not path.exists():
        raise FileNotFoundError(
            f"未找到「{name}」：{path}\n请把该文件放进数据目录（文件名固定，来源见 数据/README.md）。"
        )
    return path


def _load_checked(path: Path, name: str, required: list[str]) -> list[dict[str, str]]:
    """读表并**校验必需列都在**——列被改名/导错文件时立刻报错，绝不静默读成0算出错数字。"""
    rows = _rows_as_dicts(path, tuple(required))
    if rows:
        have = set(rows[0].keys())
        missing = [c for c in required if c not in have]
        if missing:
            raise ValueError(
                f"「{name}」缺少必需列：{missing}\n实际列：{sorted(have)}\n"
                f"可能是导出格式变了或导错了文件——请核对来源（数据/README.md），或在 config.json 的 columns 里更新列名。"
            )
    return rows


def load_project_detail(cfg: dict, root: Path | None = None) -> list[dict[str, str]]:
    c = cfg["columns"]
    return _load_checked(
        resolve_project_detail_path(cfg, root),
        "项目明细",
        [c["project_delivery_date"], c["project_revenue"], c["project_cost"]],
    )


def load_orders(cfg: dict, root: Path | None = None) -> list[dict[str, str]]:
    c = cfg["columns"]
    return _load_checked(
        _require_file(data_dir(cfg, root) / cfg["files"]["orders"], cfg["files"]["orders"]),
        "下单",
        [c["order_amount"], c["order_date"]],
    )


def load_receipts(cfg: dict, root: Path | None = None) -> list[dict[str, str]]:
    c = cfg["columns"]
    return _load_checked(
        _require_file(data_dir(cfg, root) / cfg["files"]["receipts"], cfg["files"]["receipts"]),
        "回款记录",
        [c["receipt_amount"], c["receipt_date"]],
    )


def load_inhouse(cfg: dict, root: Path | None = None) -> list[dict[str, str]]:
    c = cfg["columns"]
    return _load_checked(
        _require_file(data_dir(cfg, root) / cfg["files"]["inhouse"], cfg["files"]["inhouse"]),
        "内部译员",
        [c["inhouse_amount"], c["inhouse_date"], c["inhouse_type"]],
    )


# 收单台账：按年份 sheet + 表头文字定位列
# 2.6.3·B6：缺当年 sheet 不抛死整管；空集 + 旗标 + 告警
_LEDGER_SHEET_MISSING: dict | None = None


def ledger_sheet_missing_status() -> dict | None:
    return dict(_LEDGER_SHEET_MISSING) if _LEDGER_SHEET_MISSING else None


def clear_ledger_sheet_missing_status() -> None:
    global _LEDGER_SHEET_MISSING
    _LEDGER_SHEET_MISSING = None


class LedgerSheetMissing(Exception):
    """收单台账缺年页（可捕获；load_ledger 默认吞掉并返回空）。"""

    def __init__(self, year: str, existing: list[str] | None = None):
        self.year = str(year)
        self.existing = list(existing or [])
        super().__init__(
            f"收单台账缺 {self.year} 页（现有：{self.existing}）。找亮晶建。"
        )


def _open_ledger_sheet(path: Path, sheet_name: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise LedgerSheetMissing(sheet_name, list(wb.sheetnames))
    return wb[sheet_name]


def load_ledger(cfg: dict, sheet_name: str, root: Path | None = None) -> tuple[list, list[tuple]]:
    """返回 (表头行, 数据行)。

    2.6.3·B6：缺当年 sheet → 空集 + 体检红旗标 + 告警，不抛死整条管道。
    文件本身不存在仍 FileNotFoundError（与旧行为一致；fetch 层会先处理）。
    """
    global _LEDGER_SHEET_MISSING
    path = data_dir(cfg, root) / cfg["files"]["ledger"]
    if not path.exists():
        raise FileNotFoundError(f"未找到收单台账：{path}")
    try:
        ws = _open_ledger_sheet(path, sheet_name)
    except LedgerSheetMissing as e:
        _LEDGER_SHEET_MISSING = {
            "year": e.year,
            "existing": e.existing,
            "path": str(path),
            "banner": f"收单台账缺 {e.year} 页，找亮晶建",
        }
        log.error("收单台账缺 %s 页（现有 %s）；台账走空集，管道继续", e.year, e.existing)
        try:
            import notify

            notify.maybe_alert_text(
                cfg,
                f"【经营看板告警】收单台账缺 {e.year} 页，找亮晶建。现有 sheet：{e.existing}",
            )
        except Exception:
            pass
        return [], []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    _LEDGER_SHEET_MISSING = None
    return list(rows[0]), rows[1:]


def _manual_header_cell(h) -> str:
    """月份列名归一：datetime / 2026-1 / 2026/1 → YYYY-MM。"""
    import re

    if h is None:
        return ""
    if hasattr(h, "year") and hasattr(h, "month"):
        return f"{int(h.year):04d}-{int(h.month):02d}"
    s = str(h).strip()
    m = re.fullmatch(r"(\d{4})[-/](\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    return s


def load_manual(cfg: dict, root: Path | None = None) -> dict[str, dict[str, float]]:
    """手填与调整表（宽表：项目=行、月份=列）→ {月份'YYYY-MM': {项目: 金额float}}。
    表头形如 [项目, 归属, 备注, 2026-01, 2026-02, ...]；某项某月留空=不写入（留给"默认上月/0"逻辑）。
    维护友好：每月只在最右加一列，11 个项目始终整列可见、不会漏填。"""
    import re

    path = data_dir(cfg, root) / cfg["files"]["manual"]
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["手填与调整"] if "手填与调整" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [_manual_header_cell(h) for h in rows[0]]
    month_cols = {i: h for i, h in enumerate(header) if re.fullmatch(r"\d{4}-\d{2}", h)}
    try:
        i_item = header.index("项目")
    except ValueError:
        return {}
    out: dict[str, dict[str, float]] = {}
    for r in rows[1:]:
        item = str(r[i_item]).strip() if i_item < len(r) and r[i_item] is not None else ""
        if not item:
            continue
        for i, month in month_cols.items():
            v = r[i] if i < len(r) else None
            if v is None or str(v).strip() == "":
                continue
            out.setdefault(month, {})[item] = parse_amount(v)
    return out

