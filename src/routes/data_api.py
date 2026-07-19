"""明细/日查/排名/体检/刷新 — 从 server.create_app 纯搬家。"""

from __future__ import annotations

import time
from urllib.parse import quote

from fastapi import Body, HTTPException, Query, Request
from fastapi.responses import JSONResponse, Response

import charts
import db
from app_state import _state


def _parse_daily_range(start: str, end: str):
    """校验 ISO 日期区间；返回 (date_start, date_end)。"""
    import datetime as _dt

    try:
        s = _dt.date.fromisoformat(start)
        e = _dt.date.fromisoformat(end)
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="日期格式须为 YYYY-MM-DD") from None
    if e < s:
        raise HTTPException(status_code=400, detail="结束日期须不早于开始日期")
    if (e - s).days > 366:
        raise HTTPException(status_code=400, detail="区间最长 366 天")
    return s, e


def _daily_wan(v) -> str:
    return ("−" if v < 0 else "") + charts.fmt_wan(abs(v)) + "万"


def _format_daily_disp(d: dict, top: int) -> dict:
    """金额→显示串 + dual_rankings（就地改 d 的 days/totals/rankings）。"""
    import render as _render

    for row in d["days"]:
        row["orders_disp"], row["receipts_disp"] = _daily_wan(row.pop("orders")), _daily_wan(row.pop("receipts"))
    t = d["totals"]
    t["orders_disp"], t["receipts_disp"] = _daily_wan(t.pop("orders")), _daily_wan(t.pop("receipts"))
    dual = _render.dual_rankings_from_daily(d["rankings"], top=min(top, 10))
    for rk in d["rankings"].values():
        for it in rk["items"]:
            it["disp"] = _daily_wan(it.pop("amount"))
        if rk.get("others"):
            rk["others"]["disp"] = _daily_wan(rk["others"].pop("amount"))
        if rk.get("unfilled"):
            rk["unfilled"]["disp"] = _daily_wan(rk["unfilled"].pop("amount"))
        rk.pop("total", None)
    return dual


def _load_sales_to_bu(cfg, root):
    sales_to_bu = None
    try:
        import bu as _bu

        bucfg = _bu.load_bu_config(cfg, root)
        if bucfg and bucfg.get("bus"):
            sales_to_bu = {}
            for b in bucfg["bus"]:
                for sal in b.get("销售") or []:
                    sales_to_bu.setdefault(str(sal).strip(), b["name"])
            if not sales_to_bu:
                sales_to_bu = None
    except Exception:
        sales_to_bu = None
    return sales_to_bu


def _bu_sales_set(bucfg: dict, name: str) -> set:
    sales = set()
    for b in bucfg.get("bus") or []:
        if b.get("name") == name:
            sales = {str(x).strip() for x in (b.get("销售") or []) if str(x).strip()}
            break
    return sales


def register(app, d):  # noqa: C901  # 纯路由/装配分发壳，复杂度在子 handler
    cfg = d.cfg
    root = d.root
    _user = d.user
    _vacct = d.vacct
    _vacc_row = d.vacc_row
    _can_view_main = d.can_view_main
    _can_view_bu = d.can_view_bu
    _bu_switcher_html = d.bu_switcher_html
    _set_vcookie = d.set_vcookie
    _set_acookie = d.set_acookie
    _main_shell = d.main_shell
    _bu_shell = d.bu_shell
    _view_login_file = d.view_login_file
    _admin_login_file = d.admin_login_file
    _admin_static_html = d.admin_static_html
    _bootstrap_page = d.bootstrap_page
    _manual_items_json = d.manual_items_json
    _html_doc = d.html_doc
    _file_html_doc = d.file_html_doc
    _audit = d.audit
    _diff_accounts = d.diff_accounts
    _diff_bu_config = d.diff_bu_config
    _run_reasons = d.run_reasons

    def start_refresh_async(cfg, root=None, trigger="manual"):
        import server as _srv

        return _srv.start_refresh_async(cfg, root, trigger)

    def recompute(cfg, root=None):
        import server as _srv

        return _srv.recompute(cfg, root)

    _screenshot_png = d.screenshot_png
    _HIDE_PW_STYLE = d.HIDE_PW_STYLE
    _WRAP_OPEN = d.WRAP_OPEN

    def _detail_access(request: Request, table: str, bu: str | None):
        """明细鉴权：任务书51·B4 统一走 authz.resolve_expense_view_access(force_whitelist=False)。"""
        import authz

        return authz.resolve_expense_view_access(
            _user(request),
            _vacc_row(request),
            bu,
            cfg=cfg,
            force_whitelist=False,
            table=table,
        )

    @app.get("/api/detail_export")
    def api_detail_export(
        request: Request,
        table: str = Query("收入明细"),
        month: str | None = None,
        q: str | None = None,
        year: str | None = None,
        bu: str | None = None,
        filters: str | None = None,
        month_from: str | None = None,
        month_to: str | None = None,
    ):
        """当前筛选结果导出 Excel（.xlsx；上限 5000 行）。
        表头+行与明细页一致（看端=白名单列，管理端=全列）；月份/搜索/列筛与页面相同。"""
        force_bu, hide_salary, audience = _detail_access(request, table, bu)
        # 导出：管理员会话走 _require 语义——非管理员也可导出自己有权看的费用明细
        if _user(request) is None and not (_vacc_row(request) and table == "费用明细"):
            raise HTTPException(status_code=401, detail="需要登录")
        import io
        import openpyxl

        who = _user(request) or _vacct(request) or "?"
        _audit(cfg, root, who, ("访问", f"导出：{table}" + (f" bu={force_bu}" if force_bu else "")))
        conn = _conn()
        try:
            d = db.query_detail(
                conn,
                table,
                month,
                q,
                page=1,
                page_size=5000,
                year=year,
                bu=force_bu,
                filters=filters,
                hide_salary=hide_salary,
                audience=audience,
                month_from=month_from,
                month_to=month_to,
            )
        except KeyError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        finally:
            conn.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        # sheet 名最多 31 字，去掉 Excel 非法字符
        safe = "".join(c for c in str(table) if c not in r"[]:*?/\\")[:31] or "明细"
        ws.title = safe
        cols = d["columns"]
        ws.append(list(cols))
        for r in d["rows"]:
            ws.append([r.get(c, "") if r.get(c, "") is not None else "" for c in cols])
        # 首行粗体 + 简单列宽
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for i, col in enumerate(cols, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max(len(str(col)) + 4, 10), 28)
        bio = io.BytesIO()
        wb.save(bio)
        raw = bio.getvalue()
        day = time.strftime("%Y%m%d")
        fname = f"{safe}_{day}.xlsx"
        # RFC 5987：中文文件名用 filename*
        cd = f"attachment; filename=\"export.xlsx\"; filename*=UTF-8''{quote(fname)}"
        return Response(
            content=raw,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": cd},
        )

    @app.get("/api/detail")
    def api_detail(
        request: Request,
        table: str = Query("收入明细"),
        month: str | None = None,
        q: str | None = None,
        page: int = 1,
        page_size: int = 50,
        unclassified: bool = False,
        unfilled_dept: bool = False,
        year: str | None = None,
        bu: str | None = None,
        filters: str | None = None,
        month_from: str | None = None,
        month_to: str | None = None,
    ):
        """明细查询。管理员：全表全列。看端：费用明细白名单列（任务书41·D）+ 月区间真筛（41·E）。
        A5：查看端会话仅可查「费用明细」且必须带本 BU 过滤（铁律隔离）。
        任务书37·B7：filters=JSON 列筛（后端 SQL 分页）；B8：整体账号默隐工资。"""
        force_bu, hide_salary, audience = _detail_access(request, table, bu)
        # 任务书46·1：看端明细查询写访问留痕（不记业务内容）
        if audience in ("view", "view_bu"):
            who = _vacct(request) or "?"
            _audit(
                cfg,
                root,
                who,
                ("访问", f"看端明细：{table}" + (f" bu={force_bu}" if force_bu else "")),
            )
        conn = db.connect(cfg, root)
        try:
            try:
                return JSONResponse(
                    db.query_detail(
                        conn,
                        table,
                        month,
                        q,
                        page,
                        page_size,
                        unclassified,
                        unfilled_dept,
                        year=year,
                        bu=force_bu,
                        filters=filters,
                        hide_salary=hide_salary,
                        audience=audience,
                        month_from=month_from,
                        month_to=month_to,
                    )
                )
            except KeyError as e:
                raise HTTPException(status_code=400, detail=str(e)) from e
        finally:
            conn.close()

    @app.get("/api/detail/values")
    def api_detail_values(
        request: Request,
        table: str = Query("收入明细"),
        column: str = Query(...),
        month: str | None = None,
        q: str | None = None,
        year: str | None = None,
        bu: str | None = None,
        filters: str | None = None,
        limit: int = 200,
        month_from: str | None = None,
        month_to: str | None = None,
    ):
        """文本列去重值（Excel 多选下拉）。与 list 同鉴权/同上下文筛。"""
        force_bu, hide_salary, audience = _detail_access(request, table, bu)
        conn = db.connect(cfg, root)
        try:
            try:
                return JSONResponse(
                    db.query_detail_distinct(
                        conn,
                        table,
                        column,
                        month=month,
                        q=q,
                        year=year,
                        bu=force_bu,
                        filters=filters,
                        hide_salary=hide_salary,
                        limit=limit,
                        audience=audience,
                        month_from=month_from,
                        month_to=month_to,
                    )
                )
            except KeyError as e:
                raise HTTPException(status_code=400, detail=str(e)) from e
        finally:
            conn.close()

    @app.get("/api/detail/meta")
    def api_detail_meta(request: Request, table: str = Query("收入明细")):
        """列名+类型（text/number/date），供表头筛选 UI。看端跟白名单。"""
        _force_bu, _hs, audience = _detail_access(request, table, None)
        try:
            return {"table": table, "columns": db.detail_columns_meta(table, audience=audience)}
        except KeyError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e

    @app.get("/api/daily")
    def api_daily(request: Request, start: str = Query(""), end: str = Query(""), top: int = Query(10)):
        """按天明细（用户端「明细」入口·迭代计划13批次B）：任意日期区间的逐日下单/回款 + 期内排名。
        v7.8 起要求整体页/管理员会话（全公司口径出口，BU 会话不给——否则 BU 链接持有者可绕过页面隔离）；
        **纯只读**、无任何写路径；金额显示串全部后端算好（铁律2）。入参严格校验：ISO日期、start<=end、区间≤366天。"""
        if not _can_view_main(request):
            raise HTTPException(status_code=401, detail="请先登录看板")
        import profit as _profit

        s, e = _parse_daily_range(start, end)
        top = max(1, min(2000, int(top)))  # 排名条数：默认前10，「其余点开看明细」传 2000 拿全量
        conn = db.connect(cfg, root)
        try:
            orders = db.load_orders(cfg, conn)
            receipts = db.load_receipts(cfg, conn)
        finally:
            conn.close()
        # 与全年预渲染一致：有销售→BU 映射则多算 orders_by_bu（看端时间段查询统一按 BU）
        sales_to_bu = _load_sales_to_bu(cfg, root)
        d = _profit.compute_daily(orders, receipts, cfg["columns"], s, e, top=top, sales_to_bu=sales_to_bu)
        dual = _format_daily_disp(d, top)
        return {"start": start, "end": end, "dual_rankings": dual, **d}

    @app.get("/api/bu_daily")
    def api_bu_daily(
        request: Request,
        bu: str = Query(""),
        start: str = Query(""),
        end: str = Query(""),
        top: int = Query(10),
    ):
        """任务书39·B/C：BU 页按时间段查询（本 BU 销售过滤，零跨界全公司 /api/daily）。
        会话须能看该 BU；返回与 /api/daily 同形 dual_rankings。"""
        name = (bu or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="缺少 bu")
        if not _can_view_bu(request, name):
            raise HTTPException(status_code=401, detail="无权查看该 BU")
        import bu as _bu
        import profit as _profit

        s, e = _parse_daily_range(start, end)
        top = max(1, min(2000, int(top)))
        bucfg = _bu.load_bu_config(cfg, root) or {"bus": []}
        sales = _bu_sales_set(bucfg, name)
        if not sales and name not in {b.get("name") for b in bucfg.get("bus") or []}:
            raise HTTPException(status_code=404, detail="未知 BU")
        conn = db.connect(cfg, root)
        try:
            orders = _profit.filter_rows_by_sales(db.load_orders(cfg, conn), sales)
            receipts = _profit.filter_rows_by_sales(db.load_receipts(cfg, conn), sales)
        finally:
            conn.close()
        d = _profit.compute_daily(orders, receipts, cfg["columns"], s, e, top=top, sales_to_bu=None)
        dual = _format_daily_disp(d, top)
        return {"start": start, "end": end, "bu": name, "dual_rankings": dual, **d}

    @app.get("/api/profit_ranking")
    def api_profit_ranking(
        request: Request, dim: str = Query(""), start: str = Query(""), end: str = Query(""), top: int = Query(5000)
    ):
        """板块③「收入与毛利结构」全量明细（「其余 N 个」点开）：确认口径 收入/毛利 按客户/销售。
        与 /api/daily 同为全公司口径出口——要整体页/管理员会话（BU 会话 401，防绕过页面隔离）；
        **纯只读**；金额/毛利率显示串全部后端算好（铁律2）。入参严格校验：dim∈{customer,sales}、ISO 日期、区间≤366天。"""
        if not _can_view_main(request):
            raise HTTPException(status_code=401, detail="请先登录看板")
        name_col = {"customer": "客户", "sales": "销售"}.get(dim)
        if not name_col:
            raise HTTPException(status_code=400, detail="dim 须为 customer 或 sales")
        import datetime as _dt

        try:
            s = _dt.date.fromisoformat(start)
            e = _dt.date.fromisoformat(end)
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="日期格式须为 YYYY-MM-DD") from None
        if e < s:
            raise HTTPException(status_code=400, detail="结束日期须不早于开始日期")
        if (e - s).days > 366:
            raise HTTPException(status_code=400, detail="区间最长 366 天")
        top = max(1, min(5000, int(top)))
        conn = db.connect(cfg, root)
        try:
            project = db.load_project_detail(cfg, conn)
        finally:
            conn.close()
        import profit as _profit

        vat = cfg["tax"]["vat_rate"]
        rk = _profit.compute_profit_ranking(project, name_col, cfg["columns"], s, e, vat, top=top)

        def _wan(v):
            return ("−" if v < 0 else "") + charts.fmt_wan(abs(v)) + "万"

        def _mg(it):
            # 陆总0714：改叫「系统成本率」；按销售的率先不显示（防"人力算不算"连锁追问）
            if dim == "sales":
                return ""
            cp = it.get("cost_pct")
            return f"系统成本率 {cp:.0f}%" if cp is not None else "系统成本率 —"

        items = [
            {"name": it["name"], "revenue_disp": _wan(it["revenue"]), "margin_disp": _mg(it)} for it in rk["items"]
        ]
        if rk.get("unfilled"):
            uf = rk["unfilled"]
            items.append(
                {"name": "（未填）", "revenue_disp": _wan(uf["revenue"]), "margin_disp": _mg(uf), "unfilled": True}
            )
        return {"dim": dim, "start": start, "end": end, "items": items}

    @app.get("/api/exceptions")
    def api_exceptions(request: Request):
        """异常处理「总览」计数（管理员）。体检黄红是运行信号，留在 /api/health，不在这。"""
        _require(request)  # 同函数作用域下文定义，调用时已存在
        conn = db.connect(cfg, root)
        try:
            return db.exceptions_summary(conn)
        finally:
            conn.close()

    @app.get("/api/order_depts")
    def api_order_depts(request: Request):
        """下单表已出现过的部门清单（「下单未填部门」归类下拉用）。"""
        _require(request)
        conn = db.connect(cfg, root)
        try:
            return db.list_order_depts(conn)
        finally:
            conn.close()


    @app.post("/api/v1/client-error")
    def api_client_error(request: Request, payload: dict = Body(default={})):
        """B-5：前端全局错误只写日志。公开可写但限流/截断/无数据读。鉴权豁免：仅错误文本。"""
        import frontend_errors

        if not isinstance(payload, dict):
            payload = {}
        # 可选附带 UA（截断）
        ua = (request.headers.get("user-agent") or "")[:120]
        payload = {**payload, "ua": ua}
        return frontend_errors.record_frontend_error(payload, cfg=cfg, root=root)

    @app.get("/api/v1/client-error/stats")
    def api_client_error_stats():
        """供 healthcheck / 管理端：近 24h 前端错误计数（无敏感）。"""
        import frontend_errors

        return frontend_errors.frontend_error_stats(cfg=cfg, root=root)

    @app.get("/api/health")
    def api_health():
        """体检状态条数据源（公开：只给绿/黄/红 + 时间 + 各源行数，不含金额/客户名）。
        任务书37·B9：fetch_banners=抓数降级黄横幅（看端/管理端顶部）。"""
        conn = db.connect(cfg, root)
        try:
            run_log = db.latest_run(conn)
        finally:
            conn.close()
        meta = (_state.get("summary") or {}).get("meta", {})
        health = meta.get("health", {})
        result = (run_log or {}).get("结果")  # 黄/红/绿：管道运行日志
        body = (run_log or {}).get("体检", {}) or {}
        reasons = _run_reasons(body)  # 「黄/红」：为啥（fetch/过期调整）
        # A3：未归属销售>0 → 至少判黄 + 顶栏短原因（沿用 v8.0 机制；不覆盖已判红）
        n_un = int((meta.get("unassigned") or {}).get("count") or 0)
        if n_un > 0:
            reasons = [f"{n_un} 名销售未归属 BU（业务不进任何 BU 页，各 BU 合计小于全公司）"] + reasons
            if result in ("绿", None):  # 未判红时至少判黄（无运行日志时 result 为 None 也升黄）
                result = "黄"
        import server as _srv

        banners = _srv.build_fetch_fallback_banners(body, cfg, root)
        # 任务书46·6：可观测指标滚动窗口（进程内；缺省 0）
        metrics = _state.get("metrics") or {}
        return {
            "result": result,
            "run_time": (run_log or {}).get("时间"),
            "built_at": _state.get("built_at"),
            "sources": health.get("sources", []),
            "warnings": health.get("warnings", []),  # 「警」：数据体检（未填分类等）
            "run_reasons": reasons,
            "fetch_banners": banners,  # B9 醒目横幅；全源成功=[]
            "metrics": {
                "update_ms": metrics.get("update_ms"),
                "fetch_fail_rate": metrics.get("fetch_fail_rate"),
                "api_p95_ms": metrics.get("api_p95_ms"),
            },
        }

    def _require(request: Request) -> str:
        user = _user(request)
        if not user:
            raise HTTPException(status_code=401, detail="需要管理员登录")
        return user

    def _conn():
        return db.connect(cfg, root)

    @app.post("/api/refresh")
    def api_refresh(request: Request):
        """立即更新=完整 pipeline（fetch+重读+重建+重放），后台线程跑、立即返回（在线抓约80秒）。
        运行中互斥，重复点返回进行中；进度轮询 /api/refresh_status。"""
        _require(request)
        if not start_refresh_async(cfg, root, "manual"):
            return JSONResponse({"status": "running", "detail": "更新进行中，请稍候"}, status_code=409)
        return {"status": "started", "refreshing": _state["refreshing"]}

    @app.get("/api/refresh_status")
    def api_refresh_status(request: Request):
        _require(request)
        return {
            "running": bool(_state["refreshing"]),
            "refreshing": _state["refreshing"],
            "last": _state["last_refresh"],
            "built_at": _state["built_at"],
            "zhiyun_auto_fetch": bool(cfg.get("zhiyun_auto_fetch")),
        }
