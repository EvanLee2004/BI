#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""任务书58·R-50：费用明细按收单日期日级筛选 + 导出注明区间/口径。

真路径：db.query_detail(date_from/date_to) + GET /api/v1/vm/ledger[+export]。
"""
from __future__ import annotations

import io
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

import accounts  # noqa: E402
import db  # noqa: E402
import loaders  # noqa: E402
import server  # noqa: E402

from db import detail as detail_mod  # noqa: E402


class TestQueryDetailDateRange(unittest.TestCase):
    """单元：query_detail 按收单日期日级闭区间，同月跨日可分出不同行集。"""

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.addCleanup(shutil.rmtree, self.tmp, True)
        cfg = loaders.load_config(ROOT)
        self.cfg = dict(cfg)
        self.cfg["data_dir"] = str(self.tmp / "数据")
        (self.tmp / "数据").mkdir(parents=True)
        accounts.seed_defaults(self.cfg, self.tmp)
        self.conn = db.connect(self.cfg, self.tmp)
        self.addCleanup(self.conn.close)
        import db_write

        # 混用两种库内格式：ISO 与生产台账 YYYYMMDD
        seeds = [
            {
                "定位键": "d58-a",
                "收单月份": "2026-03",
                "收单日期": "2026-03-05",
                "含税金额": 10.0,
                "业务BU": "数据",
                "对应报表大类": "管理费用",
                "预算明细费用类型": "办公费",
                "预算归属部门": "财务",
                "事项": "三月五日",
                "业务员": "测",
                "归属月": "2026-03",
                "原值_归属月": "2026-03",
            },
            {
                "定位键": "d58-b",
                "收单月份": "2026-03",
                "收单日期": "20260320",
                "含税金额": 20.0,
                "业务BU": "数据",
                "对应报表大类": "管理费用",
                "预算明细费用类型": "办公费",
                "预算归属部门": "财务",
                "事项": "三月二十",
                "业务员": "测",
                "归属月": "2026-03",
                "原值_归属月": "2026-03",
            },
            {
                "定位键": "d58-c",
                "收单月份": "2026-04",
                "收单日期": "2026-04-01 00:00:00",
                "含税金额": 30.0,
                "业务BU": "数据",
                "对应报表大类": "管理费用",
                "预算明细费用类型": "办公费",
                "预算归属部门": "财务",
                "事项": "四月一日",
                "业务员": "测",
                "归属月": "2026-04",
                "原值_归属月": "2026-04",
            },
        ]
        db_write.insert_std_records(self.conn, "std_费用明细", seeds)
        self.conn.commit()

    def test_same_month_day_split(self):
        all_m = detail_mod.query_detail(
            self.conn, "费用明细", month_from="2026-03", month_to="2026-03", page_size=50
        )
        self.assertEqual(all_m["total"], 2, "归属月三月应 2 行")

        early = detail_mod.query_detail(
            self.conn,
            "费用明细",
            date_from="2026-03-01",
            date_to="2026-03-10",
            page_size=50,
        )
        self.assertEqual(early["total"], 1)
        self.assertIn("三月五日", str(early["rows"]))

        late = detail_mod.query_detail(
            self.conn,
            "费用明细",
            date_from="2026-03-15",
            date_to="2026-03-31",
            page_size=50,
        )
        self.assertEqual(late["total"], 1)
        self.assertIn("三月二十", str(late["rows"]))
        self.assertNotEqual(
            {r.get("事项") for r in early["rows"]},
            {r.get("事项") for r in late["rows"]},
            "同月跨日必须分出不同行集",
        )

    def test_date_range_cross_month(self):
        d = detail_mod.query_detail(
            self.conn,
            "费用明细",
            date_from="2026-03-20",
            date_to="2026-04-01",
            page_size=50,
        )
        self.assertEqual(d["total"], 2)
        matters = {r.get("事项") for r in d["rows"]}
        self.assertEqual(matters, {"三月二十", "四月一日"})

    def test_bad_date_raises(self):
        with self.assertRaises(KeyError):
            detail_mod.query_detail(self.conn, "费用明细", date_from="2026/03/01")

    def test_yyyymmdd_production_format(self):
        """生产台账 收单日期=YYYYMMDD 也能被 date_from/to 命中。"""
        d = detail_mod.query_detail(
            self.conn,
            "费用明细",
            date_from="2026-03-20",
            date_to="2026-03-20",
            page_size=50,
        )
        self.assertEqual(d["total"], 1)
        self.assertIn("三月二十", str(d["rows"]))


class TestLedgerApiDateRange(unittest.TestCase):
    """HTTP：/api/v1/vm/ledger + export 日级参数与口径说明。"""

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.addCleanup(shutil.rmtree, self.tmp, True)
        cfg = loaders.load_config(ROOT)
        self.cfg = dict(cfg)
        self.cfg["data_dir"] = str(self.tmp / "数据")
        (self.tmp / "数据").mkdir(parents=True)
        accounts.seed_defaults(self.cfg, self.tmp)
        conn = db.connect(self.cfg, self.tmp)
        try:
            import db_write

            db_write.insert_std_records(
                conn,
                "std_费用明细",
                [
                    {
                        "定位键": "api58-1",
                        "收单月份": "2026-06",
                        "收单日期": "2026-06-03",
                        "含税金额": 11.0,
                        "业务BU": "数据",
                        "对应报表大类": "管理费用",
                        "预算明细费用类型": "办公费",
                        "预算归属部门": "财务",
                        "事项": "六月三日",
                        "业务员": "测",
                        "归属月": "2026-06",
                        "原值_归属月": "2026-06",
                    },
                    {
                        "定位键": "api58-2",
                        "收单月份": "2026-06",
                        "收单日期": "2026-06-18",
                        "含税金额": 22.0,
                        "业务BU": "数据",
                        "对应报表大类": "管理费用",
                        "预算明细费用类型": "办公费",
                        "预算归属部门": "财务",
                        "事项": "六月十八",
                        "业务员": "测",
                        "归属月": "2026-06",
                        "原值_归属月": "2026-06",
                    },
                ],
            )
            conn.commit()
        finally:
            conn.close()
        self.app = server.create_app(self.cfg, self.tmp)
        from fastapi.testclient import TestClient

        self.client = TestClient(self.app)

    def _login(self):
        r = self.client.post(
            "/api/v1/login",
            json={"account": accounts.MASTER_ACCOUNT, "password": accounts.DEFAULT_ADMIN_PW},
        )
        self.assertEqual(r.status_code, 200, r.text[:300])
        return self.client

    def test_ledger_list_day_filter(self):
        c = self._login()
        r_all = c.get(
            "/api/v1/vm/ledger",
            params={"date_from": "2026-06-01", "date_to": "2026-06-30", "page_size": 50, "show_all": 1},
        )
        self.assertEqual(r_all.status_code, 200, r_all.text[:300])
        self.assertEqual(r_all.json()["total"], 2)

        r_early = c.get(
            "/api/v1/vm/ledger",
            params={"date_from": "2026-06-01", "date_to": "2026-06-10", "page_size": 50, "show_all": 1},
        )
        self.assertEqual(r_early.status_code, 200)
        self.assertEqual(r_early.json()["total"], 1)
        flat = str(r_early.json().get("rows"))
        self.assertIn("六月三日", flat)
        self.assertNotIn("六月十八", flat)

        # 关键词仍可用
        r_q = c.get(
            "/api/v1/vm/ledger",
            params={
                "date_from": "2026-06-01",
                "date_to": "2026-06-30",
                "q": "六月十八",
                "page_size": 50,
                "show_all": 1,
            },
        )
        self.assertEqual(r_q.status_code, 200)
        self.assertEqual(r_q.json()["total"], 1)

    def test_export_notes_range_and_caliber(self):
        c = self._login()
        r = c.get(
            "/api/v1/vm/ledger/export",
            params={
                "date_from": "2026-06-01",
                "date_to": "2026-06-10",
                "show_all": 0,
            },
        )
        self.assertEqual(r.status_code, 200, r.text[:200])
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        try:
            self.assertIn("口径说明", wb.sheetnames)
            note_ws = wb["口径说明"]
            rows = [tuple(c.value for c in row) for row in note_ws.iter_rows(max_row=10)]
            flat = " ".join(str(x) for row in rows for x in row if x is not None)
            self.assertIn("2026-06-01", flat)
            self.assertIn("2026-06-10", flat)
            self.assertIn("show_all", flat)
            self.assertIn("0", flat)
            # 明细 sheet 仅含区间内行
            data_ws = wb.active
            data_rows = list(data_ws.iter_rows(values_only=True))
            body = " ".join(str(x) for row in data_rows[1:] for x in (row or ()) if x is not None)
            self.assertIn("六月三日", body)
            self.assertNotIn("六月十八", body)
        finally:
            wb.close()

    def test_frontend_controls_present(self):
        led = (ROOT / "frontend" / "src" / "components" / "LedgerTable.vue").read_text(encoding="utf-8")
        self.assertIn('type="date"', led)
        self.assertIn("date_from", led)
        self.assertIn("本月", led)
        self.assertIn("返回本年", led)
        self.assertIn("ledger-this-month", led)
        self.assertNotIn("月起", led)
        dq = (ROOT / "frontend" / "src" / "components" / "DailyQuery.vue").read_text(encoding="utf-8")
        self.assertIn("dailyThisMonth", dq)
        self.assertIn("setThisMonth", dq)
        # 本月在返回默认左边：源码顺序
        self.assertLess(dq.find("dailyThisMonth"), dq.find("dailyClose"))


if __name__ == "__main__":
    unittest.main()
