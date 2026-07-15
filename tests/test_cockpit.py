#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""经营驾驶舱回归测试：口径对账 + 手填默认上月 + 前端不做金额运算守卫。跑：python tests/test_cockpit.py"""

import datetime
import sys
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

import loaders, profit, render, assets, charts, columns, validate  # noqa: E402


def _summary():
    cfg = loaders.load_config()
    today = loaders.pinned_today(cfg)
    yr = today.year
    lh, lr = loaders.load_ledger(cfg, str(yr))
    S = profit.build_summary(
        cfg,
        loaders.load_project_detail(cfg),
        loaders.load_orders(cfg),
        loaders.load_receipts(cfg),
        loaders.load_inhouse(cfg),
        lh,
        lr,
        yr,
        today,
    )
    return cfg, S


class TestReconcile(unittest.TestCase):
    """每个周期都要满足陆总的公式（对账，防口径漂移）。"""

    def test_all_periods_reconcile(self):
        cfg, S = _summary()
        vat = cfg["tax"]["vat_rate"]
        sur = cfg["tax"]["surtax_rate"]
        for key, p in S["periods"].items():
            man = p["manual"]
            led = p["ledger_expenses"]
            prod_manual = sum(
                man[k]
                for k in [
                    "PM人力成本",
                    "VM人力成本",
                    "实际内部译员成本",
                    "税费损失",
                    "技术流量成本",
                    "其他（生产成本）",
                ]
            )
            # 生产成本 = 系统直接成本 − 内部译员成本 + 手填生产成本项
            self.assertAlmostEqual(
                p["production_cost"], p["system_direct_cost"] - p["inhouse_cost"] + prod_manual, places=1, msg=key
            )
            # 毛利
            self.assertAlmostEqual(p["gross_profit"], p["revenue_net"] - p["production_cost"], places=1, msg=key)
            # 财务费用 = 台账 + 手填补充
            self.assertAlmostEqual(p["expense"]["财务费用"], led["财务费用"] + man["财务费用补充"], places=1, msg=key)
            # 营销/管理/研发 = 手填人力 + 台账
            self.assertAlmostEqual(p["expense"]["营销费用"], man["营销人力成本"] + led["市场费用"], places=1, msg=key)
            self.assertAlmostEqual(p["expense"]["研发费用"], man["研发人力成本"] + led["技术服务费"], places=1, msg=key)
            # 附加税费 = 收入 × 6% × 12%
            self.assertAlmostEqual(p["surtax"], round(p["revenue_net"] * vat * sur, 2), places=1, msg=key)
            # 税前利润 = 毛利 − 期间费用 − 附加税费 + 其他损益
            self.assertAlmostEqual(
                p["pretax_profit"],
                p["gross_profit"] - p["expense"]["total"] - p["surtax"] + p["other_pl"],
                places=1,
                msg=key,
            )


class TestManualDefault(unittest.TestCase):
    """手填：现行 default=zero，当月未填=0（不再沿用上月）。"""

    def test_no_carry_forward(self):
        cfg = loaders.load_config()
        raw = {"2024-06": {"研发人力成本": 150000, "其他损益": 5000}}  # 7月都没填
        filled = profit.build_manual_monthly(cfg, raw, 2024, 7)
        self.assertEqual(filled[(2024, 7)]["研发人力成本"], 0.0)  # zero → 不继承
        self.assertEqual(filled[(2024, 7)]["其他损益"], 0.0)
        self.assertEqual(filled[(2024, 6)]["研发人力成本"], 150000)  # 有填的月保留
        self.assertEqual(filled[(2024, 6)]["营销人力成本"], 0.0)  # 从没填过 → 0
        # 兼容：显式 default=prev 仍可沿用
        cfg2 = {"manual_items": [{"name": "X", "default": "prev"}, {"name": "Y", "default": "zero"}]}
        f2 = profit.build_manual_monthly(cfg2, {"2024-06": {"X": 10}}, 2024, 7)
        self.assertEqual(f2[(2024, 7)]["X"], 10)
        self.assertEqual(f2[(2024, 7)]["Y"], 0.0)


class TestValueGuards(unittest.TestCase):
    """坏值防线：文本金额不崩、坏日期计数、表头重名报错、手填月份列名不补零也能读。"""

    LEDGER_HEADER = ["收单月份", "收单日期", "含税金额", "业务BU", "对应报表大类", "预算明细费用类型"]

    def test_ledger_text_amount_and_bad_date(self):
        cfg = loaders.load_config()
        lcols = columns.resolve_ledger_columns(self.LEDGER_HEADER)
        rows = [
            (7, None, "1,000.50", "公共", "管理费用", "差旅费"),  # 文本金额带千分位：以前直接崩
            ("7月", None, 200, "公共", "管理费用", "差旅费"),  # 收单月份"7月"解析不出：行被剔除但必须计数
        ]
        led, _ = profit.compute_ledger_expenses(
            rows, 2024, datetime.date(2024, 7, 1), datetime.date(2024, 7, 31), cfg, lcols
        )
        self.assertAlmostEqual(led["管理费用"], 1000.50)
        date_bad, amt_bad = profit._scan_ledger_issues(rows, 2024, lcols)
        self.assertEqual(date_bad, 1)
        self.assertEqual(amt_bad, 0)

    def test_duplicate_ledger_header_raises(self):
        with self.assertRaises(ValueError):
            columns.resolve_ledger_columns(self.LEDGER_HEADER + ["含税金额"])
        with self.assertRaises(ValueError):  # 别名互撞（业务BU 与 利润归属中心 同时存在）也要报
            columns.resolve_ledger_columns(self.LEDGER_HEADER + ["利润归属中心"])

    def test_manual_month_header_no_padding(self):
        import tempfile
        import openpyxl

        with tempfile.TemporaryDirectory() as td:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "手填与调整"
            ws.append(["项目", "归属", "备注", "2024-7"])  # 手打不补零
            ws.append(["研发人力成本", "x", "", 123])
            d = Path(td) / "d"
            d.mkdir()
            wb.save(d / "手填与调整.xlsx")
            out = loaders.load_manual({"data_dir": "d", "files": {"manual": "手填与调整.xlsx"}}, Path(td))
            self.assertEqual(out["2024-07"]["研发人力成本"], 123)


class TestValidate(unittest.TestCase):
    """进门验证：好数据全绿；故意弄坏的数据必须被点名到 源+列+Excel行号。"""

    def test_current_data_passes(self):
        cfg = loaders.load_config()
        rep = validate.validate_all(cfg, loaders.pinned_today(cfg).year)
        self.assertEqual([f"{i.source}:{i.message}" for i in rep.errors], [])

    def test_broken_data_pinpointed(self):
        import shutil
        import tempfile
        import openpyxl

        cfg = loaders.load_config()
        year = loaders.pinned_today(cfg).year
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            shutil.copytree(ROOT / "数据", root / "数据")
            # ① 台账：报表大类錯别字 + 收单月份"7月"且无收单日期 + 金额文本"abc"
            p = root / "数据" / "收单台账.xlsx"
            wb = openpyxl.load_workbook(p)
            ws = wb[str(year)]
            hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
            c_cat, c_m, c_d, c_amt = (hdr.index(x) + 1 for x in ("对应报表大类", "收单月份", "收单日期", "含税金额"))
            ws.cell(row=3, column=c_cat, value="管理费")  # 口径外类别
            ws.cell(row=4, column=c_m, value="7月")  # 月份解析不出
            ws.cell(row=4, column=c_d).value = None  # 注意:cell(...,value=None)不会清空,要显式赋值
            ws.cell(row=5, column=c_amt, value="abc")  # 非数字金额
            wb.save(p)
            # ② 项目明细：交付日期坏值
            p = root / "数据" / "项目明细.xlsx"
            wb = openpyxl.load_workbook(p)
            ws = wb.active
            hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
            ws.cell(row=6, column=hdr.index("整单交付日期") + 1, value="无")
            wb.save(p)
            rep = validate.validate_all(cfg, year, root)
            msgs = "\n".join(f"[{i.source}]{i.message}" for i in rep.errors)
            self.assertIn("管理费", msgs)
            self.assertIn("第3行", msgs)
            self.assertIn("第4行", msgs)  # 收单月份坏行
            self.assertIn("第5行", msgs)  # 金额坏行
            self.assertIn("[项目明细(智云)]", msgs)
            self.assertIn("第6行", msgs)


class TestReceiptOrderRatio(unittest.TestCase):
    """A2 回款/下单比：每期 = 回款÷下单×100（无下单置 None）；逐月序列与月周期对齐。"""

    def test_period_ratio_matches_formula(self):
        _, S = _summary()
        for key, p in S["periods"].items():
            r = p["receipt_order_ratio_pct"]
            if p["orders"]:
                self.assertAlmostEqual(r, round(p["receipts"] / p["orders"] * 100, 2), places=2, msg=key)
            else:
                self.assertIsNone(r, msg=key)

    def test_monthly_series_aligns(self):
        _, S = _summary()
        month_keys = S["meta"]["tab_groups"]["月"]
        rom = S["receipt_order_monthly"]
        self.assertEqual(len(rom), len(month_keys))
        for (label, rec, order, ratio), k in zip(rom, month_keys):
            p = S["periods"][k]
            self.assertAlmostEqual(rec, p["receipts"], places=2)
            self.assertAlmostEqual(order, p["orders"], places=2)
            self.assertEqual(ratio, p["receipt_order_ratio_pct"])


class TestRenderGuards(unittest.TestCase):
    """前端守卫：自包含 + 不做金额运算 + 含关键结构。"""

    @classmethod
    def setUpClass(cls):
        cfg, S = _summary()
        cls.html = render.render_dashboard(S, cfg, assets.load_logo_base64(cfg))

    def test_no_client_side_math(self):
        # v1.4：业务 JS 在 /static/js/cockpit.js，HTML 内联脚本不再含金额运算
        for bad in ("toFixed(", "parseFloat(", "parseInt("):
            self.assertNotIn(bad, self.html, f"前端出现金额运算 {bad}，违反'客户端不算数'铁律")
        from pathlib import Path

        js = (Path(__file__).resolve().parents[1] / "static" / "js" / "cockpit.js").read_text(encoding="utf-8")
        # 允许 parseInt 仅出现在周期 key 展示类逻辑时仍禁止金额运算关键字组合；这里继续禁 toFixed/parseFloat
        for bad in ("toFixed(", "parseFloat("):
            self.assertNotIn(bad, js, f"static/js 出现金额运算 {bad}")

    def test_self_contained(self):
        # v1.4：允许同源 /static/*，禁止外链 CDN
        self.assertNotIn('src="http', self.html)
        self.assertNotIn('href="http', self.html)
        self.assertIn('href="/static/css/theme.css"', self.html)
        self.assertIn('src="/static/js/cockpit.js"', self.html)

    def test_structure(self):
        for token in (
            "基本情况",
            "经营利润",
            "管理利润表",
            "税前利润",
            "附加税费",
            "回款情况",
            "themeBtn",
            "kpi-grid",
            "甲骨易智能经营",
            "罗盘",
        ):
            self.assertIn(token, self.html, token)

    def test_dark_default(self):
        # v1.4：暗色变量唯一入口 theme.get_css() → static/css/theme.css
        import theme

        css = theme.get_css()
        self.assertIn(":root{", css)
        self.assertIn(".theme-light{", css)
        self.assertIn('href="/static/css/theme.css"', self.html)

    def test_profit_formula_strip(self):
        # 板块③下方「计算逻辑」公式条：三条公式在位（收入/毛利率/集中度），静态口径
        self.assertIn('class="pr-formula"', self.html)
        self.assertIn("计算逻辑", self.html)
        self.assertIn("交付金额", self.html)
        self.assertIn("交付收入", self.html)
        self.assertIn("÷ 1.06", self.html)
        self.assertIn("系统成本率", self.html)  # 陆总0714 改名

    def test_receipt_month_highlight_hooks(self):
        """迭代21-A：回款柱带 data-rm、周期→月份映射 data-rm-map；看端卡头已精简。"""
        html = self.html
        self.assertIn('id="rcCard"', html)
        self.assertIn("data-rm-map=", html)
        self.assertIn("data-rm-year=", html)
        self.assertNotIn("全年视角 · 选中周期高亮", html)  # 领导视角去运营旁注
        # 柱/点/数字至少有一根 data-rm（1~12）
        self.assertRegex(html, r'data-rm="\d{1,2}"')
        # 映射含年/月/季 key 形样（Python 预生成，前端不解析）；高亮逻辑在 static/js
        from pathlib import Path
        import theme

        js = (Path(__file__).resolve().parents[1] / "static" / "js" / "cockpit.js").read_text(encoding="utf-8")
        self.assertIn("_syncRmHighlight", js)
        # 前端零金额：高亮只读写 class / 读 data 属性（class 名在 CSS，唯一入口 get_css）
        css = theme.get_css()
        self.assertIn("rm-dim", css)
        self.assertIn("rm-on", css)
        self.assertIn("rm-on", js)
        # 通用选择器：任意 [data-rm-map] 卡（回款+趋势），非仅 #rcCard
        self.assertIn("[data-rm-map]", js)
        self.assertIn(".rc-rm-filter", css)

    def test_trend_month_highlight_hooks(self):
        """经营利润趋势图复用回款卡 data-rm-map / data-rm 机制。"""
        html = self.html
        self.assertIn('id="trendCard"', html)
        # 趋势卡根挂 map（与回款卡同格式）
        self.assertRegex(html, r'id="trendCard"[^>]*data-rm-map=')
        self.assertRegex(html, r'id="trendCard"[^>]*data-rm-year=')
        # 趋势图柱组带 data-rm
        from pathlib import Path
        import charts

        svg = charts.combo_bar_line_chart([("1月", 100.0, 40.0, 60.0), ("2月", 80.0, 30.0, 62.5)], None)
        self.assertIn('data-rm="1"', svg)
        self.assertIn('data-rm="2"', svg)
        # 收入柱 + 成本柱 + 毛利率点均带标识
        self.assertGreaterEqual(svg.count('data-rm="1"'), 3)


class TestUnclassifiedPeriodTip(unittest.TestCase):
    """看端整体页不再展示底部「口径提示」淡字（未分类仍进全年利润表 + 管理端体检）。"""

    def test_dashboard_hides_faint_unclassified_tip(self):
        cfg, S = _summary()
        S["meta"]["unclassified"]["expense"] = {"count": 3, "amount": 123456.0}
        html = render.render_dashboard(S, cfg, assets.load_logo_base64(cfg))
        self.assertNotIn("口径提示", html)
        self.assertNotIn("待分类费用尚未计入", html)
        self.assertNotIn("全年另有待分类台账费用未计入", html)
        self.assertNotIn("税前利润略偏高", html)

    def test_no_unclassified_still_no_tip(self):
        cfg, S = _summary()
        S["meta"]["unclassified"]["expense"] = {"count": 0, "amount": 0.0}
        html = render.render_dashboard(S, cfg, assets.load_logo_base64(cfg))
        self.assertNotIn("口径提示", html)
        self.assertNotIn("待分类费用尚未计入", html)


class TestReceiptsBudgetLayout(unittest.TestCase):
    """迭代19：陆总拍板去掉部门费用年预算卡；回款整宽，页面不再出现该卡。"""

    GRID = '<div class="grid-2e rb-grid" style="margin-top:16px"><div class="period-receipts">'
    FULL = '<div class="period-receipts" style="margin-top:16px">'

    @staticmethod
    def _render(dept_budget):
        cfg, S = _summary()
        S["meta"]["dept_budget"] = dept_budget
        return render.render_dashboard(S, cfg, assets.load_logo_base64(cfg))

    def test_no_dept_budget_card_even_when_filled(self):
        db = {"year": 2026, "rows": [{"dept": "示例部", "used": 10.0, "target": 8.0, "pct": 125.0}]}
        html = self._render(db)
        self.assertIn("回款情况", html)
        self.assertIn(self.FULL, html)
        self.assertNotIn(self.GRID, html)
        # 用户可见卡头已下线（CSS 注释等可残留，不算卡）
        self.assertNotIn("部门费用预算执行 ", html)  # 卡头带空格 tag
        self.assertNotIn("暂无部门年预算", html)
        self.assertEqual(render.render_dept_budget(db), "")

    def test_receipts_full_width_no_budget_card(self):
        html = self._render(None)
        self.assertIn("回款情况", html)
        self.assertIn(self.FULL, html)
        self.assertNotIn("暂无部门年预算", html)
        self.assertNotIn(self.GRID, html)

    def test_receipt_split_symmetric_half_half(self):
        """v1.0.4：回款卡左图/右驾驶舱各占一半（明昊拍板），右栏不再压 280px。"""
        import theme

        css = theme.get_css()
        self.assertIn(".rc-card .rc-split{display:grid;grid-template-columns:minmax(0,1fr) minmax(0,1fr)", css)
        self.assertNotIn("minmax(200px,280px)", css)
        self.assertNotIn(".rc-side{max-width:280px}", css)

    def test_receipt_body_column_layout(self):
        """迭代20-A1：图例必须在图下方（rc-body 纵向列），不得与 SVG 左右并排。"""
        import theme

        css = theme.get_css()
        self.assertIn("flex-direction:column", css.split(".rc-card .rc-body{")[1].split("}")[0])

    def test_note_texts_enlarged(self):
        """v1.0.4：口径/公式小字统一放大提亮（--note 色，chart-note/pr-formula >=13px）。"""
        import theme

        css = theme.get_css()
        self.assertIn("--note:", css)
        self.assertIn(".chart-note{font-size:13.5px;color:var(--note)", css)
        self.assertIn(".pr-formula{margin-top:14px;padding:12px 16px", css)
        self.assertNotIn(".chart-note{font-size:12.5px", css)


class TestRankingsAndRanges(unittest.TestCase):
    """板块③排名（下单按部门/销售、回款按客户）+ 自定义月区间周期。"""

    def test_compute_ranking_sorted_top_others(self):
        rows = (
            [{"名": "甲", "额": 100, "日": "2026-03-05"}] * 2
            + [{"名": "乙", "额": 350, "日": "2026-03-08"}]
            + [{"名": "", "额": 50, "日": "2026-03-09"}]  # 空名归（未填）
            + [{"名": "丙", "额": 999, "日": "2026-07-01"}]
        )  # 期外不计
        import datetime as dt

        rk = profit.compute_ranking(rows, "名", "额", "日", dt.date(2026, 3, 1), dt.date(2026, 3, 31), top=2)
        self.assertEqual([i["name"] for i in rk["items"]], ["乙", "甲"])  # 降序
        self.assertEqual(rk["items"][1], {"name": "甲", "amount": 200.0, "count": 2})
        # v7.4：（未填）不再进排名/其余，单拆 unfilled 置底展示（total 仍含它=守恒）
        self.assertIsNone(rk["others"])
        self.assertEqual(rk["unfilled"], {"amount": 50.0, "count": 1})
        self.assertAlmostEqual(rk["total"], 600.0)

    def test_period_ranges_include_month_spans(self):
        import datetime as dt, periods

        r = periods.all_period_ranges(dt.date(2026, 7, 15))
        self.assertIn("2026年1-3月", r)
        label, start, end, group = r["2026年1-3月"]
        self.assertEqual((label, group), ("2026年1~3月", "区间"))
        self.assertEqual((start, end), (dt.date(2026, 1, 1), dt.date(2026, 3, 31)))
        self.assertNotIn("2026年3-1月", r)  # 只生成 m1<m2
        self.assertNotIn("2026年7-8月", r)  # 未来月不生成
        # 组合数 = C(7,2)=21
        self.assertEqual(len([k for k in r if r[k][3] == "区间"]), 21)

    def test_rendered_rankings_and_picker(self):
        cfg, S = _summary()
        html = render.render_dashboard(S, cfg, assets.load_logo_base64(cfg))
        for token in ("下单与回款", "下单/回款 · 按销售", "下单/回款 · 按客户", "periodBtn", "ppanel", "pp-grid"):
            self.assertIn(token, html, token)
        # 迭代20-A2：回款情况卡并入板块④「下单与回款」（在③收入与毛利结构之后）
        self.assertLess(html.index("收入与毛利结构"), html.index('class="period-receipts"'))
        self.assertLess(html.index("下单与回款"), html.index('class="period-receipts"'))
        # 区间周期块已预渲染（前端只切显示、不算数）
        yr = S["meta"]["year"]
        if S["meta"]["tab_groups"]["区间"]:
            self.assertIn(f'data-blk="{S["meta"]["tab_groups"]["区间"][0]}"', html)
        # 区间周期的排名 = 成员月排名之和（口径自洽抽查：1-2月 总额 == 1月+2月）
        k12, k1, k2 = f"{yr}年1-2月", f"{yr}年1月", f"{yr}年2月"
        P = S["periods"]
        if k12 in P:
            self.assertAlmostEqual(
                P[k12]["rankings"]["orders_by_sales"]["total"],
                round(P[k1]["rankings"]["orders_by_sales"]["total"] + P[k2]["rankings"]["orders_by_sales"]["total"], 2),
                places=1,
            )


if __name__ == "__main__":
    unittest.main(verbosity=2)
