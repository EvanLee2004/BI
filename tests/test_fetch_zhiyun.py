#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""fetch_zhiyun 离线单测：解析器/翻页/必需列护栏/三态返回/xlsx 产物可被 loaders 读回。
不碰网络——post 用测试桩注入。"""

from __future__ import annotations

import json
import sys
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from ingest import fetch_zhiyun as fz  # noqa: E402


def ctrl(cid, name, options=None):
    return {"controlId": cid, "controlName": name, "options": options or []}


class TestParseCell(unittest.TestCase):
    def test_plain_and_empty(self):
        c = ctrl("c1", "文本")
        self.assertEqual(fz.parse_cell("abc", c), "abc")
        self.assertEqual(fz.parse_cell("7806.84", c), "7806.84")
        self.assertEqual(fz.parse_cell(None, c), "")
        self.assertEqual(fz.parse_cell("", c), "")

    def test_option_keys_translated(self):
        c = ctrl("c1", "单选", options=[{"key": "k1", "value": "美元"}, {"key": "k2", "value": "人民币"}])
        self.assertEqual(fz.parse_cell(json.dumps(["k1"]), c), "美元")
        self.assertEqual(fz.parse_cell(json.dumps(["k1", "k2"]), c), "美元/人民币")
        self.assertEqual(fz.parse_cell(json.dumps(["k9"]), c), "k9")  # 未知key回退

    def test_member_department_relation(self):
        c = ctrl("c1", "人")
        self.assertEqual(fz.parse_cell(json.dumps([{"fullname": "于占国"}]), c), "于占国")
        self.assertEqual(fz.parse_cell(json.dumps([{"departmentName": "Multi-language"}]), c), "Multi-language")
        self.assertEqual(fz.parse_cell(json.dumps([{"name": "北京多语"}]), c), "北京多语")

    def test_broken_json_falls_back(self):
        c = ctrl("c1", "坏")
        self.assertEqual(fz.parse_cell("[not json", c), "[not json")

    def test_cell_already_object(self):
        c = ctrl("c1", "人")
        self.assertEqual(fz.parse_cell([{"fullname": "于占国"}], c), "于占国")
        c2 = ctrl("c2", "单选", options=[{"key": "k1", "value": "美元"}])
        self.assertEqual(fz.parse_cell(["k1"], c2), "美元")


class TestRowsToRecords(unittest.TestCase):
    def test_maps_control_names(self):
        controls = [ctrl("ca", "下单日期"), ctrl("cb", "下单预估额/本币")]
        rows = [{"ca": "2026-06-23", "cb": "100.5"}, {"ca": "2026-06-24", "cb": ""}]
        recs = fz.rows_to_records(rows, controls)
        self.assertEqual(recs[0], {"下单日期": "2026-06-23", "下单预估额/本币": "100.5"})
        self.assertEqual(recs[1]["下单预估额/本币"], "")

    def test_duplicate_name_non_empty_wins(self):
        """同名列合并：空值不覆盖有值（防两个"整单交付日期"的空把有值清掉）。"""
        controls = [ctrl("c_val", "整单交付日期"), ctrl("c_empty", "整单交付日期")]
        rows = [{"c_val": "2026-06-30", "c_empty": ""}]  # 有值在前、空在后
        self.assertEqual(fz.rows_to_records(rows, controls)[0]["整单交付日期"], "2026-06-30")
        rows2 = [{"c_val": "", "c_empty": "2026-06-30"}]  # 空在前、有值在后
        self.assertEqual(fz.rows_to_records(rows2, controls)[0]["整单交付日期"], "2026-06-30")


class TestDateSinceFilter(unittest.TestCase):
    CTRLS = [{"controlId": "d1", "controlName": "下单日期", "type": 15}]

    def test_builds_filter_value_is_day_before_since(self):
        """任务书35：filterType=13 严格大于 value → value=since 前一天才含 since 当天。"""
        fc = fz.build_date_since_filter(self.CTRLS, "下单日期", "2026-01-01")
        self.assertEqual(len(fc), 1)
        self.assertEqual(fc[0]["controlId"], "d1")
        self.assertEqual(fc[0]["filterType"], 13)
        self.assertEqual(fc[0]["value"], "2025-12-31")  # 前一天，不是 2026-01-01
        self.assertEqual(fz._since_filter_value("2026-03-01"), "2026-02-28")  # 非字符串硬减

    def test_empty_when_no_since_or_missing_col(self):
        self.assertEqual(fz.build_date_since_filter(self.CTRLS, "下单日期", ""), [])
        self.assertEqual(fz.build_date_since_filter(self.CTRLS, "不存在的列", "2026-01-01"), [])

    def test_uses_first_of_duplicate_date_controls(self):
        ctrls = [
            {"controlId": "first", "controlName": "整单交付日期", "type": 15},
            {"controlId": "second", "controlName": "整单交付日期", "type": 15},
        ]
        fc = fz.build_date_since_filter(ctrls, "整单交付日期", "2026-01-01")
        self.assertEqual(fc[0]["controlId"], "first")
        self.assertEqual(len(fz.controls_with_name(ctrls, "整单交付日期")), 2)

    def test_resolve_zhiyun_since_auto_and_fixed(self):
        """任务书36·E：auto=当年元旦；写死日期兼容；空=全量。"""
        import datetime as _dt

        self.assertEqual(fz.resolve_zhiyun_since("auto", today=_dt.date(2027, 1, 15)), "2027-01-01")
        self.assertEqual(fz.resolve_zhiyun_since("AUTO", today=_dt.date(2026, 6, 1)), "2026-01-01")
        self.assertEqual(fz.resolve_zhiyun_since("2025-03-01"), "2025-03-01")
        self.assertEqual(fz.resolve_zhiyun_since(""), "")
        self.assertEqual(fz.resolve_zhiyun_since(None), "")
        # auto → filter value = 前一年 12-31
        fc = fz.build_date_since_filter(self.CTRLS, "下单日期", "auto", today=_dt.date(2027, 1, 15))
        self.assertEqual(fc[0]["value"], "2026-12-31")


class TestAuthExpiry(unittest.TestCase):
    def test_detects_expired(self):
        self.assertTrue(fz._is_auth_expired({"state": 0, "exception": "帐号已退出，请重新登录"}))
        self.assertTrue(fz._is_auth_expired({"state": "0", "message": "请登录"}))

    def test_not_expired_on_success_or_other_error(self):
        self.assertFalse(fz._is_auth_expired({"state": 1, "data": {}}))
        self.assertFalse(fz._is_auth_expired({"state": 0, "exception": "无权限查看该表"}))
        self.assertFalse(fz._is_auth_expired("非字典"))


class TestLoginGuards(unittest.TestCase):
    def test_missing_credentials_raises(self):
        from ingest import login_zhiyun as lz

        for zy in ({}, {"base_url": "http://x"}, {"base_url": "http://x", "username": "u"}):
            with self.assertRaises(lz.LoginError):
                lz.login(zy)


class TestRequiredColumns(unittest.TestCase):
    CFG = {"columns": {"order_amount": "下单预估额/本币", "order_date": "下单日期"}}

    def test_ok_and_missing(self):
        good = [{"下单预估额/本币": "1", "下单日期": "2026-01-01"}]
        self.assertEqual(fz.check_required_columns(good, self.CFG, "orders"), [])
        bad = [{"别的列": "x"}]
        self.assertEqual(fz.check_required_columns(bad, self.CFG, "orders"), ["下单预估额/本币", "下单日期"])
        self.assertTrue(fz.check_required_columns([], self.CFG, "orders"))  # 空=缺


class TestPagination(unittest.TestCase):
    def test_pages_until_short_page(self):
        pages = [[{"i": n} for n in range(fz.PAGE_SIZE)], [{"i": "last"}]]
        total = fz.PAGE_SIZE + 1
        calls = []

        def post(path, body):
            calls.append((body["pageIndex"], body.get("notGetTotal")))
            d = {"data": pages[body["pageIndex"] - 1]}
            if body["pageIndex"] == 1:
                d["count"] = total
            return {"data": d}

        rows = fz.fetch_all_rows(post, "ws1", "app1")
        self.assertEqual(len(rows), total)
        self.assertEqual(calls[0], (1, False))  # 首页要 total
        self.assertEqual(calls[1], (2, True))

    def test_empty_table(self):
        rows = fz.fetch_all_rows(lambda p, b: {"data": {"data": [], "count": 0}}, "ws1", "app1")
        self.assertEqual(rows, [])

    def test_runaway_pagination_raises(self):
        full = [{"i": 1}] * fz.PAGE_SIZE  # 永远回满页 → 必须撞上限而不是死循环
        with self.assertRaises(RuntimeError):
            fz.fetch_all_rows(lambda p, b: {"data": {"data": full, "count": 10**9}}, "ws1", "app1")

    def test_total_mismatch_raises(self):
        """任务书30·0.5 / 35：total=2500 只给 2 页残缺 → 拒收。"""
        # 2 满页 = 2000 < 2500
        pages = {
            1: [{"i": n} for n in range(fz.PAGE_SIZE)],
            2: [{"i": n} for n in range(fz.PAGE_SIZE)],  # 仍满页，但我们用 count 对账
        }
        # 模拟第 3 页空：实际上 2 满页后继续翻——改成 page2 short wrong total
        pages[2] = [{"i": n} for n in range(100)]  # short page, total rows=1100 != 2500

        def post(path, body):
            idx = body["pageIndex"]
            d = {"data": pages.get(idx, [])}
            if idx == 1:
                d["count"] = 2500
            return {"data": d}

        with self.assertRaises(RuntimeError) as cm:
            fz.fetch_all_rows(post, "ws1", "app1")
        self.assertIn("行数对账失败", str(cm.exception))
        self.assertIn("2500", str(cm.exception))

    def test_write_empty_records_raises(self):
        import tempfile

        with tempfile.TemporaryDirectory() as td:
            with self.assertRaises(ValueError):
                fz.write_records_xlsx([], Path(td) / "x.xlsx")


class TestZhiyunDefaults(unittest.TestCase):
    """内置连接默认（2026-07-13 明昊拍板进公开库）：文件缺失也有完整连接信息；文件非空值覆盖默认。"""

    def _cfg(self, tmp):
        return {"data_dir": str(tmp)}

    def test_missing_file_yields_defaults(self):
        import tempfile

        with tempfile.TemporaryDirectory() as td:
            zy = fz._load_zhiyun_cfg(self._cfg(Path(td)), Path(td))
            self.assertEqual(zy["base_url"], fz.ZHIYUN_DEFAULTS["base_url"])
            self.assertEqual(zy["app_id"], fz.ZHIYUN_DEFAULTS["app_id"])
            for s in ("orders", "receipts", "project_detail", "inhouse"):
                self.assertTrue(zy["tables"][s]["worksheetId"], f"{s} 表ID应有默认值")
            self.assertEqual(zy["tables"]["inhouse"]["min_rows"], 1000)  # 权限护栏门槛随默认走
            self.assertFalse(zy.get("username"))  # 账号密码绝不内置

    def test_file_overrides_defaults_and_blank_ignored(self):
        import json as _json
        import tempfile

        with tempfile.TemporaryDirectory() as td:
            (Path(td) / "智云配置.json").write_text(
                _json.dumps(
                    {
                        "username": "u1",
                        "password": "p1",
                        "base_url": "http://other:1",
                        "app_id": "",  # 空值不覆盖默认
                        "tables": {"orders": {"worksheetId": "custom-ws"}, "receipts": {"worksheetId": ""}},
                    }
                ),
                encoding="utf-8",
            )
            zy = fz._load_zhiyun_cfg(self._cfg(Path(td)), Path(td))
            self.assertEqual(zy["base_url"], "http://other:1")  # 文件非空值胜出
            self.assertEqual(zy["app_id"], fz.ZHIYUN_DEFAULTS["app_id"])  # 空串不覆盖
            self.assertEqual(zy["tables"]["orders"]["worksheetId"], "custom-ws")
            self.assertEqual(
                zy["tables"]["receipts"]["worksheetId"], fz.ZHIYUN_DEFAULTS["tables"]["receipts"]["worksheetId"]
            )
            self.assertEqual(zy["username"], "u1")

    def test_save_session_creates_missing_file(self):
        """连接走内置默认（无文件）时登录成功也要能持久化 token，否则每轮更新重登。"""
        import json as _json
        import tempfile

        with tempfile.TemporaryDirectory() as td:
            cfg = self._cfg(Path(td))
            fz._save_session(cfg, Path(td), "TOK1", "ACC1")
            d = _json.loads((Path(td) / "智云配置.json").read_text(encoding="utf-8"))
            self.assertEqual((d["md_pss_id"], d["account_id"]), ("TOK1", "ACC1"))


class TestFetchSourceStates(unittest.TestCase):
    """三态与产物端到端（临时目录、post 桩）。"""

    def _cfg(self, tmp):
        return {
            "data_dir": str(tmp),
            "files": {
                "orders": "下单.xlsx",
                "receipts": "回款记录.xlsx",
                "project_detail_stem": "项目明细",
                "inhouse": "内部译员.xlsx",
            },
            "columns": {"order_amount": "下单预估额/本币", "order_date": "下单日期"},
        }

    def _zy(self):
        return {
            "base_url": "http://x",
            "app_id": "app1",
            "account_id": "acc1",
            "md_pss_id": "cookie",
            "tables": {"orders": {"worksheetId": "ws1"}},
        }

    def _post_ok(self, path, body):
        if path.endswith("getWorksheetInfo"):
            return {
                "data": {
                    "template": {
                        "controls": [ctrl("ca", "下单日期"), ctrl("cb", "下单预估额/本币"), ctrl("cc", "客户名称")]
                    }
                }
            }
        # GetFilterRows：count 与 data 对齐（批次0.5 对账）
        return {"data": {"data": [{"ca": "2026-06-01", "cb": "12.5", "cc": "客户A"}], "count": 1}}

    def test_blank_base_url_no_local(self):
        """服务器地址为空（正常合并流程出不来，守卫直接传入）→ 不抓、no_source。"""
        import tempfile

        with tempfile.TemporaryDirectory() as td:
            r = fz.fetch_source(self._cfg(Path(td)), "orders", root=Path(td), zy={"base_url": "", "tables": {}})
            self.assertEqual(r["status"], "no_source")
            self.assertIn("服务器地址为空", r["detail"])

    def test_blank_base_url_with_local_fallback(self):
        import tempfile

        with tempfile.TemporaryDirectory() as td:
            (Path(td) / "下单.xlsx").write_bytes(b"PK\x03\x04")
            r = fz.fetch_source(self._cfg(Path(td)), "orders", root=Path(td), zy={"base_url": "", "tables": {}})
            self.assertEqual(r["status"], "local_fallback")

    def test_fetched_and_readable(self):
        import tempfile
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as td:
            cfg = self._cfg(Path(td))
            r = fz.fetch_source(cfg, "orders", root=Path(td), post=self._post_ok, zy=self._zy())
            self.assertEqual(r["status"], "fetched", r["detail"])
            ws = load_workbook(Path(td) / "下单.xlsx").active
            rows = list(ws.values)
            self.assertEqual(rows[0], ("下单日期", "下单预估额/本币", "客户名称"))
            self.assertEqual(rows[1], ("2026-06-01", "12.5", "客户A"))

    def test_missing_required_column_not_written(self):
        import tempfile

        def post(path, body):
            if path.endswith("getWorksheetInfo"):
                return {"data": {"template": {"controls": [ctrl("cc", "客户名称")]}}}
            return {"data": {"data": [{"cc": "客户A"}], "count": 1}}

        with tempfile.TemporaryDirectory() as td:
            r = fz.fetch_source(self._cfg(Path(td)), "orders", root=Path(td), post=post, zy=self._zy())
            self.assertEqual(r["status"], "no_source")
            self.assertIn("缺必需列", r["detail"])
            self.assertFalse((Path(td) / "下单.xlsx").exists())  # 坏产物绝不落盘

    def test_since_boundary_day_included_in_filter_and_rows(self):
        """语义：since=2026-01-01 时，发出的 filter value=2025-12-31；归属日=since 的行在桩数据中可被抓进。"""
        import tempfile

        seen_filters = []

        def post(path, body):
            if path.endswith("getWorksheetInfo"):
                return {
                    "data": {
                        "template": {
                            "controls": [
                                ctrl("ca", "下单日期"),
                                ctrl("cb", "下单预估额/本币"),
                            ]
                        }
                    }
                }
            seen_filters.append(body.get("filterControls") or [])
            # 桩：两行，含 since 当天与之后
            rows = [
                {"ca": "2026-01-01", "cb": "1"},
                {"ca": "2026-01-02", "cb": "2"},
            ]
            return {"data": {"data": rows, "count": 2}}

        with tempfile.TemporaryDirectory() as td:
            cfg = self._cfg(Path(td))
            cfg["zhiyun_since"] = "2026-01-01"
            r = fz.fetch_source(cfg, "orders", root=Path(td), post=post, zy=self._zy())
            self.assertEqual(r["status"], "fetched", r["detail"])
            self.assertEqual(r.get("rows"), 2)
            self.assertTrue(seen_filters)
            fc = seen_filters[0]
            self.assertEqual(fc[0]["value"], "2025-12-31")
            # 产物含 since 当天行
            from openpyxl import load_workbook

            vals = list(load_workbook(Path(td) / "下单.xlsx").active.values)
            dates = {v[0] for v in vals[1:]}
            self.assertIn("2026-01-01", dates)

    def test_row_drop_warns_yellow_not_block(self):
        """骤降：上次 100 本次 50、阈值 30% → warnings，status 仍 fetched。"""
        import tempfile

        with tempfile.TemporaryDirectory() as td:
            cfg = self._cfg(Path(td))
            cfg["zhiyun_row_drop_ratio"] = 0.3
            fz.save_last_row_count(cfg, "orders", 100, root=Path(td))
            r = fz.fetch_source(cfg, "orders", root=Path(td), post=self._post_ok, zy=self._zy())
            self.assertEqual(r["status"], "fetched")
            self.assertTrue(r.get("warnings"))
            self.assertTrue(any("骤降" in w for w in r["warnings"]))
            # 成功后更新上次行数
            self.assertEqual(fz.load_last_row_counts(cfg, Path(td))["orders"], 1)

    def test_duplicate_date_control_warns(self):
        import tempfile

        def post(path, body):
            if path.endswith("getWorksheetInfo"):
                return {
                    "data": {
                        "template": {
                            "controls": [
                                ctrl("c1", "下单日期"),
                                ctrl("c2", "下单日期"),  # 同名
                                ctrl("cb", "下单预估额/本币"),
                            ]
                        }
                    }
                }
            return {"data": {"data": [{"c1": "2026-06-01", "c2": "", "cb": "1"}], "count": 1}}

        with tempfile.TemporaryDirectory() as td:
            r = fz.fetch_source(self._cfg(Path(td)), "orders", root=Path(td), post=post, zy=self._zy())
            self.assertEqual(r["status"], "fetched")
            self.assertTrue(any("同名控件" in w for w in (r.get("warnings") or [])))

    def test_min_rows_guard_blocks_permission_starved_account(self):
        """行数门槛：抓到的行数 < tables.<源>.min_rows（=账号行级权限不足，如亮晶号在
        任务表只看得到『我的任务』85行）→ 降级、绝不用残缺数据覆盖现有文件。"""
        import tempfile

        zy = self._zy()
        zy["tables"]["orders"]["min_rows"] = 100  # 桩只返回1行 < 100
        with tempfile.TemporaryDirectory() as td:
            old = Path(td) / "下单.xlsx"
            old.write_bytes(b"OLD")
            r = fz.fetch_source(self._cfg(Path(td)), "orders", root=Path(td), post=self._post_ok, zy=zy)
            self.assertEqual(r["status"], "local_fallback")
            self.assertIn("门槛", r["detail"])
            self.assertEqual(old.read_bytes(), b"OLD")  # 现有文件原样保留

    def test_fetch_all_unreachable_server_fast_fallback(self):
        """连通性探测：内网不可达 → 四源整体快速降级（不逐源等超时）。"""
        import json as _json
        import tempfile

        orig = fz._server_reachable
        fz._server_reachable = lambda base_url, timeout=5: False
        try:
            with tempfile.TemporaryDirectory() as td:
                cfg = self._cfg(Path(td))
                (Path(td) / "下单.xlsx").write_bytes(b"PK\x03\x04")
                (Path(td) / "智云配置.json").write_text(
                    _json.dumps(
                        {
                            "base_url": "http://10.9.9.9",
                            "username": "u",
                            "password": "p",
                            "app_id": "a",
                            "account_id": "acc",
                            "tables": {s: {"worksheetId": "w"} for s in fz.SOURCES},
                        }
                    ),
                    encoding="utf-8",
                )
                res = fz.fetch_all(cfg, root=Path(td))
                self.assertEqual(set(res), set(fz.SOURCES))
                self.assertEqual(res["orders"]["status"], "local_fallback")  # 有本地文件
                self.assertEqual(res["receipts"]["status"], "no_source")  # 无本地文件
                self.assertIn("不可达", res["orders"]["detail"])
        finally:
            fz._server_reachable = orig

    def test_fetch_all_login_failure_fast_fallback_single_attempt(self):
        """登录失败：四源整体快速降级，且只试一次登录（不逐源重试——慢+密码错反复试有锁号风险）。"""
        import json as _json
        import tempfile
        from ingest import login_zhiyun

        calls = {"n": 0}

        def boom(zy, headless=True):
            calls["n"] += 1
            raise login_zhiyun.LoginError("账号或密码错误")

        orig_login, orig_reach = login_zhiyun.login, fz._server_reachable
        login_zhiyun.login, fz._server_reachable = boom, (lambda b, timeout=5: True)
        try:
            with tempfile.TemporaryDirectory() as td:
                cfg = self._cfg(Path(td))
                (Path(td) / "下单.xlsx").write_bytes(b"PK\x03\x04")
                (Path(td) / "智云配置.json").write_text(
                    _json.dumps(
                        {
                            "base_url": "http://x",
                            "username": "u",
                            "password": "bad",
                            "app_id": "a",
                            "account_id": "acc",
                            "md_pss_id": "",
                            "tables": {s: {"worksheetId": "w"} for s in fz.SOURCES},
                        }
                    ),
                    encoding="utf-8",
                )
                res = fz.fetch_all(cfg, root=Path(td))
                self.assertEqual(calls["n"], 1)  # 只登录一次
                self.assertEqual(res["orders"]["status"], "local_fallback")
                self.assertIn("登录失败", res["orders"]["detail"])
        finally:
            login_zhiyun.login, fz._server_reachable = orig_login, orig_reach

    def test_make_post_no_relogin_loop_after_failure(self):
        """共享 post：token 失效且重登失败后，后续调用不再反复起浏览器登录。"""
        from ingest import login_zhiyun

        calls = {"n": 0}

        def boom(zy, headless=True):
            calls["n"] += 1
            raise login_zhiyun.LoginError("密码错")

        orig = login_zhiyun.login
        login_zhiyun.login = boom
        try:
            import types
            import sys

            zy = {"base_url": "http://x", "account_id": "a", "md_pss_id": "DEAD", "username": "u", "password": "bad"}

            # 桩掉 requests：永远返回 state==0 需登录
            class _R:
                status_code = 200

                def json(self):
                    return {"state": 0, "exception": "请重新登录"}

                def raise_for_status(self):
                    pass

            fake = types.ModuleType("requests")
            fake.post = lambda *a, **k: _R()
            sys.modules["requests"] = fake
            try:
                post = fz._make_post(zy, cfg={}, root=None)
                for _ in range(3):
                    with self.assertRaises(Exception):
                        post("Worksheet/GetFilterRows", {})
            finally:
                del sys.modules["requests"]
            self.assertEqual(calls["n"], 1)  # 三次调用只登录一次
        finally:
            login_zhiyun.login = orig

    def test_api_error_keeps_old_file(self):
        import tempfile

        def post(path, body):
            raise RuntimeError("boom")

        with tempfile.TemporaryDirectory() as td:
            old = Path(td) / "下单.xlsx"
            old.write_bytes(b"OLD")
            r = fz.fetch_source(self._cfg(Path(td)), "orders", root=Path(td), post=post, zy=self._zy())
            self.assertEqual(r["status"], "local_fallback")
            self.assertEqual(old.read_bytes(), b"OLD")  # 旧文件原样保留


if __name__ == "__main__":
    unittest.main(verbosity=2)
