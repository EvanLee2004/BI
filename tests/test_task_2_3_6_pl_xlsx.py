#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""2.3.6→2.4.0 管理利润表 Excel 导出：单 sheet 一页化 + 周期敏感 + HTTP 鉴权隔离。"""
from __future__ import annotations

import io
import json
import os
import shutil
import sys
import tempfile
import unittest
from pathlib import Path
from urllib.parse import quote

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

import accounts  # noqa: E402
import bu  # noqa: E402
import core  # noqa: E402
import db  # noqa: E402
import loaders  # noqa: E402
import openpyxl  # noqa: E402
import server  # noqa: E402
from export_pl_xlsx import build_pl_xlsx_bytes, pl_xlsx_filename  # noqa: E402
from viewmodels.packers import pack_pl_by_period  # noqa: E402

FAKE = ROOT / "_golden_data"


def _load_golden_summary():
    if not FAKE.exists():
        raise unittest.SkipTest("缺 _golden_data")
    cfg = dict(loaders.load_config(ROOT))
    cfg["data_dir"] = "_golden_data"
    cfg["db_path"] = "_golden_data/看板.db"
    cfg["zhiyun_auto_fetch"] = False
    today = loaders.pinned_today(cfg)
    conn = db.connect(cfg, ROOT)
    try:
        summary = core.summary_from_conn(cfg, conn, today)
    finally:
        conn.close()
    return summary


def _data_start_row(ws) -> int:
    """表头「科目」所在行（抬头块之后）。"""
    for r in range(1, min(ws.max_row, 30) + 1):
        if ws.cell(row=r, column=1).value == "科目":
            return r
    raise AssertionError("未找到表头「科目」")


class TestBuildPlXlsxPure(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.summary = _load_golden_summary()
        cls.yk = cls.summary["meta"]["year_key"]
        cls.periods = list((cls.summary.get("periods") or {}).keys())

    def test_single_sheet_only(self):
        """2.4.0：全程一张 sheet，无「构成_*」「导出说明」独立页。"""
        raw = build_pl_xlsx_bytes(
            self.summary,
            period_key=self.yk,
            is_bu=False,
            scope_label="整体",
            version="2.4.0",
            export_time="2026-07-23 12:00:00",
        )
        self.assertIsInstance(raw, (bytes, bytearray))
        self.assertGreater(len(raw), 100)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        self.assertEqual(len(wb.worksheets), 1, wb.sheetnames)
        self.assertEqual(wb.sheetnames, ["管理利润表"])
        self.assertFalse(any(n.startswith("构成_") for n in wb.sheetnames))
        self.assertNotIn("导出说明", wb.sheetnames)

    def test_main_sheet_bold_header_and_nested_details(self):
        raw = build_pl_xlsx_bytes(
            self.summary,
            period_key=self.yk,
            is_bu=False,
            scope_label="整体",
            version="2.4.0",
            export_time="2026-07-23 12:00:00",
        )
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        self.assertEqual(len(wb.worksheets), 1)
        ws = wb["管理利润表"]
        # 抬头块
        blob_meta = " ".join(
            str(ws.cell(row=r, column=c).value or "")
            for r in range(1, 8)
            for c in range(1, 3)
        )
        self.assertIn("甲骨易经营看板", blob_meta)
        self.assertIn("整体", blob_meta)
        self.assertIn(self.yk, blob_meta)
        self.assertIn("2.4.0", blob_meta)

        hdr = _data_start_row(ws)
        self.assertEqual(ws.cell(row=hdr, column=1).value, "科目")
        self.assertTrue(ws.cell(row=hdr, column=1).font.bold)

        packed = pack_pl_by_period(self.summary, is_bu=False)[self.yk]
        pack_rows = list(packed.get("rows") or [])
        pack_details = dict(packed.get("details") or {})

        # 扫描数据区：大类加粗；有 open_key 的明细紧随其后且非加粗+有缩进
        r = hdr + 1
        for pr in pack_rows:
            name = pr.get("name") or ""
            cell = ws.cell(row=r, column=1)
            self.assertEqual(cell.value, name, f"row {r} expected category {name}")
            self.assertTrue(cell.font and cell.font.bold, f"category {name} should be bold")
            # 金额与 pack 的 amt_disp 一致
            self.assertEqual(
                ws.cell(row=r, column=2).value,
                pr.get("amt_disp") if pr.get("amt_disp") is not None else "",
                f"amt mismatch for {name}",
            )
            r += 1
            ok = pr.get("open_key")
            if not ok:
                continue
            block = pack_details.get(str(ok)) or {}
            for ln in block.get("lines") or []:
                if not isinstance(ln, dict):
                    continue
                dcell = ws.cell(row=r, column=1)
                self.assertEqual(dcell.value, ln.get("name") or "", f"detail name @row {r}")
                self.assertFalse(
                    bool(dcell.font and dcell.font.bold),
                    f"detail row {r} should NOT be bold",
                )
                # 缩进
                al = dcell.alignment
                indent = int(getattr(al, "indent", 0) or 0) if al else 0
                expect_indent = 2 if ln.get("sub") else 1
                self.assertEqual(indent, expect_indent, f"indent @row {r} for {dcell.value}")
                self.assertEqual(
                    ws.cell(row=r, column=2).value,
                    ln.get("amt_disp") if ln.get("amt_disp") is not None else "",
                )
                r += 1

    def test_sampled_amounts_match_pack_disp(self):
        """至少抽 3 个数字：页面/pack disp == 单元格。"""
        raw = build_pl_xlsx_bytes(
            self.summary, period_key=self.yk, is_bu=False, scope_label="整体"
        )
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb.active
        packed = pack_pl_by_period(self.summary, is_bu=False)[self.yk]
        samples: list[tuple[str, str]] = []
        for pr in packed.get("rows") or []:
            if pr.get("amt_disp") and not pr.get("is_pct"):
                samples.append((pr["name"], pr["amt_disp"]))
            if len(samples) >= 3:
                break
        self.assertGreaterEqual(len(samples), 3)
        # 建 name→amt 映射（含明细）
        cell_map: dict[str, str] = {}
        for row in ws.iter_rows(min_row=1, max_col=2):
            n, a = row[0].value, row[1].value
            if n and a is not None and n not in ("科目", "产品", "VERSION", "范围", "周期", "导出时间", "口径"):
                cell_map[str(n)] = str(a)
        for name, amt in samples:
            self.assertIn(name, cell_map, f"missing {name}")
            self.assertEqual(cell_map[name], amt, f"{name}: cell={cell_map[name]} pack={amt}")

    def test_period_sensitivity(self):
        pkeys = [k for k in self.periods if k != self.yk]
        self.assertGreaterEqual(len(pkeys), 1, "need ≥2 periods in golden")
        p2 = pkeys[0]
        raw1 = build_pl_xlsx_bytes(
            self.summary, period_key=self.yk, is_bu=False, scope_label="整体"
        )
        raw2 = build_pl_xlsx_bytes(
            self.summary, period_key=p2, is_bu=False, scope_label="整体"
        )
        wb1 = openpyxl.load_workbook(io.BytesIO(raw1))
        wb2 = openpyxl.load_workbook(io.BytesIO(raw2))
        self.assertEqual(len(wb1.worksheets), 1)
        self.assertEqual(len(wb2.worksheets), 1)
        ws1, ws2 = wb1.active, wb2.active
        hdr1, hdr2 = _data_start_row(ws1), _data_start_row(ws2)
        amts1 = [ws1.cell(row=r, column=2).value for r in range(hdr1 + 1, ws1.max_row + 1)]
        amts2 = [ws2.cell(row=r, column=2).value for r in range(hdr2 + 1, ws2.max_row + 1)]
        names1 = {ws1.cell(row=r, column=1).value for r in range(hdr1 + 1, ws1.max_row + 1)}
        names2 = {ws2.cell(row=r, column=1).value for r in range(hdr2 + 1, ws2.max_row + 1)}
        self.assertTrue(
            amts1 != amts2 or names1 != names2,
            f"expected period diff: {self.yk} vs {p2}",
        )

    def test_unknown_period_raises(self):
        with self.assertRaises(KeyError):
            build_pl_xlsx_bytes(
                self.summary,
                period_key="1999年不存在",
                is_bu=False,
                scope_label="整体",
            )

    def test_filename_pattern(self):
        fn = pl_xlsx_filename(scope_label="游戏", period_key="2026年Q1", day="20260723")
        self.assertEqual(fn, "管理利润表_游戏_2026年Q1_20260723.xlsx")
        fn2 = pl_xlsx_filename(scope_label="整体", period_key="2026年", day="20260723")
        self.assertIn("管理利润表_整体_", fn2)


class TestPlXlsxHttp(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.summary = _load_golden_summary()
        cls.yk = cls.summary["meta"]["year_key"]
        cls.tmp = Path(tempfile.mkdtemp())
        (cls.tmp / "数据").mkdir(exist_ok=True)
        cls.cfg = dict(loaders.load_config(ROOT))
        cls.cfg["data_dir"] = "数据"
        cls.cfg["db_path"] = "数据/看板.db"
        cls.cfg["zhiyun_auto_fetch"] = False
        # BU 配置
        bus = [
            {"name": "甲BU", "销售": ["销售A"]},
            {"name": "乙BU", "销售": ["销售B"]},
        ]
        p = bu.config_path(cls.cfg, cls.tmp)
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(json.dumps({"bus": bus}, ensure_ascii=False), encoding="utf-8")
        accounts.save_accounts(
            cls.cfg,
            cls.tmp,
            [
                {
                    "账号": "lushasha",
                    "显示名": "管理员甲",
                    "权限": "管理员",
                    "密码": server.DEFAULT_PW,
                },
                {
                    "账号": "overall",
                    "显示名": "整体甲",
                    "权限": "整体",
                    "密码": server.DEFAULT_VIEW_PW,
                },
                {
                    "账号": "user_a",
                    "显示名": "甲负责人",
                    "权限": "BU",
                    "可见BU": ["甲BU"],
                    "密码": server.DEFAULT_VIEW_PW,
                },
                {
                    "账号": "user_b",
                    "显示名": "乙负责人",
                    "权限": "BU",
                    "可见BU": ["乙BU"],
                    "密码": server.DEFAULT_VIEW_PW,
                },
            ],
        )
        # 最小 DB
        conn = db.connect(cls.cfg, cls.tmp)
        conn.commit()
        conn.close()

        server._state["summary"] = cls.summary
        server._state["has_data"] = True
        server._state["user_html"] = "ready"
        # BU 页：summary 与整体同源（golden 无分 BU 时足够验权限+xlsx 路径）
        server._state["bu_pages"] = {
            "甲BU": {"summary": cls.summary, "name": "甲BU"},
            "乙BU": {"summary": cls.summary, "name": "乙BU"},
        }
        cls.app = server.create_app(cls.cfg, root=cls.tmp)
        cls._prev_offline = os.environ.get("KANBAN_OFFLINE")
        os.environ["KANBAN_OFFLINE"] = "1"

    @classmethod
    def tearDownClass(cls):
        if cls._prev_offline is None:
            os.environ.pop("KANBAN_OFFLINE", None)
        else:
            os.environ["KANBAN_OFFLINE"] = cls._prev_offline
        shutil.rmtree(cls.tmp, ignore_errors=True)

    def _client(self):
        from fastapi.testclient import TestClient

        return TestClient(self.app, follow_redirects=False)

    def _login(self, account: str):
        c = self._client()
        r = c.post(
            "/login",
            data={"account": account, "password": server.DEFAULT_VIEW_PW},
            follow_redirects=False,
        )
        self.assertIn(r.status_code, (200, 302, 303), r.text[:300])
        return c

    def test_overall_export_200_and_content_type(self):
        c = self._login("overall")
        r = c.get("/api/export/pl.xlsx", params={"blk": self.yk})
        self.assertEqual(r.status_code, 200, r.text[:400])
        ct = r.headers.get("content-type", "")
        self.assertIn("spreadsheet", ct)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        self.assertEqual(len(wb.worksheets), 1, wb.sheetnames)
        self.assertIn("管理利润表", wb.sheetnames)
        self.assertFalse(any(n.startswith("构成_") for n in wb.sheetnames))
        # dual mount
        r2 = c.get("/export/pl.xlsx", params={"blk": self.yk})
        self.assertEqual(r2.status_code, 200)

    def test_unauthenticated_401(self):
        c = self._client()
        r = c.get("/api/export/pl.xlsx", params={"blk": self.yk})
        self.assertEqual(r.status_code, 401)

    def test_unknown_blk_400(self):
        c = self._login("overall")
        r = c.get("/api/export/pl.xlsx", params={"blk": "1999年不存在"})
        self.assertEqual(r.status_code, 400)

    def test_bu_isolation_403_and_own_200(self):
        c = self._login("user_a")
        # 他人 BU
        r = c.get(f"/bu/{quote('乙BU')}/export/pl.xlsx", params={"blk": self.yk})
        self.assertEqual(r.status_code, 403, r.text[:300])
        # 自己的 BU
        r2 = c.get(f"/bu/{quote('甲BU')}/export/pl.xlsx", params={"blk": self.yk})
        self.assertEqual(r2.status_code, 200, r2.text[:400])
        self.assertIn("spreadsheet", r2.headers.get("content-type", ""))
        wb = openpyxl.load_workbook(io.BytesIO(r2.content))
        self.assertEqual(len(wb.worksheets), 1, wb.sheetnames)
        self.assertIn("管理利润表", wb.sheetnames)
        # 抬头块含范围
        meta = " ".join(
            str(ws_cell.value or "")
            for row in wb.active.iter_rows(min_row=1, max_row=8, max_col=2)
            for ws_cell in row
        )
        self.assertIn("甲BU", meta)

    def test_frontend_pl_export_button_source(self):
        """静态守卫：按钮在 PLTable，非顶栏；snapshot 隐藏；URL 跟 period。"""
        pl = (ROOT / "frontend/src/components/PLTable.vue").read_text(encoding="utf-8")
        top = (ROOT / "frontend/src/components/TopBarActions.vue").read_text(encoding="utf-8")
        self.assertIn("导出 Excel", pl)
        self.assertIn("export/pl.xlsx", pl)
        self.assertIn("store.period", pl)
        self.assertIn("snapshotMode", pl)
        self.assertIn("pl-export-excel", pl)
        # 顶栏仍是 HTML 快照，不得被换成 pl.xlsx 唯一导出
        self.assertIn("export.html", top)
        self.assertNotIn("export/pl.xlsx", top)


if __name__ == "__main__":
    unittest.main()
